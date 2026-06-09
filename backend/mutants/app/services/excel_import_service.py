"""Импорт объектов проекта из Excel."""

from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from dataclasses import dataclass
from typing import Any, Literal
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import use_fast_commit_for_current_transaction
from app.core.dependencies import CurrentPrincipal
from app.models.electrical_calculation import ElectricalCalculation
from app.models.project_object import ProjectObject
from app.models.specification import Specification
from app.reference_data.loader import (
    INSULATION_MATERIAL_RESELECTION_MESSAGE,
    list_insulation_materials,
)
from app.services.project_object_params import normalize_project_object_params
from app.services.project_service import (
    ProjectAccessError,
    ProjectNotFoundError,
    ProjectService,
)
from app.services.spreadsheet_safety import append_safe_row

PIPE_SHEET_NAMES = {"трубопроводы", "трубы", "pipes"}
TANK_SHEET_NAMES = {"резервуары", "ёмкости", "емкости", "tanks"}
IMPORT_COMMIT_BATCH_SIZE = 25
ImportMode = Literal["append", "merge", "replace"]

# Алиасы для колонки «Тип» в CSV (различает трубу/резервуар в одном файле)
TYPE_ALIASES: dict[str, str] = {
    "труба": "pipe",
    "трубопровод": "pipe",
    "pipe": "pipe",
    "резервуар": "tank",
    "ёмкость": "tank",
    "емкость": "tank",
    "бак": "tank",
    "tank": "tank",
}

GENERIC_MATERIAL_ALIASES: dict[str, str] = {
    "минеральная вата": "mineral_wool",
    "мин вата": "mineral_wool",
    "мин. вата": "mineral_wool",
    "минвата": "mineral_wool",
    "mineral_wool": "mineral_wool",
    "пеностекло": "foam_glass",
    "foam_glass": "foam_glass",
    "пенополиуретан": "polyurethane",
    "ппу": "polyurethane",
    "polyurethane": "polyurethane",
    "пенополистирол": "polystyrene",
    "полистирол": "polystyrene",
    "polystyrene": "polystyrene",
    "аэрогель": "aerogel",
    "aerogel": "aerogel",
    "силикат кальция": "calcium_silicate",
    "calcium_silicate": "calcium_silicate",
}

SPECIAL_MATERIAL_ALIASES: dict[str, str] = {
    "другое": "other",
    "other": "other",
}

PIPE_MATERIAL_ALIASES: dict[str, str] = {
    "углеродистая сталь": "carbon_steel",
    "сталь углеродистая": "carbon_steel",
    "carbon_steel": "carbon_steel",
    "нержавеющая сталь": "stainless_304",
    "stainless_304": "stainless_304",
    "медь": "copper",
    "copper": "copper",
    "алюминий": "aluminum",
    "aluminum": "aluminum",
    "пластик": "plastic",
    "plastic": "plastic",
}

PLACEMENT_ALIASES: dict[str, str] = {
    "на открытом воздухе": "outdoor",
    "улица": "outdoor",
    "надземная": "outdoor",
    "outdoor": "outdoor",
    "в помещении": "indoor",
    "помещение": "indoor",
    "indoor": "indoor",
    "подземно": "underground",
    "подземная": "underground",
    "underground": "underground",
}

GROUND_ALIASES: dict[str, str] = {
    "сухой песок": "dry_sand",
    "dry_sand": "dry_sand",
    "влажный песок": "wet_sand",
    "wet_sand": "wet_sand",
    "глина": "clay",
    "clay": "clay",
    "другое": "custom",
    "custom": "custom",
}

CLIMATE_BASIS_ALIASES: dict[str, str] = {
    "0,92": "t_0_92",
    "0.92": "t_0_92",
    "t_0_92": "t_0_92",
    "t0.92": "t_0_92",
    "t 0.92": "t_0_92",
    "0,98": "t_0_98",
    "0.98": "t_0_98",
    "t_0_98": "t_0_98",
    "t0.98": "t_0_98",
    "t 0.98": "t_0_98",
    "абс. мин.": "t_abs_min",
    "абс мин": "t_abs_min",
    "абсолютный минимум": "t_abs_min",
    "t_abs_min": "t_abs_min",
}

INSULATION_TEMPERATURE_BASIS_ALIASES: dict[str, str] = {
    "indoor": "indoor",
    "помещение": "indoor",
    "в помещении": "indoor",
    "outdoor_summer": "outdoor_summer",
    "улица лето": "outdoor_summer",
    "открытый воздух лето": "outdoor_summer",
    "открытый воздух, лето": "outdoor_summer",
    "outdoor_winter": "outdoor_winter",
    "улица зима": "outdoor_winter",
    "открытый воздух зима": "outdoor_winter",
    "открытый воздух, зима": "outdoor_winter",
    "channel": "channel",
    "канал": "channel",
    "tunnel": "tunnel",
    "тоннель": "tunnel",
    "technical_subfloor": "technical_subfloor",
    "техническое подполье": "technical_subfloor",
    "подполье": "technical_subfloor",
    "attic": "attic",
    "чердак": "attic",
    "basement": "basement",
    "подвал": "basement",
}

SHAPE_ALIASES: dict[str, str] = {
    "цилиндр": "cylindrical",
    "цилиндрический": "cylindrical",
    "cylindrical": "cylindrical",
    "параллелепипед": "rectangular",
    "прямоугольный": "rectangular",
    "прямоуг": "rectangular",
    "rectangular": "rectangular",
    "шар": "spherical",
    "сфера": "spherical",
    "сферический": "spherical",
    "spherical": "spherical",
}


# Заголовок колонки типа в CSV
TYPE_HEADERS: set[str] = {"тип", "type"}


# Возможные заголовки (нормализованные) → каноническое имя поля
PIPE_HEADERS: dict[str, str] = {
    "наименование": "name",
    "название": "name",
    "имя": "name",
    "диаметр, мм": "outer_diameter_mm",
    "диаметр мм": "outer_diameter_mm",
    "диаметр": "outer_diameter_mm",
    "ø, мм": "outer_diameter_mm",
    "ø мм": "outer_diameter_mm",
    "длина, м": "pipe_length",
    "длина м": "pipe_length",
    "длина": "pipe_length",
    "l, м": "pipe_length",
    "толщина изоляции, мм": "insulation_thickness_mm",
    "толщина изоляции мм": "insulation_thickness_mm",
    "толщина изоляции": "insulation_thickness_mm",
    "δ, мм": "insulation_thickness_mm",
    "δ мм": "insulation_thickness_mm",
    "материал изоляции": "insulation_material",
    "код материала изоляции": "insulation_material",
    "материал": "insulation_material",
    "т° среды": "ambient_temperature",
    "т среды": "ambient_temperature",
    "t° среды": "ambient_temperature",
    "t среды": "ambient_temperature",
    "температура среды": "ambient_temperature",
    "т° продукта": "process_temperature",
    "т продукта": "process_temperature",
    "t° продукта": "process_temperature",
    "t продукта": "process_temperature",
    "температура продукта": "process_temperature",
    "требуемая температура": "process_temperature",
    "требуемая температура трубы": "process_temperature",
    "температура поддержания": "process_temperature",
    "т° поддержания": "process_temperature",
    "t° поддержания": "process_temperature",
    "t проп": "vapor_temperature",
    "t проп.": "vapor_temperature",
    "t° проп": "vapor_temperature",
    "t° пропарки": "vapor_temperature",
    "t проп., °c": "vapor_temperature",
    "температура пропарки": "vapor_temperature",
    "t3": "maintain_temperature",
    "t3, °c": "maintain_temperature",
    "t3 поддержания": "maintain_temperature",
    "температура поддержания t3": "maintain_temperature",
    "макс. t° окр. среды": "max_ambient_temperature",
    "макс t° окр. среды": "max_ambient_temperature",
    "макс. допуст. t° продукта": "max_process_temperature",
    "макс допуст t° продукта": "max_process_temperature",
    "толщина стенки, мм": "wall_thickness_mm",
    "толщина стенки мм": "wall_thickness_mm",
    "толщина стенки": "wall_thickness_mm",
    "материал трубы": "pipe_material",
    "λ трубы": "pipe_lambda",
    "лямбда трубы": "pipe_lambda",
    "количество слоёв изоляции": "insulation_layer_count",
    "кол-во слоёв из": "insulation_layer_count",
    "слоёв из": "insulation_layer_count",
    "материал 2-го слоя": "second_insulation_material",
    "толщина 2-го слоя, мм": "second_insulation_thickness_mm",
    "толщина 2-го слоя": "second_insulation_thickness_mm",
    "λ 1-го слоя": "first_insulation_lambda",
    "λ 2-го слоя": "second_insulation_lambda",
    "материал 3-го слоя": "third_insulation_material",
    "толщина 3-го слоя, мм": "third_insulation_thickness_mm",
    "толщина 3-го слоя": "third_insulation_thickness_mm",
    "λ 3-го слоя": "third_insulation_lambda",
    "материал покрытия": "insulation_cover_material",
    "размещение": "placement",
    "размещение трубопровода": "placement",
    "глубина прокладки": "burial_depth",
    "грунт": "ground_type",
    "λ грунта": "ground_conductivity",
    "kзап": "safety_factor",
    "k зап": "safety_factor",
    "коэффициент запаса": "safety_factor",
    "рабочее напряжение": "supply_voltage",
    "мин. t включения, °c": "min_switch_temperature",
    "мин t включения": "min_switch_temperature",
    "минимальная температура включения": "min_switch_temperature",
    "климатический регион": "climate_region",
    "регион климата": "climate_region",
    "климатический город": "climate_city",
    "город климата": "climate_city",
    "климат": "climate_city",
    "ключ климата": "climate_key",
    "climate_key": "climate_key",
    "обеспеченность климата": "climate_temperature_basis",
    "температура климата": "climate_temperature_basis",
    "режим температуры изоляции": "insulation_temperature_basis",
    "tm изоляции": "insulation_temperature_basis",
    "tм изоляции": "insulation_temperature_basis",
    "задвижки": "valve_count",
    "фланцы": "flange_count",
    "опоры": "support_count",
    "l экв., м": "local_element_equiv_length",
    "l экв. м": "local_element_equiv_length",
    "эквивалентная длина локального элемента": "local_element_equiv_length",
    "эквивалентная длина, м": "local_element_equiv_length",
}

TANK_HEADERS: dict[str, str] = {
    "наименование": "name",
    "название": "name",
    "форма": "shape",
    "диаметр, мм": "diameter_mm",
    "диаметр мм": "diameter_mm",
    "диаметр": "diameter_mm",
    "длина, мм": "length_mm",
    "длина мм": "length_mm",
    "длина": "length_mm",
    "ширина, мм": "width_mm",
    "ширина мм": "width_mm",
    "ширина": "width_mm",
    "высота, мм": "height_mm",
    "высота мм": "height_mm",
    "высота": "height_mm",
    "толщина изоляции, мм": "insulation_thickness_mm",
    "толщина изоляции мм": "insulation_thickness_mm",
    "толщина изоляции": "insulation_thickness_mm",
    "δ, мм": "insulation_thickness_mm",
    "δ мм": "insulation_thickness_mm",
    "материал изоляции": "insulation_material",
    "код материала изоляции": "insulation_material",
    "материал": "insulation_material",
    "т° среды": "ambient_temperature",
    "т среды": "ambient_temperature",
    "t° среды": "ambient_temperature",
    "t среды": "ambient_temperature",
    "температура среды": "ambient_temperature",
    "т° продукта": "process_temperature",
    "т продукта": "process_temperature",
    "t° продукта": "process_temperature",
    "t продукта": "process_temperature",
    "температура продукта": "process_temperature",
    "требуемая температура": "process_temperature",
    "требуемая температура трубы": "process_temperature",
    "температура поддержания": "process_temperature",
    "т° поддержания": "process_temperature",
    "t° поддержания": "process_temperature",
    "t проп": "vapor_temperature",
    "t проп.": "vapor_temperature",
    "t° проп": "vapor_temperature",
    "t° пропарки": "vapor_temperature",
    "t проп., °c": "vapor_temperature",
    "температура пропарки": "vapor_temperature",
    "t3": "maintain_temperature",
    "t3, °c": "maintain_temperature",
    "t3 поддержания": "maintain_temperature",
    "температура поддержания t3": "maintain_temperature",
    "макс. t° окр. среды": "max_ambient_temperature",
    "макс t° окр. среды": "max_ambient_temperature",
    "макс. допуст. t° продукта": "max_process_temperature",
    "макс допуст t° продукта": "max_process_temperature",
    "количество слоёв изоляции": "insulation_layer_count",
    "кол-во слоёв из": "insulation_layer_count",
    "слоёв из": "insulation_layer_count",
    "материал 2-го слоя": "second_insulation_material",
    "толщина 2-го слоя, мм": "second_insulation_thickness_mm",
    "толщина 2-го слоя": "second_insulation_thickness_mm",
    "λ 1-го слоя": "first_insulation_lambda",
    "λ 2-го слоя": "second_insulation_lambda",
    "материал 3-го слоя": "third_insulation_material",
    "толщина 3-го слоя, мм": "third_insulation_thickness_mm",
    "толщина 3-го слоя": "third_insulation_thickness_mm",
    "λ 3-го слоя": "third_insulation_lambda",
    "материал покрытия": "insulation_cover_material",
    "размещение": "placement",
    "размещение резервуара": "placement",
    "глубина прокладки": "burial_depth",
    "грунт": "ground_type",
    "λ грунта": "ground_conductivity",
    "kзап": "safety_factor",
    "k зап": "safety_factor",
    "коэффициент запаса": "safety_factor",
    "рабочее напряжение": "supply_voltage",
    "мин. t включения, °c": "min_switch_temperature",
    "мин t включения": "min_switch_temperature",
    "минимальная температура включения": "min_switch_temperature",
    "климатический регион": "climate_region",
    "регион климата": "climate_region",
    "климатический город": "climate_city",
    "город климата": "climate_city",
    "климат": "climate_city",
    "ключ климата": "climate_key",
    "climate_key": "climate_key",
    "обеспеченность климата": "climate_temperature_basis",
    "температура климата": "climate_temperature_basis",
    "режим температуры изоляции": "insulation_temperature_basis",
    "tm изоляции": "insulation_temperature_basis",
    "tм изоляции": "insulation_temperature_basis",
    "q доп., вт": "q_additional",
    "q доп, вт": "q_additional",
    "дополнительные теплопотери, вт": "q_additional",
}


class ExcelImportError(Exception):
    """Ошибка импорта Excel/CSV."""


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > settings.MAX_XLSX_FILES:
                raise ExcelImportError(
                    f"XLSX содержит слишком много внутренних файлов: {len(infos)}"
                )
            total_uncompressed = sum(info.file_size for info in infos)
            if total_uncompressed > settings.MAX_XLSX_UNCOMPRESSED_BYTES:
                raise ExcelImportError(
                    "XLSX слишком большой после распаковки " f"({total_uncompressed // 1024} КБ)"
                )
            for info in infos:
                if info.compress_size and info.file_size / info.compress_size > 100:
                    raise ExcelImportError(
                        "XLSX похож на zip-bomb: слишком высокий коэффициент сжатия"
                    )
    except zipfile.BadZipFile as exc:
        raise ExcelImportError("Файл не является корректным XLSX-архивом") from exc


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, int | float):
        return float(v)
    s = str(v).replace(",", ".").strip()
    match = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if match:
        s = match.group(0)
    try:
        return float(s)
    except ValueError:
        return None


@dataclass(frozen=True)
class _MaterialResolution:
    material: str | None = None
    raw: str | None = None
    family: str | None = None
    needs_reselection: bool = False


def _selectable_insulation_entries() -> list[dict[str, Any]]:
    return [
        entry
        for entry in list_insulation_materials()
        if entry.get("selectable") is not False and entry.get("deprecated") is not True
    ]


def _resolve_material_entry(v: Any) -> _MaterialResolution:
    key = _norm(v)
    if not key:
        return _MaterialResolution()
    raw = str(v).strip()
    if key in SPECIAL_MATERIAL_ALIASES:
        return _MaterialResolution(material=SPECIAL_MATERIAL_ALIASES[key], raw=raw)
    if key in GENERIC_MATERIAL_ALIASES:
        return _MaterialResolution(
            raw=raw,
            family=GENERIC_MATERIAL_ALIASES[key],
            needs_reselection=True,
        )

    entries = _selectable_insulation_entries()
    by_code = {str(entry["material"]).lower(): entry for entry in entries}
    code_match = by_code.get(key)
    if code_match is not None:
        return _MaterialResolution(material=str(code_match["material"]), raw=raw)

    name_matches = [entry for entry in entries if _norm(entry.get("name")) == key]
    if len(name_matches) == 1:
        return _MaterialResolution(material=str(name_matches[0]["material"]), raw=raw)
    if len(name_matches) > 1:
        family = str(name_matches[0].get("material_family") or name_matches[0]["material"])
        return _MaterialResolution(raw=raw, family=family, needs_reselection=True)
    return _MaterialResolution()


def _resolve_material(v: Any) -> str | None:
    return _resolve_material_entry(v).material


def _apply_insulation_material_resolution(
    params: dict[str, Any],
    field: str,
    resolution: _MaterialResolution,
) -> None:
    if resolution.material:
        params[field] = resolution.material
        return
    if not resolution.needs_reselection:
        return
    params["needs_material_reselection"] = True
    params[f"{field}_raw"] = resolution.raw
    params[f"{field}_family"] = resolution.family
    params[f"{field}_warning"] = INSULATION_MATERIAL_RESELECTION_MESSAGE


def _resolve_pipe_material(v: Any) -> str | None:
    key = _norm(v)
    if not key:
        return None
    return PIPE_MATERIAL_ALIASES.get(key)


def _resolve_alias(v: Any, aliases: dict[str, str]) -> str | None:
    key = _norm(v)
    if not key:
        return None
    return aliases.get(key)


def _resolve_climate_basis(v: Any) -> str | None:
    key = _norm(v)
    if not key:
        return None
    return CLIMATE_BASIS_ALIASES.get(key)


def _resolve_shape(v: Any) -> str | None:
    key = _norm(v)
    if not key:
        return None
    # Первое слово тоже пробуем (если запись «цилиндрический бак»)
    return SHAPE_ALIASES.get(key) or SHAPE_ALIASES.get(key.split()[0] if key else "")


def _apply_common_srs_params(params: dict[str, Any], row: dict[str, Any]) -> None:
    """Добавляет необязательные SRS-поля объекта из Excel/CSV в params."""
    placement = _resolve_alias(row.get("placement"), PLACEMENT_ALIASES)
    if placement:
        params["placement"] = placement
        params["location"] = "indoor" if placement == "indoor" else "outdoor"
    burial_depth = _to_float(row.get("burial_depth"))
    if burial_depth is not None:
        params["burial_depth"] = burial_depth
    ground_type = _resolve_alias(row.get("ground_type"), GROUND_ALIASES)
    if ground_type:
        params["ground_type"] = ground_type
    ground_conductivity = _to_float(row.get("ground_conductivity"))
    if ground_conductivity is not None:
        params["ground_conductivity"] = ground_conductivity
    safety_factor = _to_float(row.get("safety_factor"))
    if safety_factor is not None:
        params["safety_factor"] = safety_factor
    min_switch_temperature = _to_float(row.get("min_switch_temperature"))
    if min_switch_temperature is not None:
        params["min_switch_temperature"] = min_switch_temperature
    supply_voltage = _to_float(row.get("supply_voltage"))
    if supply_voltage is not None:
        params["supply_voltage"] = supply_voltage
    vapor_temperature = _to_float(row.get("vapor_temperature"))
    if vapor_temperature is not None:
        params["vapor_temperature"] = vapor_temperature
    maintain_temperature = _to_float(row.get("maintain_temperature"))
    if maintain_temperature is not None:
        params["maintain_temperature"] = maintain_temperature
    max_ambient = _to_float(row.get("max_ambient_temperature"))
    if max_ambient is not None:
        params["max_ambient_temperature"] = max_ambient
    max_process = _to_float(row.get("max_process_temperature"))
    if max_process is not None:
        params["max_process_temperature"] = max_process
    cover = row.get("insulation_cover_material")
    if cover and str(cover).strip():
        params["insulation_cover_material"] = str(cover).strip()
    climate_key = row.get("climate_key")
    if climate_key and str(climate_key).strip():
        key_value = str(climate_key).strip()
        params["climate_key"] = key_value
        if "|||" in key_value:
            region, city = key_value.split("|||", 1)
            params.setdefault("climate_region", region.strip())
            params.setdefault("climate_city", city.strip())
    climate_region = row.get("climate_region")
    if climate_region and str(climate_region).strip():
        params["climate_region"] = str(climate_region).strip()
    climate_city = row.get("climate_city")
    if climate_city and str(climate_city).strip():
        params["climate_city"] = str(climate_city).strip()
    climate_basis = _resolve_climate_basis(row.get("climate_temperature_basis"))
    if climate_basis:
        params["climate_temperature_basis"] = climate_basis
    insulation_temperature_basis = _resolve_alias(
        row.get("insulation_temperature_basis"),
        INSULATION_TEMPERATURE_BASIS_ALIASES,
    )
    if insulation_temperature_basis:
        params["insulation_temperature_basis"] = insulation_temperature_basis


def _apply_layered_insulation(params: dict[str, Any], row: dict[str, Any]) -> None:
    count = int(_to_float(row.get("insulation_layer_count")) or 1)
    count = min(max(count, 1), 3)
    params["insulation_layer_count"] = str(count)

    layers: list[dict[str, Any]] = []
    first_layer: dict[str, Any] = {}
    if params.get("insulation_thickness") is not None:
        first_layer["thickness"] = params["insulation_thickness"]
    if params.get("insulation_material") is not None:
        first_layer["material"] = params["insulation_material"]
    first_lambda = _to_float(row.get("first_insulation_lambda"))
    if first_lambda is not None:
        first_layer["conductivity"] = first_lambda
    if first_layer:
        layers.append(first_layer)

    if count >= 2:
        material2_resolution = _resolve_material_entry(row.get("second_insulation_material"))
        material2 = material2_resolution.material
        thickness2 = _to_float(row.get("second_insulation_thickness_mm"))
        lambda2 = _to_float(row.get("second_insulation_lambda"))
        if (
            material2
            or material2_resolution.needs_reselection
            or thickness2 is not None
            or lambda2 is not None
        ):
            while len(layers) < 1:
                layers.append({})
            layer2: dict[str, Any] = {}
            if thickness2 is not None:
                layer2["thickness"] = thickness2 / 1000.0
            if material2:
                layer2["material"] = material2
            elif material2_resolution.needs_reselection:
                params["needs_material_reselection"] = True
                layer2["material_raw"] = material2_resolution.raw
                layer2["material_family"] = material2_resolution.family
                layer2["material_warning"] = INSULATION_MATERIAL_RESELECTION_MESSAGE
            if lambda2 is not None:
                layer2["conductivity"] = lambda2
            layers.append(layer2)

    if count >= 3:
        material3_resolution = _resolve_material_entry(row.get("third_insulation_material"))
        material3 = material3_resolution.material
        thickness3 = _to_float(row.get("third_insulation_thickness_mm"))
        lambda3 = _to_float(row.get("third_insulation_lambda"))
        if (
            material3
            or material3_resolution.needs_reselection
            or thickness3 is not None
            or lambda3 is not None
        ):
            while len(layers) < 2:
                layers.append({})
            layer3: dict[str, Any] = {}
            if thickness3 is not None:
                layer3["thickness"] = thickness3 / 1000.0
            if material3:
                layer3["material"] = material3
            elif material3_resolution.needs_reselection:
                params["needs_material_reselection"] = True
                layer3["material_raw"] = material3_resolution.raw
                layer3["material_family"] = material3_resolution.family
                layer3["material_warning"] = INSULATION_MATERIAL_RESELECTION_MESSAGE
            if lambda3 is not None:
                layer3["conductivity"] = lambda3
            layers.append(layer3)

    if layers:
        params["insulation_layers"] = layers


def _read_sheet(ws: Any, header_map: dict[str, str]) -> list[dict[str, Any]]:
    """Читает лист, мапит заголовки → канонические имена, возвращает список строк-словарей.

    Строки возвращаются с ключом ``_row`` — номер строки в Excel (1-based).
    """
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    mapped_cols: list[tuple[int, str]] = []
    for idx, h in enumerate(header_row):
        key = header_map.get(_norm(h))
        if key:
            mapped_cols.append((idx, key))
    result: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        if len(result) >= settings.MAX_IMPORT_ROWS:
            raise ExcelImportError(f"Превышен лимит строк импорта: {settings.MAX_IMPORT_ROWS}")
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        item: dict[str, Any] = {"_row": row_idx}
        for idx, key in mapped_cols:
            if idx < len(row):
                item[key] = row[idx]
        result.append(item)
    return result


def _build_pipe_params(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    """Строит params-словарь для трубы (м, не мм). Возвращает (params, error)."""
    d_mm = _to_float(row.get("outer_diameter_mm"))
    L = _to_float(row.get("pipe_length"))
    ins_mm = _to_float(row.get("insulation_thickness_mm"))
    material_resolution = _resolve_material_entry(row.get("insulation_material"))
    t_a = _to_float(row.get("ambient_temperature"))
    t_p = _to_float(row.get("process_temperature"))

    params: dict[str, Any] = {}
    if d_mm is not None:
        params["outer_diameter"] = d_mm / 1000.0
    if L is not None:
        params["pipe_length"] = L
    if ins_mm is not None:
        params["insulation_thickness"] = ins_mm / 1000.0
    _apply_insulation_material_resolution(params, "insulation_material", material_resolution)
    if t_a is not None:
        params["ambient_temperature"] = t_a
    if t_p is not None:
        params["process_temperature"] = t_p
    wall_mm = _to_float(row.get("wall_thickness_mm"))
    if wall_mm is not None:
        params["wall_thickness"] = wall_mm / 1000.0
    pipe_material = _resolve_pipe_material(row.get("pipe_material"))
    if pipe_material:
        params["pipe_material"] = pipe_material
    pipe_lambda = _to_float(row.get("pipe_lambda"))
    if pipe_lambda is not None:
        params["pipe_lambda"] = pipe_lambda
    for count_field in ("valve_count", "flange_count", "support_count"):
        value = _to_float(row.get(count_field))
        if value is not None:
            params[count_field] = int(value)
    local_element_equiv_length = _to_float(row.get("local_element_equiv_length"))
    if local_element_equiv_length is not None:
        params["local_element_equiv_length"] = local_element_equiv_length
    local_count = sum(
        int(params.get(k, 0) or 0) for k in ("valve_count", "flange_count", "support_count")
    )
    if local_count:
        params["num_local_elements"] = local_count
    _apply_common_srs_params(params, row)
    _apply_layered_insulation(params, row)
    name = row.get("name")
    if name and str(name).strip():
        params["name"] = str(name).strip()
    return params, None


def _build_tank_params(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    shape = _resolve_shape(row.get("shape"))
    if not shape and row.get("shape") not in (None, ""):
        return None, "Не указана или не распознана форма (цилиндр / параллелепипед / шар)"

    ins_mm = _to_float(row.get("insulation_thickness_mm"))
    material_resolution = _resolve_material_entry(row.get("insulation_material"))
    t_a = _to_float(row.get("ambient_temperature"))
    t_p = _to_float(row.get("process_temperature"))

    params: dict[str, Any] = {}
    if shape:
        params["shape"] = shape
    if ins_mm is not None:
        params["insulation_thickness"] = ins_mm / 1000.0
    _apply_insulation_material_resolution(params, "insulation_material", material_resolution)
    if t_a is not None:
        params["ambient_temperature"] = t_a
    if t_p is not None:
        params["process_temperature"] = t_p

    if shape == "cylindrical":
        d_mm = _to_float(row.get("diameter_mm"))
        h_mm = _to_float(row.get("height_mm"))
        if d_mm is not None:
            params["diameter"] = d_mm / 1000.0
        if h_mm is not None:
            params["height"] = h_mm / 1000.0
    elif shape == "rectangular":
        L_mm = _to_float(row.get("length_mm"))
        W_mm = _to_float(row.get("width_mm"))
        H_mm = _to_float(row.get("height_mm"))
        if L_mm is not None:
            params["length"] = L_mm / 1000.0
        if W_mm is not None:
            params["width"] = W_mm / 1000.0
        if H_mm is not None:
            params["height"] = H_mm / 1000.0
    elif shape == "spherical":
        d_mm = _to_float(row.get("diameter_mm"))
        if d_mm is not None:
            params["diameter"] = d_mm / 1000.0

    _apply_common_srs_params(params, row)
    _apply_layered_insulation(params, row)
    q_additional = _to_float(row.get("q_additional"))
    if q_additional is not None:
        params["q_additional"] = q_additional

    name = row.get("name")
    if name and str(name).strip():
        params["name"] = str(name).strip()
    return params, None


def _parse_csv(content: bytes) -> list[tuple[str, list[dict[str, Any]]]]:
    """Парсит CSV-файл.

    Возвращает список пар (sheet_label, rows) по типу:
    [('Трубопроводы', [...]), ('Резервуары', [...])].
    CSV должен содержать колонку «Тип» со значениями «труба» / «резервуар».
    Автодетект разделителя (``,``, ``;``, ``\t``). Кодировки: UTF-8 / UTF-8-BOM / CP1251.
    """
    import csv

    # Определяем кодировку
    text: str | None = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ExcelImportError("Не удалось определить кодировку CSV (ожидается UTF-8 или CP1251)")

    # Автодетект разделителя по первой строке
    first_line = text.splitlines()[0] if text.strip() else ""
    if not first_line:
        raise ExcelImportError("CSV-файл пустой")
    counts = {d: first_line.count(d) for d in (";", ",", "\t")}
    delimiter = max(counts, key=counts.get) if any(counts.values()) else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        raise ExcelImportError("CSV-файл пустой")
    if len(all_rows) - 1 > settings.MAX_IMPORT_ROWS:
        raise ExcelImportError(f"Превышен лимит строк импорта: {settings.MAX_IMPORT_ROWS}")

    header = all_rows[0]
    # Индекс колонки «Тип»
    type_idx = next((i for i, h in enumerate(header) if _norm(h) in TYPE_HEADERS), None)
    if type_idx is None:
        raise ExcelImportError(
            "В CSV не найдена колонка «Тип» (со значениями «труба»/«резервуар»). "
            "Используйте шаблон CSV или файл Excel."
        )

    # Разделяем строки по типу
    pipe_rows_raw: list[tuple[int, list]] = []
    tank_rows_raw: list[tuple[int, list]] = []
    for row_idx, row in enumerate(all_rows[1:], start=2):
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        if type_idx >= len(row):
            continue
        t = TYPE_ALIASES.get(_norm(row[type_idx]))
        if t == "pipe":
            pipe_rows_raw.append((row_idx, row))
        elif t == "tank":
            tank_rows_raw.append((row_idx, row))

    def build_mapped(
        rows_raw: list[tuple[int, list]], header_map: dict[str, str]
    ) -> list[dict[str, Any]]:
        # Определяем маппинг колонок общего header
        mapped_cols: list[tuple[int, str]] = []
        for idx, h in enumerate(header):
            key = header_map.get(_norm(h))
            if key:
                mapped_cols.append((idx, key))
        out: list[dict[str, Any]] = []
        for row_idx, row in rows_raw:
            item: dict[str, Any] = {"_row": row_idx}
            for idx, key in mapped_cols:
                if idx < len(row):
                    val = row[idx]
                    # Пустые строки — None
                    item[key] = val if (val is not None and str(val).strip() != "") else None
            out.append(item)
        return out

    result: list[tuple[str, list[dict[str, Any]]]] = []
    if pipe_rows_raw:
        result.append(("Трубопроводы (CSV)", build_mapped(pipe_rows_raw, PIPE_HEADERS)))
    if tank_rows_raw:
        result.append(("Резервуары (CSV)", build_mapped(tank_rows_raw, TANK_HEADERS)))
    return result


async def _project_import_state(db: AsyncSession, project_id: UUID) -> tuple[int, int]:
    result = await db.execute(
        select(func.count(ProjectObject.id), func.max(ProjectObject.sort_order)).where(
            ProjectObject.project_id == project_id
        )
    )
    count, max_sort = result.one()
    return int(count or 0), int(max_sort if max_sort is not None else -1) + 1


def _normalize_name_for_dedupe(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _dedupe_key(object_type: str, params: dict[str, Any]) -> str:
    key_params = dict(params)
    name = _normalize_name_for_dedupe(key_params.pop("name", ""))
    payload = json.dumps(
        key_params,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    return f"{object_type}:{name}:{digest}"


def _object_type_for_dedupe(value: Any) -> str:
    return str(getattr(value, "value", value))


async def _existing_dedupe_keys(db: AsyncSession, project_id: UUID) -> set[str]:
    result = await db.execute(
        select(ProjectObject.object_type, ProjectObject.params).where(
            ProjectObject.project_id == project_id
        )
    )
    return {
        _dedupe_key(_object_type_for_dedupe(object_type), params or {})
        for object_type, params in result.all()
    }


async def _replace_project_objects(db: AsyncSession, project_id: UUID) -> None:
    await db.execute(
        delete(ElectricalCalculation).where(ElectricalCalculation.project_id == project_id)
    )
    await db.execute(delete(Specification).where(Specification.project_id == project_id))
    await db.execute(delete(ProjectObject).where(ProjectObject.project_id == project_id))
    await db.flush()


def _validate_import_mode(mode: str) -> ImportMode:
    normalized = (mode or "merge").strip().lower()
    if normalized not in {"append", "merge", "replace"}:
        raise ExcelImportError(
            "Некорректный режим импорта: " f"{mode!r} (допустимо: append, merge, replace)"
        )
    return normalized  # type: ignore[return-value]


async def _ensure_import_access(
    db: AsyncSession,
    project_id: UUID,
    principal: CurrentPrincipal,
) -> None:
    project_service = ProjectService(db)
    project = await project_service.get_project_basic(project_id, principal)
    project_service._check_owner(project, principal)


async def _commit_object_batch(
    db: AsyncSession,
    batch: list[tuple[ProjectObject, dict[str, Any]]],
    sheet_label: str,
) -> tuple[int, list[UUID], list[dict[str, Any]]]:
    if not batch:
        return 0, [], []
    objects = [item[0] for item in batch]
    try:
        db.add_all(objects)
        await db.flush()
        object_ids = [obj.id for obj in objects]
        await use_fast_commit_for_current_transaction(db)
        await db.commit()
        return len(objects), object_ids, []
    except SQLAlchemyError as exc:
        await db.rollback()
        return await _commit_object_batch_row_by_row(db, batch, sheet_label, exc)


async def _commit_object_batch_row_by_row(
    db: AsyncSession,
    batch: list[tuple[ProjectObject, dict[str, Any]]],
    sheet_label: str,
    batch_error: SQLAlchemyError,
) -> tuple[int, list[UUID], list[dict[str, Any]]]:
    created = 0
    object_ids: list[UUID] = []
    errors: list[dict[str, Any]] = []
    for obj, row in batch:
        retry_obj = ProjectObject(
            project_id=obj.project_id,
            object_type=obj.object_type,
            sort_order=obj.sort_order,
            params=obj.params,
        )
        try:
            db.add(retry_obj)
            await db.flush()
            object_ids.append(retry_obj.id)
            await use_fast_commit_for_current_transaction(db)
            await db.commit()
            created += 1
        except Exception as exc:
            await db.rollback()
            message = f"{type(exc).__name__}: {exc}"
            if not message.strip():
                message = f"{type(batch_error).__name__}: {batch_error}"
            errors.append({"sheet": sheet_label, "row": row["_row"], "message": message})
    return created, object_ids, errors


async def _add_rows(
    db: AsyncSession,
    project_id: UUID,
    sheet_label: str,
    rows: list[dict[str, Any]],
    object_type: str,
    next_sort: int,
    current_count: int,
    dedupe_keys: set[str] | None = None,
) -> tuple[int, int, int, list[dict[str, Any]], list[UUID], int, int]:
    """Создаёт объекты из распарсенных строк.

    Расчёт теплопотерь здесь намеренно не запускается: импорт должен быстро
    сохранить всё распознанное и отдать новые объекты в один фоновый batch.
    """
    created = 0
    skipped_duplicates = 0
    skipped_limit = 0
    created_object_ids: list[UUID] = []
    errors: list[dict[str, Any]] = []
    batch: list[tuple[ProjectObject, dict[str, Any]]] = []
    builder = _build_pipe_params if object_type == "pipe" else _build_tank_params

    async def flush_batch() -> None:
        nonlocal batch, created, current_count
        attempted = len(batch)
        batch_created, object_ids, batch_errors = await _commit_object_batch(
            db,
            batch,
            sheet_label,
        )
        created += batch_created
        created_object_ids.extend(object_ids)
        errors.extend(batch_errors)
        current_count -= attempted - batch_created
        batch = []

    for row_index, row in enumerate(rows):
        if current_count >= settings.GUEST_MAX_OBJECTS_PER_PROJECT:
            skipped_limit = len(rows) - row_index
            errors.append(
                {
                    "sheet": sheet_label,
                    "row": row["_row"],
                    "message": (
                        "Достигнут лимит объектов в проекте "
                        f"({settings.GUEST_MAX_OBJECTS_PER_PROJECT}). "
                        f"Пропущено строк: {skipped_limit}."
                    ),
                }
            )
            break
        params, err = builder(row)
        if err or params is None:
            errors.append(
                {"sheet": sheet_label, "row": row["_row"], "message": err or "Ошибка парсинга"}
            )
            continue
        try:
            normalized_params = normalize_project_object_params(object_type, params)
            if dedupe_keys is not None:
                key = _dedupe_key(object_type, normalized_params)
                if key in dedupe_keys:
                    skipped_duplicates += 1
                    continue
                dedupe_keys.add(key)
            obj = ProjectObject(
                project_id=project_id,
                object_type=object_type,
                sort_order=next_sort,
                params=normalized_params,
            )
            batch.append((obj, row))
            current_count += 1
            next_sort += 1
            if len(batch) >= IMPORT_COMMIT_BATCH_SIZE:
                await flush_batch()
        except Exception as exc:
            errors.append(
                {
                    "sheet": sheet_label,
                    "row": row["_row"],
                    "message": f"{type(exc).__name__}: {exc}",
                }
            )
    await flush_batch()
    return (
        created,
        next_sort,
        current_count,
        errors,
        created_object_ids,
        skipped_duplicates,
        skipped_limit,
    )


async def import_objects_from_csv(
    db: AsyncSession,
    project_id: UUID,
    principal: CurrentPrincipal,
    content: bytes,
    mode: str = "merge",
) -> dict[str, Any]:
    """Импортирует объекты из CSV-файла. Требуется колонка «Тип»."""
    import_mode = _validate_import_mode(mode)
    await _ensure_import_access(db, project_id, principal)

    sheets = _parse_csv(content)
    if not sheets:
        raise ExcelImportError(
            "В CSV не найдено ни одной строки с распознанным типом (труба/резервуар)."
        )

    if import_mode == "replace":
        await _replace_project_objects(db, project_id)
        current_count, next_sort = 0, 0
        dedupe_keys = None
    else:
        current_count, next_sort = await _project_import_state(db, project_id)
        dedupe_keys = (
            await _existing_dedupe_keys(db, project_id) if import_mode == "merge" else None
        )

    total_created = 0
    skipped_duplicates = 0
    skipped_limit = 0
    all_errors: list[dict[str, Any]] = []
    created_object_ids: list[UUID] = []
    for sheet_label, rows in sheets:
        obj_type = "pipe" if "Трубопровод" in sheet_label else "tank"
        (
            created,
            next_sort,
            current_count,
            errors,
            object_ids,
            skipped,
            limit_skipped,
        ) = await _add_rows(
            db,
            project_id,
            sheet_label,
            rows,
            obj_type,
            next_sort,
            current_count,
            dedupe_keys=dedupe_keys,
        )
        total_created += created
        skipped_duplicates += skipped
        skipped_limit += limit_skipped
        all_errors.extend(errors)
        created_object_ids.extend(object_ids)

    if import_mode == "replace" and not created_object_ids:
        await db.commit()

    return {
        "created": total_created,
        "skipped_duplicates": skipped_duplicates,
        "skipped_limit": skipped_limit,
        "mode": import_mode,
        "errors": all_errors,
        "created_object_ids": created_object_ids,
    }


async def import_objects_from_excel(
    db: AsyncSession,
    project_id: UUID,
    principal: CurrentPrincipal,
    content: bytes,
    mode: str = "merge",
) -> dict[str, Any]:
    """Импортирует объекты из xlsx-файла в проект.

    Возвращает сводку: {"created": N, "errors": [{"sheet", "row", "message"}]}.
    """
    import_mode = _validate_import_mode(mode)
    _validate_xlsx_archive(content)
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportError(f"Не удалось открыть файл: {exc}") from exc
    if len(wb.sheetnames) > settings.MAX_IMPORT_SHEETS:
        raise ExcelImportError(f"Превышен лимит листов импорта: {settings.MAX_IMPORT_SHEETS}")

    # Проверяем доступ к проекту
    try:
        await _ensure_import_access(db, project_id, principal)
    except (ProjectNotFoundError, ProjectAccessError):
        raise

    created = 0
    skipped_duplicates = 0
    skipped_limit = 0
    errors: list[dict[str, Any]] = []
    created_object_ids: list[UUID] = []
    found_sheet = False
    parsed_sheets: list[tuple[str, str, list[dict[str, Any]]]] = []

    for sheet in wb.sheetnames:
        norm = _norm(sheet)
        if norm in PIPE_SHEET_NAMES:
            found_sheet = True
            ws = wb[sheet]
            rows = _read_sheet(ws, PIPE_HEADERS)
            parsed_sheets.append((sheet, "pipe", rows))
        elif norm in TANK_SHEET_NAMES:
            found_sheet = True
            ws = wb[sheet]
            rows = _read_sheet(ws, TANK_HEADERS)
            parsed_sheets.append((sheet, "tank", rows))

    if not found_sheet:
        raise ExcelImportError(
            "В файле не найдены листы «Трубопроводы» или «Резервуары». "
            "Используйте шаблон (кнопка «Скачать шаблон»)."
        )

    if import_mode == "replace":
        await _replace_project_objects(db, project_id)
        current_count, next_sort = 0, 0
        dedupe_keys = None
    else:
        current_count, next_sort = await _project_import_state(db, project_id)
        dedupe_keys = (
            await _existing_dedupe_keys(db, project_id) if import_mode == "merge" else None
        )

    for sheet, object_type, rows in parsed_sheets:
        (
            added,
            next_sort,
            current_count,
            sheet_errors,
            object_ids,
            skipped,
            limit_skipped,
        ) = await _add_rows(
            db,
            project_id,
            sheet,
            rows,
            object_type,
            next_sort,
            current_count,
            dedupe_keys=dedupe_keys,
        )
        created += added
        skipped_duplicates += skipped
        skipped_limit += limit_skipped
        errors.extend(sheet_errors)
        created_object_ids.extend(object_ids)

    if import_mode == "replace" and not created_object_ids:
        await db.commit()

    return {
        "created": created,
        "skipped_duplicates": skipped_duplicates,
        "skipped_limit": skipped_limit,
        "mode": import_mode,
        "errors": errors,
        "created_object_ids": created_object_ids,
    }


def _material_label(material: Any) -> str:
    if not material:
        return ""
    material_str = str(material)
    for entry in list_insulation_materials():
        if entry.get("material") == material_str:
            return str(entry.get("name") or material_str)
    return material_str


def _material_status(params: dict[str, Any]) -> str:
    if params.get("needs_material_reselection") is True:
        return "Требует уточнения"
    if params.get("insulation_material"):
        return "Конкретный материал"
    return ""


def _material_warning(params: dict[str, Any]) -> str:
    if params.get("needs_material_reselection") is True:
        return INSULATION_MATERIAL_RESELECTION_MESSAGE
    return ""


SHAPE_LABELS_RU: dict[str, str] = {
    "cylindrical": "Цилиндр",
    "rectangular": "Параллелепипед",
    "spherical": "Шар",
}


def _to_export_mm(value: Any) -> float | str:
    if value in (None, ""):
        return ""
    try:
        result = round(float(value) * 1000, 3)
    except (TypeError, ValueError):
        return str(value)
    return result or ""


def build_objects_xlsx(objects: list[Any]) -> bytes:
    """Экспорт объектов проекта в формат, round-trip-совместимый с импортом.

    Пишет листы `Трубопроводы` и `Резервуары` с теми же колонками, что и
    `build_template_xlsx` — файл можно сохранить локально и через
    `POST /objects/import-excel` загрузить обратно.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_pipe = wb.active
    ws_pipe.title = "Трубопроводы"
    pipe_cols = [
        "Наименование",
        "Диаметр, мм",
        "Длина, м",
        "Толщина изоляции, мм",
        "Материал изоляции",
        "Код материала изоляции",
        "Статус материала изоляции",
        "Комментарий материала изоляции",
        "T° среды",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        "Режим температуры изоляции",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Задвижки",
        "Фланцы",
        "Опоры",
        "L экв., м",
    ]
    for c, h in enumerate(pipe_cols, start=1):
        cell = ws_pipe.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")

    ws_tank = wb.create_sheet("Резервуары")
    tank_cols = [
        "Наименование",
        "Форма",
        "Диаметр, мм",
        "Длина, мм",
        "Ширина, мм",
        "Высота, мм",
        "Толщина изоляции, мм",
        "Материал изоляции",
        "Код материала изоляции",
        "Статус материала изоляции",
        "Комментарий материала изоляции",
        "T° среды",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        "Режим температуры изоляции",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Q доп., Вт",
    ]
    for c, h in enumerate(tank_cols, start=1):
        cell = ws_tank.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")

    for obj in objects:
        params = obj.params or {}
        name = params.get("name") or ""
        material_code = params.get("insulation_material", "")
        material = _material_label(material_code) or str(
            params.get("insulation_material_raw") or ""
        )
        if obj.object_type == "pipe":
            append_safe_row(
                ws_pipe,
                [
                    name,
                    _to_export_mm(params.get("outer_diameter")),
                    params.get("pipe_length") or "",
                    _to_export_mm(params.get("insulation_thickness")),
                    material,
                    material_code,
                    _material_status(params),
                    _material_warning(params),
                    params.get("ambient_temperature", ""),
                    params.get("process_temperature", ""),
                    params.get("vapor_temperature", ""),
                    params.get("placement", ""),
                    params.get("insulation_temperature_basis", ""),
                    params.get("climate_region", ""),
                    params.get("climate_city", ""),
                    params.get("climate_key", ""),
                    params.get("climate_temperature_basis", ""),
                    params.get("safety_factor", ""),
                    params.get("min_switch_temperature", ""),
                    params.get("valve_count", ""),
                    params.get("flange_count", ""),
                    params.get("support_count", ""),
                    params.get("local_element_equiv_length", ""),
                ],
            )
        elif obj.object_type == "tank":
            shape_code = params.get("shape") or "cylindrical"
            shape = SHAPE_LABELS_RU.get(shape_code, shape_code)

            def to_mm(k, _params=params):
                return _to_export_mm(_params.get(k))

            append_safe_row(
                ws_tank,
                [
                    name,
                    shape,
                    to_mm("diameter"),
                    to_mm("length"),
                    to_mm("width"),
                    to_mm("height"),
                    to_mm("insulation_thickness"),
                    material,
                    material_code,
                    _material_status(params),
                    _material_warning(params),
                    params.get("ambient_temperature", ""),
                    params.get("process_temperature", ""),
                    params.get("vapor_temperature", ""),
                    params.get("placement", ""),
                    params.get("insulation_temperature_basis", ""),
                    params.get("climate_region", ""),
                    params.get("climate_city", ""),
                    params.get("climate_key", ""),
                    params.get("climate_temperature_basis", ""),
                    params.get("safety_factor", ""),
                    params.get("min_switch_temperature", ""),
                    params.get("q_additional", ""),
                ],
            )

    ws_pipe.column_dimensions["A"].width = 24
    for col_idx in range(2, len(pipe_cols) + 1):
        ws_pipe.column_dimensions[get_column_letter(col_idx)].width = 18
    ws_tank.column_dimensions["A"].width = 24
    for col_idx in range(2, len(tank_cols) + 1):
        ws_tank.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_xlsx() -> bytes:
    """Формирует xlsx-шаблон с двумя листами и примерами."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_pipe = wb.active
    ws_pipe.title = "Трубопроводы"
    pipe_cols = [
        "Наименование",
        "Диаметр, мм",
        "Длина, м",
        "Толщина изоляции, мм",
        "Материал изоляции",
        "Код материала изоляции",
        "Статус материала изоляции",
        "Комментарий материала изоляции",
        "T° среды",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        "Режим температуры изоляции",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Задвижки",
        "Фланцы",
        "Опоры",
        "L экв., м",
    ]
    for c, h in enumerate(pipe_cols, start=1):
        cell = ws_pipe.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")
    ws_pipe.append(
        [
            "Пример DN100",
            108,
            50,
            50,
            "Плиты минераловатные прошивные",
            "mineral_wool_boards_120",
            "",
            "",
            -20,
            80,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            2,
            2,
            2,
            1.5,
        ]
    )
    ws_pipe.append(
        [
            "Пример DN50",
            57,
            20,
            40,
            "Теплоизоляционные изделия из пенополиуретана",
            "polyurethane_products_50",
            "",
            "",
            -30,
            60,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            0,
            0,
            0,
            "",
        ]
    )
    ws_pipe.column_dimensions["A"].width = 24
    for col_idx in range(2, len(pipe_cols) + 1):
        ws_pipe.column_dimensions[get_column_letter(col_idx)].width = 18

    ws_tank = wb.create_sheet("Резервуары")
    tank_cols = [
        "Наименование",
        "Форма",
        "Диаметр, мм",
        "Длина, мм",
        "Ширина, мм",
        "Высота, мм",
        "Толщина изоляции, мм",
        "Материал изоляции",
        "Код материала изоляции",
        "Статус материала изоляции",
        "Комментарий материала изоляции",
        "T° среды",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        "Режим температуры изоляции",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Q доп., Вт",
    ]
    for c, h in enumerate(tank_cols, start=1):
        cell = ws_tank.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")
    ws_tank.append(
        [
            "Бак цил.",
            "Цилиндр",
            2000,
            "",
            "",
            3000,
            80,
            "Плиты минераловатные прошивные",
            "mineral_wool_boards_120",
            "",
            "",
            -20,
            80,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            0,
        ]
    )
    ws_tank.append(
        [
            "Бак прям.",
            "Параллелепипед",
            "",
            5000,
            3000,
            4000,
            80,
            "Плиты минераловатные прошивные",
            "mineral_wool_boards_120",
            "",
            "",
            -20,
            80,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            0,
        ]
    )
    ws_tank.append(
        [
            "Шаровой",
            "Шар",
            1500,
            "",
            "",
            "",
            60,
            "Изделия из ППУ",
            "polyurethane_products_50",
            "",
            "",
            -20,
            60,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            0,
        ]
    )
    ws_tank.column_dimensions["A"].width = 24
    for col_idx in range(2, len(tank_cols) + 1):
        ws_tank.column_dimensions[get_column_letter(col_idx)].width = 18

    ws_info = wb.create_sheet("Справка")
    ws_info["A1"] = "Подсказки по импорту"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A3"] = "Материалы изоляции: используйте конкретный код и плотность."
    ws_info["A3"].font = Font(bold=True)
    for i, mat in enumerate(
        [
            "mineral_wool_boards_120 — Плиты минераловатные прошивные, ρ 120",
            "polyurethane_products_50 — Изделия из ППУ, ρ 50",
            "polystyrene_products_50 — Изделия из полистирола, ρ 50",
            "mineral_wool_cylinders_100 — Цилиндры минераловатные, ρ 100",
            "fiberglass_mats_70 — Маты стеклянного штапельного волокна, ρ 70",
            "k_flex_st — Кайманфлекс K-Flex ST, ρ 60-80",
        ],
        start=4,
    ):
        ws_info.cell(row=i, column=1, value=mat)
    ws_info["A11"] = "Формы резервуара:"
    ws_info["A11"].font = Font(bold=True)
    ws_info["A12"] = "Цилиндр — требуются Диаметр и Высота"
    ws_info["A13"] = "Параллелепипед — требуются Длина, Ширина, Высота"
    ws_info["A14"] = "Шар — требуется Диаметр"
    ws_info.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_csv() -> bytes:
    """Формирует CSV-шаблон: один файл, колонка «Тип» разделяет трубы и резервуары."""
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", lineterminator="\n")
    writer.writerow(
        [
            "Тип",
            "Наименование",
            "Форма",
            "Диаметр, мм",
            "Длина, мм",
            "Ширина, мм",
            "Высота, мм",
            "Длина, м",  # для трубы — длина в метрах
            "Толщина изоляции, мм",
            "Материал изоляции",
            "Код материала изоляции",
            "Статус материала изоляции",
            "Комментарий материала изоляции",
            "T° среды",
            "T° продукта",
            "T проп., °C",
            "Размещение",
            "Режим температуры изоляции",
            "Климатический регион",
            "Климатический город",
            "Ключ климата",
            "Обеспеченность климата",
            "Kзап",
            "Мин. T включения, °C",
            "Задвижки",
            "Фланцы",
            "Опоры",
            "L экв., м",
            "Q доп., Вт",
        ]
    )
    # Для трубы: диаметр в мм, длина в м, форма/ширина/высота пустые
    writer.writerow(
        [
            "труба",
            "Пример DN100",
            "",
            108,
            "",
            "",
            "",
            50,
            50,
            "Плиты минераловатные прошивные",
            "mineral_wool_boards_120",
            "",
            "",
            -20,
            80,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            2,
            2,
            2,
            1.5,
            "",
        ]
    )
    writer.writerow(
        [
            "труба",
            "Пример DN50",
            "",
            57,
            "",
            "",
            "",
            20,
            40,
            "Теплоизоляционные изделия из пенополиуретана",
            "polyurethane_products_50",
            "",
            "",
            -30,
            60,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            0,
            0,
            0,
            "",
            "",
        ]
    )
    # Резервуары: заполняем Форма + нужные габариты в мм
    writer.writerow(
        [
            "резервуар",
            "Бак цил.",
            "Цилиндр",
            2000,
            "",
            "",
            3000,
            "",
            80,
            "Плиты минераловатные прошивные",
            "mineral_wool_boards_120",
            "",
            "",
            -20,
            80,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            "",
            "",
            "",
            "",
            0,
        ]
    )
    writer.writerow(
        [
            "резервуар",
            "Бак прям.",
            "Параллелепипед",
            "",
            5000,
            3000,
            4000,
            "",
            80,
            "Теплоизоляционные изделия из пенополиуретана",
            "polyurethane_products_50",
            "",
            "",
            -20,
            60,
            "",
            "outdoor",
            "outdoor_winter",
            "",
            "",
            "",
            "",
            1.1,
            -20,
            "",
            "",
            "",
            "",
            0,
        ]
    )
    writer.writerow(
        [
            "резервуар",
            "Шар",
            "Шар",
            1500,
            "",
            "",
            "",
            "",
            60,
            "Теплоизоляционные изделия из пенополиуретана",
            "polyurethane_products_50",
            "",
            "",
            -20,
            50,
            "",
            "outdoor",
            "outdoor_winter",
        ]
    )
    # UTF-8 BOM чтобы Excel правильно открывал файл
    return ("\ufeff" + buf.getvalue()).encode("utf-8")
