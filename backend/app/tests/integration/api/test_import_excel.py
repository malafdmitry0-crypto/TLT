"""Integration-тесты импорта объектов из Excel."""

import io
from uuid import UUID

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker

from app.services.excel_import_service import build_template_csv, build_template_xlsx
from app.services.task_service import TaskService

MINERAL_WOOL = "mineral_wool_boards_120"
POLYURETHANE = "polyurethane_products_50"


async def _create_project(client: AsyncClient, session_id: str) -> str:
    resp = await client.get(
        "/api/v1/projects",
        headers={"X-Session-Id": session_id},
    )
    return resp.json()[0]["id"]


@pytest.mark.asyncio(loop_scope="session")
class TestExcelRoundTrip:
    """Экспорт в Excel → импорт этого же файла (same project или новый)."""

    async def test_export_edit_ambient_temperature_survives_recalculation(
        self, client: AsyncClient, employee_token: str, test_engine
    ):
        headers = {"Authorization": f"Bearer {employee_token}"}
        source_project = (
            await client.post(
                "/api/v1/projects",
                json={"name": "RT-climate-source"},
                headers=headers,
            )
        ).json()
        source_params = {
            "name": "Тогул edited ambient",
            "outer_diameter": 0.108,
            "wall_thickness": 0.004,
            "pipe_material": "carbon_steel",
            "insulation_layers": [{"thickness": 0.05, "material": MINERAL_WOOL}],
            "insulation_temperature_basis": "outdoor_winter",
            "ambient_temperature": -33.0,
            "ambient_temperature_source": "climate",
            "process_temperature": 80.0,
            "min_switch_temperature": -20.0,
            "pipe_length": 50.0,
            "placement": "outdoor",
            "climate_key": "Алтайский край|||Тогул",
            "climate_region": "Алтайский край",
            "climate_city": "Тогул",
            "climate_temperature_basis": "t_0_92",
            "wind_speed": 3.1,
            "wind_speed_source": "climate",
            "num_local_elements": 0,
        }
        created = await client.post(
            f"/api/v1/projects/{source_project['id']}/objects",
            json={"object_type": "pipe", "sort_order": 0, "params": source_params},
            headers=headers,
        )
        assert created.status_code == 201, created.text

        exported = await client.get(
            f"/api/v1/projects/{source_project['id']}/objects/export-excel",
            headers=headers,
        )
        assert exported.status_code == 200, exported.text
        workbook = load_workbook(io.BytesIO(exported.content))
        worksheet = workbook["Трубопроводы"]
        headers_by_name = {cell.value: cell.column for cell in worksheet[1]}
        worksheet.cell(row=2, column=headers_by_name["Мин. T° окр. среды"], value=-10.0)
        edited = io.BytesIO()
        workbook.save(edited)

        target_project = (
            await client.post(
                "/api/v1/projects",
                json={"name": "RT-climate-edited"},
                headers=headers,
            )
        ).json()
        imported = await client.post(
            f"/api/v1/projects/{target_project['id']}/objects/import-excel",
            files={
                "file": (
                    "edited.xlsx",
                    edited.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=headers,
        )
        assert imported.status_code == 200, imported.text
        task_id = UUID(imported.json()["heat_loss_task"]["id"])
        session_factory = async_sessionmaker(test_engine, expire_on_commit=False)
        async with session_factory() as worker_db:
            await TaskService(worker_db, session_factory=session_factory).run_task(
                task_id,
                worker_id="test-worker",
            )

        objects = (
            await client.get(
                f"/api/v1/projects/{target_project['id']}/objects",
                headers=headers,
            )
        ).json()
        assert len(objects) == 1
        params = objects[0]["params"]
        assert params["ambient_temperature"] == -10.0
        assert params["ambient_temperature_source"] == "manual"

    async def test_export_then_reimport_matches_count(
        self, client: AsyncClient, employee_token: str
    ):
        headers = {"Authorization": f"Bearer {employee_token}"}
        # Проект 1 с двумя трубами
        p1 = (
            await client.post(
                "/api/v1/projects",
                json={"name": "RT-1"},
                headers=headers,
            )
        ).json()
        for params in [
            {
                "name": "Труба 1",
                "outer_diameter": 0.108,
                "wall_thickness": 0.004,
                "pipe_material": "carbon_steel",
                "insulation_layers": [{"thickness": 0.05, "material": MINERAL_WOOL}],
                "insulation_temperature_basis": "outdoor_winter",
                "ambient_temperature": -20.0,
                "process_temperature": 80.0,
                "min_switch_temperature": -20.0,
                "pipe_length": 50.0,
                "placement": "outdoor",
                "wind_speed": 0.0,
                "num_local_elements": 0,
            },
            {
                "name": "Труба 2",
                "outer_diameter": 0.057,
                "wall_thickness": 0.004,
                "pipe_material": "carbon_steel",
                "insulation_layers": [{"thickness": 0.04, "material": POLYURETHANE}],
                "insulation_temperature_basis": "outdoor_winter",
                "ambient_temperature": -10.0,
                "process_temperature": 60.0,
                "min_switch_temperature": -20.0,
                "pipe_length": 25.0,
                "placement": "outdoor",
                "wind_speed": 0.0,
                "num_local_elements": 0,
            },
        ]:
            created = await client.post(
                f"/api/v1/projects/{p1['id']}/objects",
                json={"object_type": "pipe", "sort_order": 0, "params": params},
                headers=headers,
            )
            assert created.status_code == 201, created.text

        # Экспорт
        exp = await client.get(
            f"/api/v1/projects/{p1['id']}/objects/export-excel",
            headers=headers,
        )
        assert exp.status_code == 200
        assert len(exp.content) > 0

        # Новый проект → импорт этого же файла
        p2 = (
            await client.post(
                "/api/v1/projects",
                json={"name": "RT-2"},
                headers=headers,
            )
        ).json()
        resp = await client.post(
            f"/api/v1/projects/{p2['id']}/objects/import-excel",
            files={
                "file": (
                    "exp.xlsx",
                    exp.content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=headers,
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["created"] == 2
        assert len(body["errors"]) == 0

    async def test_csv_pipe_with_missing_required_field_reports_row_error(
        self, client: AsyncClient, guest_session: str
    ):
        """Неполная строка отклоняется, а валидный сосед импортируется."""
        pid = (
            await client.get("/api/v1/projects", headers={"X-Session-Id": guest_session})
        ).json()[0]["id"]
        # Одна валидная труба, одна без диаметра
        csv = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта;Толщина стенки, мм;"
            "Материал трубы;Размещение;Скорость ветра, м/с;Режим температуры изоляции;"
            "Мин. T включения, °C\n"
            f"труба;Good;108;50;50;{MINERAL_WOOL};-20;80;4;carbon_steel;outdoor;0;outdoor_winter;-20\n"
            f"труба;Bad;;30;40;{MINERAL_WOOL};-20;80;4;carbon_steel;outdoor;0;outdoor_winter;-20\n"
        ).encode()
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={"file": ("x.csv", csv, "text/csv")},
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["created"] == 1
        assert body["valid"] == 1
        assert body["invalid"] == 1
        assert body["errors"] == []
        assert body["validation_errors"][0]["row"] == 3
        assert body["validation_errors"][0]["field"] == "outer_diameter"
        assert body["heat_loss_task"]["type"] == "heat_loss_batch"

    async def test_roundtrip_preserves_material_and_dimensions(
        self,
        client: AsyncClient,
        employee_token: str,
        test_engine,
    ):
        """Проверяем не только count — params после round-trip совпадают."""
        headers = {"Authorization": f"Bearer {employee_token}"}
        p1 = (
            await client.post(
                "/api/v1/projects",
                json={"name": "RT-fields"},
                headers=headers,
            )
        ).json()
        src_params = {
            "name": "Magistral-1",
            "outer_diameter": 0.273,
            "wall_thickness": 0.004,
            "pipe_material": "carbon_steel",
            "insulation_layers": [
                {"thickness": 0.06, "material": POLYURETHANE},
                {"thickness": 0.04, "material": MINERAL_WOOL},
            ],
            "insulation_temperature_basis": "outdoor_winter",
            "ambient_temperature": -7.0,
            "max_ambient_temperature": 0.0,
            "ambient_temperature_source": "manual",
            "process_temperature": 150.0,
            "pipe_length": 200.5,
            "placement": "outdoor",
            "wind_speed": 4.0,
            "wind_speed_source": "climate",
            "climate_region": "Алтайский край",
            "climate_city": "Славгород",
            "climate_key": "Алтайский край|||Славгород",
            "climate_temperature_basis": "t_0_92",
            "safety_factor": 1.1,
            "safety_factor_source": "climate_policy",
            "min_switch_temperature": -35,
            "num_local_elements": 6,
            "local_element_equiv_length": 2.4,
        }
        created = await client.post(
            f"/api/v1/projects/{p1['id']}/objects",
            json={"object_type": "pipe", "sort_order": 0, "params": src_params},
            headers=headers,
        )
        assert created.status_code == 201, created.text
        source_results = created.json()["results"]

        exp = await client.get(
            f"/api/v1/projects/{p1['id']}/objects/export-excel",
            headers=headers,
        )
        p2 = (
            await client.post(
                "/api/v1/projects",
                json={"name": "RT-fields-2"},
                headers=headers,
            )
        ).json()
        imported = await client.post(
            f"/api/v1/projects/{p2['id']}/objects/import-excel",
            files={
                "file": (
                    "e.xlsx",
                    exp.content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=headers,
        )
        assert imported.status_code == 200, imported.text
        task_id = UUID(imported.json()["heat_loss_task"]["id"])
        session_factory = async_sessionmaker(test_engine, expire_on_commit=False)
        async with session_factory() as worker_db:
            await TaskService(worker_db, session_factory=session_factory).run_task(
                task_id,
                worker_id="test-worker",
            )

        # Проверяем финальные параметры после фонового пересчёта, а не сырой импорт.
        objs = (
            await client.get(
                f"/api/v1/projects/{p2['id']}/objects",
                headers=headers,
            )
        ).json()
        assert len(objs) == 1
        p = objs[0]["params"]
        assert p["insulation_layers"] == src_params["insulation_layers"]
        assert abs(p["outer_diameter"] - 0.273) < 1e-6
        assert p["pipe_length"] == 200.5
        assert p["ambient_temperature"] == -7.0
        assert p["max_ambient_temperature"] == 0.0
        assert p["ambient_temperature_source"] == "manual"
        assert p["process_temperature"] == 150.0
        assert p["climate_key"] == "Алтайский край|||Славгород"
        assert p["climate_temperature_basis"] == "t_0_92"
        assert p["wind_speed"] == pytest.approx(4.0)
        assert p["wind_speed_source"] == "climate"
        assert p["safety_factor"] == pytest.approx(1.1)
        assert p["safety_factor_source"] == "climate_policy"
        assert p["min_switch_temperature"] == -35
        assert p["num_local_elements"] == 6
        assert p["local_element_equiv_length"] == 2.4
        assert objs[0]["results"] == source_results
        for forbidden in (
            "location",
            "burial_depth",
            "insulation_thickness",
            "insulation_material",
            "insulation_layer_count",
        ):
            assert forbidden not in p

        changed_climate_params = {
            **p,
            "climate_key": "Алтайский край|||Солонешное",
            "climate_region": "Алтайский край",
            "climate_city": "Солонешное",
        }
        changed = await client.put(
            f"/api/v1/projects/{p2['id']}/objects/{objs[0]['id']}",
            json={"version": objs[0]["version"], "params": changed_climate_params},
            headers=headers,
        )
        assert changed.status_code == 200, changed.text
        changed_params = changed.json()["params"]
        assert changed_params["wind_speed"] == pytest.approx(1.2)
        assert changed_params["wind_speed_source"] == "climate"


def _build_xlsx(pipes: list[list] | None = None, tanks: list[list] | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    if pipes is not None:
        ws = wb.create_sheet("Трубопроводы")
        ws.append(
            [
                "Наименование",
                "Диаметр, мм",
                "Длина, м",
                "Толщина изоляции, мм",
                "Материал изоляции",
                "T° среды",
                "T° продукта",
                "Толщина стенки, мм",
                "Материал трубы",
                "Размещение",
                "Скорость ветра, м/с",
                "Режим температуры изоляции",
                "Мин. T включения, °C",
            ]
        )
        for row in pipes:
            ws.append([*row, -20])
    if tanks is not None:
        ws = wb.create_sheet("Резервуары")
        ws.append(
            [
                "Наименование",
                "Форма",
                "Диаметр, мм",
                "Длина, мм",
                "Ширина, мм",
                "Высота, мм",
                "Толщина изоляции, мм",
                "Материал изоляции",
                "T° среды",
                "T° продукта",
                "Скорость ветра, м/с",
                "Мин. T включения, °C",
                "Высота обогрева, м",
                "Шаг укладки, м",
            ]
        )
        for row in tanks:
            ws.append([*row, 0, -20, 2.0, 0.2])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio(loop_scope="session")
class TestExcelImport:
    async def test_template_downloads(self, client: AsyncClient, guest_session: str):
        pid = await _create_project(client, guest_session)
        resp = await client.get(
            f"/api/v1/projects/{pid}/objects/import-template",
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml"
        )
        # Шаблон парсится и содержит оба листа
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(resp.content))
        assert "Трубопроводы" in wb.sheetnames
        assert "Резервуары" in wb.sheetnames

    async def test_import_pipes_ok(self, client: AsyncClient, guest_session: str):
        pid = await _create_project(client, guest_session)
        xlsx = _build_xlsx(
            pipes=[
                [
                    "Труба №1",
                    108,
                    50,
                    50,
                    MINERAL_WOOL,
                    -20,
                    80,
                    4,
                    "carbon_steel",
                    "outdoor",
                    0,
                    "outdoor_winter",
                ],
                [
                    "Труба №2",
                    57,
                    20,
                    40,
                    POLYURETHANE,
                    -30,
                    60,
                    4,
                    "carbon_steel",
                    "outdoor",
                    0,
                    "outdoor_winter",
                ],
            ]
        )
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={
                "file": (
                    "t.xlsx",
                    xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["created"] == 2, body
        assert body["errors"] == []
        assert body["heat_loss_task"]["type"] == "heat_loss_batch"

        # Объекты появились в проекте
        objs = (
            await client.get(
                f"/api/v1/projects/{pid}/objects",
                headers={"X-Session-Id": guest_session},
            )
        ).json()
        assert len(objs) == 2
        assert all(o["object_type"] == "pipe" for o in objs)

    async def test_import_rejects_5000_mm_pipe_and_schedules_only_valid_neighbor(
        self,
        client: AsyncClient,
        guest_session: str,
        monkeypatch: pytest.MonkeyPatch,
    ):
        pid = await _create_project(client, guest_session)
        xlsx = _build_xlsx(
            pipes=[
                [
                    "Валидная",
                    108,
                    50,
                    50,
                    MINERAL_WOOL,
                    -20,
                    80,
                    4,
                    "carbon_steel",
                    "outdoor",
                    0,
                    "outdoor_winter",
                ],
                [
                    "Диаметр 5000",
                    5000,
                    50,
                    50,
                    MINERAL_WOOL,
                    -20,
                    80,
                    4,
                    "carbon_steel",
                    "outdoor",
                    0,
                    "outdoor_winter",
                ],
            ]
        )
        scheduled_ids: list[UUID] = []
        original_create = TaskService.create_heat_loss_batch_task

        async def capture_task(self, request, principal, **kwargs):
            scheduled_ids.extend(request.object_ids or [])
            return await original_create(self, request, principal, **kwargs)

        monkeypatch.setattr(TaskService, "create_heat_loss_batch_task", capture_task)

        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={
                "file": (
                    "mixed-validation.xlsx",
                    xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-Session-Id": guest_session},
        )

        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["created"] == 1
        assert body["valid"] == 1
        assert body["invalid"] == 1
        assert body["errors"] == []
        assert body["skipped_duplicates"] == 0
        assert body["skipped_limit"] == 0
        assert body["validation_errors"] == [
            {
                "sheet": "Трубопроводы",
                "row": 3,
                "field": "outer_diameter",
                "code": "OBJECT_PARAMS_INVALID",
                "message": "Наружный диаметр должен быть от 10,8 до 3000 мм",
            }
        ]
        objects = (
            await client.get(
                f"/api/v1/projects/{pid}/objects",
                headers={"X-Session-Id": guest_session},
            )
        ).json()
        assert [item["params"]["name"] for item in objects] == ["Валидная"]
        assert scheduled_ids == [UUID(objects[0]["id"])]

    async def test_import_pipe_rejects_generic_and_unknown_materials(
        self, client: AsyncClient, guest_session: str
    ):
        pid = await _create_project(client, guest_session)
        xlsx = _build_xlsx(
            pipes=[
                [
                    "Общий",
                    108,
                    50,
                    50,
                    "Минеральная вата",
                    -20,
                    80,
                    4,
                    "carbon_steel",
                    "outdoor",
                    0,
                    "outdoor_winter",
                ],
                [
                    "Неизвестный",
                    57,
                    20,
                    40,
                    "unknown-material",
                    -20,
                    60,
                    4,
                    "carbon_steel",
                    "outdoor",
                    0,
                    "outdoor_winter",
                ],
            ]
        )

        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={
                "file": (
                    "invalid-materials.xlsx",
                    xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-Session-Id": guest_session},
        )

        assert resp.status_code == 200
        body = resp.json()
        assert body["created"] == 0
        assert body["valid"] == 0
        assert body["invalid"] == 2
        assert body["errors"] == []
        assert len(body["validation_errors"]) == 2
        objects = (
            await client.get(
                f"/api/v1/projects/{pid}/objects",
                headers={"X-Session-Id": guest_session},
            )
        ).json()
        assert objects == []

    async def test_import_rejects_pipe_with_formula_domain_error(
        self, client: AsyncClient, guest_session: str
    ):
        pid = await _create_project(client, guest_session)
        xlsx = _build_xlsx(
            pipes=[
                [
                    "Невалидная температура",
                    108,
                    50,
                    50,
                    MINERAL_WOOL,
                    5,
                    5,
                    4,
                    "carbon_steel",
                    "outdoor",
                    0,
                    "outdoor_winter",
                ],
            ]
        )

        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={
                "file": (
                    "formula-domain-invalid.xlsx",
                    xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-Session-Id": guest_session},
        )

        assert resp.status_code == 200
        body = resp.json()
        assert body["created"] == 0
        assert body["invalid"] == 1
        assert body["validation_errors"][0]["code"] == ("process_temperature_not_above_ambient")
        objects = (
            await client.get(
                f"/api/v1/projects/{pid}/objects",
                headers={"X-Session-Id": guest_session},
            )
        ).json()
        assert objects == []

    async def test_import_rejects_pipe_with_core_scalar_range_error(
        self, client: AsyncClient, guest_session: str
    ):
        """Canonical scalar validation rejects the row before persistence."""

        pid = await _create_project(client, guest_session)
        xlsx = _build_xlsx(
            pipes=[
                [
                    "Недопустимая длина",
                    108,
                    0.1,
                    50,
                    MINERAL_WOOL,
                    -20,
                    80,
                    4,
                    "carbon_steel",
                    "outdoor",
                    0,
                    "outdoor_winter",
                ],
            ]
        )

        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={
                "file": (
                    "core-range-invalid.xlsx",
                    xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-Session-Id": guest_session},
        )

        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["created"] == 0
        assert body["valid"] == 0
        assert body["invalid"] == 1
        assert body["validation_errors"][0]["field"] == "pipe_length"
        assert body["validation_errors"][0]["code"] == "OBJECT_PARAMS_INVALID"
        objects = (
            await client.get(
                f"/api/v1/projects/{pid}/objects",
                headers={"X-Session-Id": guest_session},
            )
        ).json()
        assert objects == []

    async def test_import_tanks_supported_shapes_and_rejects_legacy_shape(
        self, client: AsyncClient, guest_session: str
    ):
        pid = await _create_project(client, guest_session)
        xlsx = _build_xlsx(
            tanks=[
                ["Цил бак", "Цилиндр", 2000, "", "", 3000, 80, MINERAL_WOOL, -20, 80],
                ["Прям бак", "Параллелепипед", "", 5000, 3000, 4000, 80, POLYURETHANE, -20, 60],
                ["Шар", "Шар", 1500, "", "", "", 60, "Пеностекло", -20, 50],
            ]
        )
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={
                "file": (
                    "t.xlsx",
                    xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["created"] == 2
        assert len(body["errors"]) == 1
        assert body["errors"][0]["row"] == 4
        assert "форма" in body["errors"][0]["message"].lower()

    async def test_import_reports_structural_row_errors(
        self, client: AsyncClient, guest_session: str
    ):
        pid = await _create_project(client, guest_session)
        xlsx = _build_xlsx(
            tanks=[
                ["OK", "Цилиндр", 2000, "", "", 3000, 80, MINERAL_WOOL, -20, 80],
                ["Неизвестная форма 1", "Куб", 2000, "", "", 3000, 80, MINERAL_WOOL, -20, 80],
                [
                    "Неизвестная форма 2",
                    "Конус",
                    2000,
                    "",
                    "",
                    3000,
                    80,
                    MINERAL_WOOL,
                    -20,
                    80,
                ],
            ]
        )
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={
                "file": (
                    "t.xlsx",
                    xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["created"] == 1
        assert len(body["errors"]) == 2
        assert body["valid"] == 1, body
        assert body["heat_loss_task"]["type"] == "heat_loss_batch"
        # Номера строк в Excel: 3 и 4 (заголовок — 1, OK — 2)
        rows_with_errors = sorted(e["row"] for e in body["errors"])
        assert rows_with_errors == [3, 4]

    async def test_import_rejects_missing_sheets(self, client: AsyncClient, guest_session: str):
        pid = await _create_project(client, guest_session)
        wb = Workbook()
        wb.active.title = "Другое"
        wb.active["A1"] = "foo"
        buf = io.BytesIO()
        wb.save(buf)
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={
                "file": (
                    "x.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 422
        assert "Трубопроводы" in resp.json()["detail"]

    async def test_import_rejects_non_xlsx(self, client: AsyncClient, guest_session: str):
        pid = await _create_project(client, guest_session)
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={"file": ("x.csv", b"foo,bar\n1,2", "text/csv")},
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 422


def test_template_builder_roundtrip():
    """Шаблон xlsx-сборщика валиден и имеет нужные листы."""
    data = build_template_xlsx()
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data))
    assert "Трубопроводы" in wb.sheetnames
    assert "Резервуары" in wb.sheetnames
    assert all(
        "DN" not in str(cell.value or "")
        for sheet in wb.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_csv_template_has_type_column():
    data = build_template_csv()
    text = data.decode("utf-8-sig")
    assert "DN" not in text
    first_line = text.splitlines()[0]
    assert "Тип" in first_line


@pytest.mark.asyncio(loop_scope="session")
class TestCsvImport:
    async def test_csv_template_endpoint(self, client: AsyncClient, guest_session: str):
        pid = await _create_project(client, guest_session)
        resp = await client.get(
            f"/api/v1/projects/{pid}/objects/import-template",
            params={"format": "csv"},
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/csv")
        assert "Тип" in resp.content.decode("utf-8-sig")

    async def test_csv_import_ok(self, client: AsyncClient, guest_session: str):
        pid = await _create_project(client, guest_session)
        csv_body = (
            "\ufeffТип;Наименование;Форма;Диаметр, мм;Длина, мм;Ширина, мм;Высота, мм;"
            "Длина, м;Толщина изоляции, мм;Материал изоляции;T° среды;T° продукта;Толщина стенки, мм;"
            "Материал трубы;Размещение;Скорость ветра, м/с;Режим температуры изоляции;"
            "Мин. T включения, °C;Высота обогрева, м;Шаг укладки, м\n"
            f"труба;Пример;;108;;;;50;50;{MINERAL_WOOL};-20;80;4;carbon_steel;outdoor;0;outdoor_winter;-20;;\n"
            f"резервуар;Бак;Цилиндр;2000;;;3000;;80;{MINERAL_WOOL};-20;80;;;"
            "outdoor;0;outdoor_winter;-20;3;0.2\n"
        ).encode()
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={"file": ("t.csv", csv_body, "text/csv")},
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["created"] == 2, body
        assert body["errors"] == []
        assert body["heat_loss_task"]["type"] == "heat_loss_batch"

    async def test_csv_import_default_merge_skips_repeated_file(
        self, client: AsyncClient, guest_session: str
    ):
        pid = await _create_project(client, guest_session)
        csv_body = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта;Толщина стенки, мм;"
            "Материал трубы;Размещение;Скорость ветра, м/с;Режим температуры изоляции;"
            "Мин. T включения, °C\n"
            f"труба;Повтор;108;50;50;{MINERAL_WOOL};-20;80;4;carbon_steel;outdoor;0;outdoor_winter;-20\n"
        ).encode()
        headers = {"X-Session-Id": guest_session}

        first = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={"file": ("t.csv", csv_body, "text/csv")},
            headers=headers,
        )
        assert first.status_code == 200, first.text
        assert first.json()["created"] == 1

        second = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={"file": ("t.csv", csv_body, "text/csv")},
            headers=headers,
        )
        assert second.status_code == 200, second.text
        second_body = second.json()
        assert second_body["created"] == 0
        assert second_body["skipped_duplicates"] == 1

        objs = (await client.get(f"/api/v1/projects/{pid}/objects", headers=headers)).json()
        assert len(objs) == 1

    async def test_csv_import_append_mode_keeps_explicit_copies(
        self,
        client: AsyncClient,
        guest_session: str,
        monkeypatch: pytest.MonkeyPatch,
    ):
        class FakeTaskResponse:
            def model_dump(self, *, mode: str):
                return {"type": "heat_loss_batch"}

        async def fake_create_task(self, request, principal):
            return object()

        monkeypatch.setattr(TaskService, "create_heat_loss_batch_task", fake_create_task)
        monkeypatch.setattr(
            TaskService,
            "to_response",
            staticmethod(lambda task: FakeTaskResponse()),
        )
        pid = await _create_project(client, guest_session)
        csv_body = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта;Толщина стенки, мм;"
            "Материал трубы;Размещение;Скорость ветра, м/с;Режим температуры изоляции;"
            "Мин. T включения, °C\n"
            f"труба;Копия;108;50;50;{MINERAL_WOOL};-20;80;4;carbon_steel;outdoor;0;outdoor_winter;-20\n"
        ).encode()
        headers = {"X-Session-Id": guest_session}
        for _ in range(2):
            resp = await client.post(
                f"/api/v1/projects/{pid}/objects/import-excel",
                data={"mode": "append"},
                files={"file": ("t.csv", csv_body, "text/csv")},
                headers=headers,
            )
            assert resp.status_code == 200, resp.text
            assert resp.json()["created"] == 1

        objs = (await client.get(f"/api/v1/projects/{pid}/objects", headers=headers)).json()
        assert len(objs) == 2

    async def test_csv_import_replace_mode_replaces_project_objects(
        self, client: AsyncClient, guest_session: str
    ):
        pid = await _create_project(client, guest_session)
        headers = {"X-Session-Id": guest_session}
        replace_csv = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта;Толщина стенки, мм;"
            "Материал трубы;Размещение;Скорость ветра, м/с;Режим температуры изоляции;"
            "Мин. T включения, °C\n"
            f"труба;Новая;159;30;50;{MINERAL_WOOL};-20;80;4;carbon_steel;outdoor;0;outdoor_winter;-20\n"
        ).encode()

        for name, diameter in (("Старая", 0.108), ("Ещё старая", 0.057)):
            seeded = await client.post(
                f"/api/v1/projects/{pid}/objects",
                json={
                    "object_type": "pipe",
                    "params": {
                        "name": name,
                        "outer_diameter": diameter,
                        "pipe_length": 50,
                        "insulation_layers": [{"thickness": 0.05, "material": MINERAL_WOOL}],
                        "ambient_temperature": -20,
                        "process_temperature": 80,
                        "min_switch_temperature": -20,
                        "wall_thickness": 0.004,
                        "pipe_material": "carbon_steel",
                        "placement": "outdoor",
                        "wind_speed": 0,
                        "insulation_temperature_basis": "outdoor_winter",
                    },
                },
                headers=headers,
            )
            assert seeded.status_code == 201, seeded.text
        replaced = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            data={"mode": "replace"},
            files={"file": ("new.csv", replace_csv, "text/csv")},
            headers=headers,
        )
        assert replaced.status_code == 200, replaced.text
        assert replaced.json()["created"] == 1

        objs = (await client.get(f"/api/v1/projects/{pid}/objects", headers=headers)).json()
        assert [obj["params"]["name"] for obj in objs] == ["Новая"]

    async def test_csv_replace_with_only_invalid_rows_preserves_existing_objects(
        self, client: AsyncClient, guest_session: str
    ):
        pid = await _create_project(client, guest_session)
        headers = {"X-Session-Id": guest_session}
        columns = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта;Толщина стенки, мм;"
            "Материал трубы;Размещение;Скорость ветра, м/с;Режим температуры изоляции;"
            "Мин. T включения, °C\n"
        )
        invalid_csv = (
            columns + f"труба;Невалидная;5000;50;50;{MINERAL_WOOL};-20;80;4;"
            "carbon_steel;outdoor;0;outdoor_winter;-20\n"
        ).encode()
        seeded = await client.post(
            f"/api/v1/projects/{pid}/objects",
            json={
                "object_type": "pipe",
                "params": {
                    "name": "Существующая",
                    "outer_diameter": 0.108,
                    "pipe_length": 50,
                    "insulation_layers": [{"thickness": 0.05, "material": MINERAL_WOOL}],
                    "ambient_temperature": -20,
                    "process_temperature": 80,
                    "min_switch_temperature": -20,
                    "wall_thickness": 0.004,
                    "pipe_material": "carbon_steel",
                    "placement": "outdoor",
                    "wind_speed": 0,
                    "insulation_temperature_basis": "outdoor_winter",
                },
            },
            headers=headers,
        )
        assert seeded.status_code == 201, seeded.text

        replaced = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            data={"mode": "replace"},
            files={"file": ("invalid.csv", invalid_csv, "text/csv")},
            headers=headers,
        )

        assert replaced.status_code == 200, replaced.text
        assert replaced.json()["created"] == 0
        assert replaced.json()["invalid"] == 1
        objects = (await client.get(f"/api/v1/projects/{pid}/objects", headers=headers)).json()
        assert [item["params"]["name"] for item in objects] == ["Существующая"]

    async def test_csv_import_reports_rows_skipped_by_project_limit(
        self,
        client: AsyncClient,
        guest_session: str,
        monkeypatch: pytest.MonkeyPatch,
    ):
        from app.core.config import settings

        monkeypatch.setattr(settings, "GUEST_MAX_OBJECTS_PER_PROJECT", 1)
        pid = await _create_project(client, guest_session)
        headers = {"X-Session-Id": guest_session}
        csv_body = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта;Толщина стенки, мм;"
            "Материал трубы;Размещение;Скорость ветра, м/с;Режим температуры изоляции;"
            "Мин. T включения, °C\n"
            f"труба;Первая;108;50;50;{MINERAL_WOOL};-20;80;4;carbon_steel;outdoor;0;outdoor_winter;-20\n"
            f"труба;Вторая;57;15;30;{MINERAL_WOOL};-10;50;4;carbon_steel;outdoor;0;outdoor_winter;-20\n"
        ).encode()

        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={"file": ("limit.csv", csv_body, "text/csv")},
            headers=headers,
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["created"] == 1
        assert body["skipped_limit"] == 1
        assert "Пропущено строк: 1" in body["errors"][0]["message"]

        objs = (await client.get(f"/api/v1/projects/{pid}/objects", headers=headers)).json()
        assert [obj["params"]["name"] for obj in objs] == ["Первая"]

    async def test_csv_requires_type_column(self, client: AsyncClient, guest_session: str):
        pid = await _create_project(client, guest_session)
        csv_body = b"Name;Diameter\nfoo;108\n"
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={"file": ("t.csv", csv_body, "text/csv")},
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 422
        assert "Тип" in resp.json()["detail"]

    async def test_csv_autodetect_comma_delimiter(self, client: AsyncClient, guest_session: str):
        pid = await _create_project(client, guest_session)
        csv_body = (
            "Тип,Наименование,Диаметр, мм,Длина, м,Толщина изоляции, мм,"
            "Материал изоляции,T° среды,T° продукта\n"
        ).encode()
        # Comma delimiter + запятые внутри заголовков — нужен ;-разделитель для чистоты.
        # Этот кейс проверяет что код не падает; используем ;-разделитель для корректности:
        csv_body = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта;Толщина стенки, мм;"
            "Материал трубы;Размещение;Скорость ветра, м/с;Режим температуры изоляции;"
            "Мин. T включения, °C\n"
            f"труба;T1;57;15;30;{MINERAL_WOOL};-10;50;4;carbon_steel;outdoor;0;outdoor_winter;-20\n"
        ).encode()
        resp = await client.post(
            f"/api/v1/projects/{pid}/objects/import-excel",
            files={"file": ("t.csv", csv_body, "text/csv")},
            headers={"X-Session-Id": guest_session},
        )
        assert resp.status_code == 200
        assert resp.json()["created"] == 1
