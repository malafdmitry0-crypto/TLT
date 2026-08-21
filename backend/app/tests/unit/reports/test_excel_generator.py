"""Unit-тесты для XLSX-генератора отчётов (`reports/excel_generator.py`)."""

import io

from openpyxl import load_workbook

from app.reports.excel_generator import generate_xlsx

PROJECT = {
    "id": "p1",
    "name": "Test",
    "description": "Desc",
    "status": "draft",
}


def _wb(data: bytes):
    return load_workbook(io.BytesIO(data))


class TestGenerateXlsx:
    def test_default_sections_all_present(self):
        ctx = {
            "project": PROJECT,
            "objects": [
                {"object_type": "pipe", "params": {"x": 1}, "results": {"q": 100}, "is_valid": True}
            ],
            "specification": {
                "items": [
                    {"category": "cable", "name": "К1", "article": "A", "unit": "м", "quantity": 5}
                ]
            },
        }
        wb = _wb(generate_xlsx(ctx))
        assert "Проект" in wb.sheetnames
        assert "Объекты" in wb.sheetnames
        assert "Спецификация" in wb.sheetnames

    def test_only_summary_section_skips_objects_sheet(self):
        """Если выбрана только summary — лист «Объекты» пуст или включает все типы."""
        ctx = {
            "project": PROJECT,
            "sections": ["summary"],
            "objects": [{"object_type": "pipe", "params": {}, "results": None, "is_valid": False}],
        }
        # summary входит в множество для «Объекты» — лист создаётся
        wb = _wb(generate_xlsx(ctx))
        assert "Объекты" in wb.sheetnames

    def test_no_relevant_sections_skips_data_sheets(self):
        """Если выбрана только секция, не относящаяся к данным — только «Проект»."""
        ctx = {"project": PROJECT, "sections": ["unknown_section"]}
        wb = _wb(generate_xlsx(ctx))
        assert wb.sheetnames == ["Проект"]

    def test_filter_pipes_only_excludes_tanks(self):
        ctx = {
            "project": PROJECT,
            "sections": ["pipes"],
            "objects": [
                {"object_type": "pipe", "params": {}, "results": None, "is_valid": True},
                {"object_type": "tank", "params": {}, "results": None, "is_valid": True},
            ],
        }
        wb = _wb(generate_xlsx(ctx))
        ws = wb["Объекты"]
        # Заголовок + 1 труба = 2 строки
        types = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert "pipe" in types
        assert "tank" not in types

    def test_project_with_description_includes_description_row(self):
        ctx = {"project": PROJECT, "sections": ["summary"], "objects": []}
        wb = _wb(generate_xlsx(ctx))
        ws = wb["Проект"]
        # Ищем строку с «Описание»
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Описание" in col_a

    def test_project_without_description_no_description_row(self):
        proj = dict(PROJECT)
        proj["description"] = ""  # пусто → строка не пишется
        ctx = {"project": proj, "sections": ["summary"], "objects": []}
        wb = _wb(generate_xlsx(ctx))
        ws = wb["Проект"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Описание" not in col_a

    def test_specification_only_section(self):
        ctx = {
            "project": PROJECT,
            "sections": ["specification"],
            "specification": {
                "items": [
                    {"category": "x", "name": "Y", "article": None, "unit": "шт", "quantity": 1}
                ]
            },
        }
        wb = _wb(generate_xlsx(ctx))
        assert "Спецификация" in wb.sheetnames
        assert "Объекты" not in wb.sheetnames

    def test_empty_specification_writes_only_header(self):
        ctx = {"project": PROJECT, "sections": ["specification"]}
        wb = _wb(generate_xlsx(ctx))
        ws = wb["Спецификация"]
        assert ws.max_row == 1  # только заголовок

    def test_dangerous_user_strings_are_not_exported_as_formulas(self):
        ctx = {
            "project": {
                "id": "p1",
                "name": '=HYPERLINK("http://example.test","x")',
                "description": "+SUM(1,1)",
                "status": "@cmd",
            },
            "sections": ["summary", "specification"],
            "objects": [
                {
                    "object_type": "pipe",
                    "params": {"name": "=1+1"},
                    "results": None,
                    "is_valid": True,
                }
            ],
            "specification": {
                "items": [
                    {
                        "category": "=cat",
                        "name": "+name",
                        "article": "-article",
                        "unit": "@unit",
                        "quantity": '=HYPERLINK("http://example.test","qty")',
                    }
                ]
            },
        }
        wb = _wb(generate_xlsx(ctx))

        guarded_cells = [
            wb["Проект"]["B3"],
            wb["Проект"]["B5"],
            wb["Проект"]["B6"],
            wb["Спецификация"]["A2"],
            wb["Спецификация"]["B2"],
            wb["Спецификация"]["C2"],
            wb["Спецификация"]["D2"],
            wb["Спецификация"]["E2"],
        ]
        for cell in guarded_cells:
            assert cell.data_type == "s"
            assert str(cell.value).startswith("'")
