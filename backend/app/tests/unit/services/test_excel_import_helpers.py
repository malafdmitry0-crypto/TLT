"""Unit-тесты чистых функций-парсеров excel_import_service.

Методология: table-driven testing. Эти функции получают произвольные строки
из Excel (локалями, запятыми, unicode) и должны корректно превращать их в
питоновские типы. Любая ошибка здесь = неправильный расчёт.
"""

import io
import zipfile

import pytest
from sqlalchemy.exc import SQLAlchemyError

from app.core.config import settings
from app.models.project_object import ProjectObject
from app.services.excel_import_service import (
    import_objects_from_csv,
    import_objects_from_excel,
)
from app.services.heat_loss_application import apply_climate_policy
from app.services.object_spreadsheet import persistence, preparation
from app.services.object_spreadsheet.export import build_objects_xlsx
from app.services.object_spreadsheet.mapping import (
    GENERIC_MATERIAL_ALIASES,
    SHAPE_ALIASES,
    SPECIAL_MATERIAL_ALIASES,
    _mark_edited_climate_temperature_as_manual,
    _norm,
    _resolve_material,
    _resolve_material_entry,
    _resolve_shape,
    _to_float,
)
from app.services.object_spreadsheet.parsing import (
    ExcelImportError,
    _parse_csv,
    _parse_excel_workbook,
    _validate_xlsx_archive,
)
from app.services.object_spreadsheet.persistence import (
    _commit_object_batch,
    _commit_object_batch_row_by_row,
)
from app.services.object_spreadsheet.pipe_mapping import _build_pipe_params
from app.services.object_spreadsheet.tank_mapping import _build_tank_params
from app.services.object_spreadsheet.templates import build_template_csv, build_template_xlsx


class TestNorm:
    @pytest.mark.parametrize(
        "inp,expected",
        [
            ("Наименование", "наименование"),
            ("  Длина, м  ", "длина, м"),
            ("T°  среды", "t° среды"),
            ("", ""),
            (None, ""),
            (42, "42"),
        ],
    )
    def test_variants(self, inp, expected):
        assert _norm(inp) == expected


class TestXlsxArchiveGuard:
    def test_rejects_large_uncompressed_archive(self, monkeypatch):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("xl/worksheets/sheet1.xml", "x" * 2048)

        monkeypatch.setattr(settings, "MAX_XLSX_UNCOMPRESSED_BYTES", 1024)

        with pytest.raises(ExcelImportError, match="распаковки"):
            _validate_xlsx_archive(buf.getvalue())


class TestBuildObjectsXlsxSafety:
    @pytest.mark.parametrize(
        ("object_type", "placement", "minimum", "maximum", "expected_maximum"),
        [
            ("pipe", "outdoor", -20.0, 0.0, 0.0),
            ("pipe", "indoor", -20.0, -5.0, -5.0),
            ("pipe", "underground", -20.0, 15.0, None),
            ("tank", "outdoor", -20.0, 0.0, 0.0),
            ("tank", "underground", -20.0, 15.0, 15.0),
        ],
    )
    def test_objects_xlsx_roundtrips_distinct_ambient_bounds(
        self,
        object_type: str,
        placement: str,
        minimum: float,
        maximum: float,
        expected_maximum: float | None,
    ):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        source = {
            "name": "Ambient bounds",
            "shape": "cylindrical",
            "ambient_temperature": minimum,
            "max_ambient_temperature": maximum,
            "ambient_temperature_source": "manual",
            "placement": placement,
        }
        content = build_objects_xlsx([SimpleNamespace(object_type=object_type, params=source)])
        sheet_name = "Трубопроводы" if object_type == "pipe" else "Резервуары"
        worksheet = load_workbook(io.BytesIO(content), data_only=True)[sheet_name]
        exported_row = dict(
            zip(
                [cell.value for cell in worksheet[1]],
                [cell.value for cell in worksheet[2]],
                strict=True,
            )
        )
        if placement == "underground" and object_type == "pipe":
            assert exported_row["Мин. T° окр. среды"] is None
            assert exported_row["Макс. T° окр. среды"] is None
            assert exported_row["Источник T° среды"] is None
        else:
            assert exported_row["Мин. T° окр. среды"] == minimum
            assert exported_row["Макс. T° окр. среды"] == maximum
            assert exported_row["Источник T° среды"] == "manual"
        [(label, parsed_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == object_type
        ]
        builder = _build_pipe_params if object_type == "pipe" else _build_tank_params
        params, err = builder(rows[0])

        assert label in {"Трубопроводы", "Резервуары"}
        assert parsed_type == object_type
        assert err is None
        assert params is not None
        if placement == "underground" and object_type == "pipe":
            assert "ambient_temperature" not in params
        else:
            assert params["ambient_temperature"] == minimum
        if expected_maximum is None:
            assert "max_ambient_temperature" not in params
        else:
            assert params["max_ambient_temperature"] == expected_maximum

    @pytest.mark.parametrize("object_type", ["pipe", "tank"])
    def test_objects_xlsx_keeps_blank_ambient_maximum_absent(self, object_type: str):
        from types import SimpleNamespace

        content = build_objects_xlsx(
            [
                SimpleNamespace(
                    object_type=object_type,
                    params={
                        "shape": "cylindrical",
                        "ambient_temperature": -20.0,
                        "placement": "outdoor",
                    },
                )
            ]
        )
        [(_label, _parsed_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == object_type
        ]
        builder = _build_pipe_params if object_type == "pipe" else _build_tank_params
        params, err = builder(rows[0])

        assert err is None
        assert params is not None
        assert "max_ambient_temperature" not in params

    def test_dangerous_object_strings_are_not_exported_as_formulas(self):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        obj = SimpleNamespace(
            object_type="pipe",
            params={
                "name": '=HYPERLINK("http://example.test","x")',
                "outer_diameter": 0.108,
                "pipe_length": "+SUM(1,1)",
                "wall_thickness": 0.004,
                "pipe_material": "carbon_steel",
                "insulation_layers": [{"thickness": 0.05, "material": "+SUM(1,1)"}],
                "ambient_temperature": "-2+3",
                "process_temperature": "=1+1",
                "vapor_temperature": "@cmd",
            },
        )

        wb = load_workbook(io.BytesIO(build_objects_xlsx([obj])), data_only=False)
        ws = wb["Трубопроводы"]
        cells_by_header = {
            header.value: ws.cell(row=2, column=index)
            for index, header in enumerate(ws[1], start=1)
        }
        for header in (
            "Наименование",
            "Длина, м",
            "Код материала изоляции",
            "Мин. T° окр. среды",
            "T° продукта",
            "T проп., °C",
        ):
            cell = cells_by_header[header]
            assert cell.data_type == "s"
            assert str(cell.value).startswith("'")

    def test_objects_xlsx_does_not_crash_on_malicious_dimension_strings(self):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        obj = SimpleNamespace(
            object_type="tank",
            params={
                "name": "tank",
                "shape": "rectangular",
                "diameter": "=1+1",
                "length": "+SUM(1,1)",
                "width": "-2+3",
                "height": "@cmd",
                "insulation_layers": [{"thickness": "=10", "material": "mineral_wool_boards_120"}],
                "ambient_temperature": -20,
                "process_temperature": 80,
            },
        )

        wb = load_workbook(io.BytesIO(build_objects_xlsx([obj])), data_only=False)
        ws = wb["Резервуары"]
        for cell_ref in ("C2", "D2", "E2", "F2", "G2"):
            cell = ws[cell_ref]
            assert cell.data_type == "s"
            assert str(cell.value).startswith("'")

    def test_objects_xlsx_exports_extended_roundtrip_fields(self):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        obj = SimpleNamespace(
            object_type="pipe",
            params={
                "name": "pipe",
                "outer_diameter": 0.108,
                "pipe_length": 50,
                "wall_thickness": 0.004,
                "pipe_material": "carbon_steel",
                "insulation_layers": [{"thickness": 0.05, "material": "mineral_wool_boards_120"}],
                "ambient_temperature": -20,
                "process_temperature": 80,
                "placement": "outdoor",
                "wind_speed": 3.5,
                "climate_region": "ХМАО",
                "climate_city": "Сургут",
                "climate_key": "ХМАО|||Сургут",
                "climate_temperature_basis": "t_0_92",
                "safety_factor": 1.2,
                "min_switch_temperature": -35,
                "num_local_elements": 6,
                "local_element_equiv_length": 2.4,
            },
        )

        wb = load_workbook(io.BytesIO(build_objects_xlsx([obj])), data_only=True)
        ws = wb["Трубопроводы"]
        headers = [cell.value for cell in ws[1]]
        row = dict(zip(headers, [cell.value for cell in ws[2]], strict=True))

        assert row["Ключ климата"] == "ХМАО|||Сургут"
        assert row["Обеспеченность климата"] == "t_0_92"
        assert row["Мин. T включения, °C"] == -35
        assert row["Количество локальных элементов"] == 6
        assert row["L экв., м"] == 2.4
        assert all(
            forbidden not in header.casefold()
            for header in headers
            for forbidden in ("location", "burial_depth", "количество слоёв")
        )

    @pytest.mark.parametrize(
        ("ambient_source", "expected_temperature"),
        [("manual", -7.0), ("climate", -33.0)],
    )
    def test_pipe_export_roundtrip_preserves_ambient_temperature_source(
        self,
        ambient_source: str,
        expected_temperature: float,
    ):
        from types import SimpleNamespace

        source = {
            "name": "Pipe climate provenance",
            "outer_diameter": 0.108,
            "pipe_length": 50.0,
            "ambient_temperature": -33.0 if ambient_source == "climate" else -7.0,
            "ambient_temperature_source": ambient_source,
            "process_temperature": 80.0,
            "placement": "outdoor",
            "climate_key": "Алтайский край|||Тогул",
            "climate_region": "Алтайский край",
            "climate_city": "Тогул",
            "climate_temperature_basis": "t_0_92",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="pipe", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "pipe"
        ]
        params, err = _build_pipe_params(rows[0])

        assert err is None
        assert params is not None
        assert params["ambient_temperature_source"] == ambient_source
        after_policy = apply_climate_policy("pipe", params)
        assert after_policy["ambient_temperature"] == pytest.approx(expected_temperature)
        assert after_policy["ambient_temperature_source"] == ambient_source

    @pytest.mark.parametrize("object_type", ["pipe", "tank"])
    @pytest.mark.parametrize("edited_temperature", [-10.0, 0.0])
    def test_exported_climate_temperature_edit_becomes_manual_override(
        self, object_type: str, edited_temperature: float
    ):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        source = {
            "name": "Edited climate temperature",
            "outer_diameter": 0.108,
            "pipe_length": 50.0,
            "shape": "cylindrical",
            "diameter": 2.0,
            "height": 3.0,
            "ambient_temperature": -33.0,
            "ambient_temperature_source": "climate",
            "process_temperature": 80.0,
            "placement": "outdoor",
            "climate_key": "Алтайский край|||Тогул",
            "climate_region": "Алтайский край",
            "climate_city": "Тогул",
            "climate_temperature_basis": "t_0_92",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type=object_type, params=source)])
        workbook = load_workbook(io.BytesIO(content))
        sheet_name = "Трубопроводы" if object_type == "pipe" else "Резервуары"
        worksheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        worksheet.cell(row=2, column=headers["Мин. T° окр. среды"], value=edited_temperature)
        edited = io.BytesIO()
        workbook.save(edited)

        [(_label, _parsed_type, rows)] = [
            parsed
            for parsed in _parse_excel_workbook(edited.getvalue())
            if parsed[1] == object_type
        ]
        builder = _build_pipe_params if object_type == "pipe" else _build_tank_params
        params, err = builder(rows[0])

        assert err is None
        assert params is not None
        assert params["ambient_temperature"] == edited_temperature
        assert params["ambient_temperature_source"] == "manual"
        after_policy = apply_climate_policy(object_type, params)
        assert after_policy["ambient_temperature"] == edited_temperature
        assert after_policy["ambient_temperature_source"] == "manual"

    def test_underground_pipe_ignores_ambient_temperature_even_with_climate_source(self):
        params, err = _build_pipe_params(
            {
                "outer_diameter_mm": 108,
                "pipe_length": 50,
                "ambient_temperature": -33,
                "ambient_temperature_source": "climate",
                "placement": "underground",
                "pipe_centerline_depth": 1.0,
                "climate_key": "Алтайский край|||Тогул",
                "climate_region": "Алтайский край",
                "climate_city": "Тогул",
                "climate_temperature_basis": "t_0_92",
            }
        )

        assert err is None
        assert params is not None
        assert "ambient_temperature" not in params
        assert "ambient_temperature_source" not in params
        after_policy = apply_climate_policy("pipe", params)
        assert "ambient_temperature" not in after_policy
        assert "ambient_temperature_source" not in after_policy

    @pytest.mark.parametrize(
        ("source", "expected_source"),
        [(None, None), ("manual", "manual"), ("climate", "climate")],
    )
    def test_climate_temperature_provenance_does_not_guess_without_known_reference(
        self, source: str | None, expected_source: str | None
    ):
        params = {
            "outer_diameter": 0.108,
            "placement": "outdoor",
            "ambient_temperature": -10.0,
            "ambient_temperature_source": source,
        }
        if source == "climate":
            params["climate_key"] = "unknown|||Unknown"

        _mark_edited_climate_temperature_as_manual("pipe", params)

        assert params["ambient_temperature_source"] == expected_source

    def test_tank_export_roundtrip_preserves_manual_ambient_temperature(self):
        from types import SimpleNamespace

        source = {
            "name": "Tank climate provenance",
            "shape": "cylindrical",
            "diameter": 2.0,
            "height": 3.0,
            "ambient_temperature": -7.0,
            "ambient_temperature_source": "manual",
            "process_temperature": 80.0,
            "placement": "outdoor",
            "climate_key": "Москва|||Москва",
            "climate_region": "Москва",
            "climate_city": "Москва",
            "climate_temperature_basis": "t_0_92",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="tank", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "tank"
        ]
        params, err = _build_tank_params(rows[0])

        assert err is None
        assert params is not None
        assert params["ambient_temperature_source"] == "manual"
        after_policy = apply_climate_policy("tank", params)
        assert after_policy["ambient_temperature"] == pytest.approx(-7.0)
        assert after_policy["ambient_temperature_source"] == "manual"

    def test_pipe_export_roundtrip_preserves_climate_wind_and_safety_sources(self):
        from types import SimpleNamespace

        source = {
            "name": "Pipe climate provenance",
            "outer_diameter": 0.108,
            "pipe_length": 50.0,
            "process_temperature": 80.0,
            "placement": "outdoor",
            "climate_key": "Алтайский край|||Славгород",
            "climate_region": "Алтайский край",
            "climate_city": "Славгород",
            "climate_temperature_basis": "t_0_92",
            "wind_speed": 4.0,
            "wind_speed_source": "climate",
            "safety_factor": 1.1,
            "safety_factor_source": "climate_policy",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="pipe", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "pipe"
        ]
        params, err = _build_pipe_params(rows[0])

        assert err is None
        assert params is not None
        assert params["wind_speed_source"] == "climate"
        assert params["safety_factor_source"] == "climate_policy"

        params.update(
            climate_key="Алтайский край|||Солонешное",
            climate_region="Алтайский край",
            climate_city="Солонешное",
        )
        after_policy = apply_climate_policy("pipe", params)
        assert after_policy["wind_speed"] == pytest.approx(1.2)
        assert after_policy["wind_speed_source"] == "climate"
        assert after_policy["safety_factor"] == pytest.approx(1.1)
        assert after_policy["safety_factor_source"] == "climate_policy"

    def test_tank_export_roundtrip_preserves_wind_and_safety_sources(self):
        from types import SimpleNamespace

        source = {
            "name": "Tank climate provenance",
            "shape": "cylindrical",
            "diameter": 2.0,
            "height": 3.0,
            "process_temperature": 80.0,
            "placement": "outdoor",
            "wind_speed": 1.7,
            "wind_speed_source": "climate",
            "safety_factor": 1.1,
            "safety_factor_source": "climate_policy",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="tank", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "tank"
        ]
        params, err = _build_tank_params(rows[0])

        assert err is None
        assert params is not None
        assert params["wind_speed"] == pytest.approx(1.7)
        assert params["wind_speed_source"] == "climate"
        assert params["safety_factor"] == pytest.approx(1.1)
        assert params["safety_factor_source"] == "climate_policy"

    def test_pipe_export_round_trips_canonical_underground_layers(self):
        from types import SimpleNamespace

        obj = SimpleNamespace(
            object_type="pipe",
            params={
                "name": "UG-1",
                "outer_diameter": 0.108,
                "pipe_length": 25.0,
                "wall_thickness": 0.004,
                "pipe_lambda": 45.0,
                "insulation_layers": [
                    {"thickness": 0.03, "material": "mineral_wool_boards_120"},
                    {"thickness": 0.02, "material": "polyurethane_products_50"},
                ],
                "process_temperature": 80.0,
                "placement": "underground",
                "pipe_centerline_depth": 1.4,
                "ground_temperature": 4.0,
                "ground_type": "clay",
                "ground_conductivity": 1.5,
                "num_local_elements": 2,
                "local_element_equiv_length": 1.2,
            },
        )

        content = build_objects_xlsx([obj])
        [(label, object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "pipe"
        ]
        params, err = _build_pipe_params(rows[0])

        assert label == "Трубопроводы"
        assert object_type == "pipe"
        assert err is None
        assert params is not None
        assert params["insulation_layers"] == obj.params["insulation_layers"]
        assert params["pipe_centerline_depth"] == 1.4
        assert params["ground_temperature"] == 4.0
        assert params["ground_temperature_source"] == "manual"
        assert params["ground_conductivity_source"] == "reference"
        assert params["num_local_elements"] == 2
        assert params["local_element_equiv_length"] == 1.2
        assert params["pipe_lambda"] == 45.0
        assert "pipe_material" not in params
        for forbidden in (
            "location",
            "burial_depth",
            "insulation_thickness",
            "insulation_material",
            "insulation_layer_count",
        ):
            assert forbidden not in params

    def test_pipe_export_omits_manual_alpha_and_round_trips_manual_insulation(self):
        from types import SimpleNamespace

        from app.services.project_object_params import prepare_project_object_params

        source = {
            "name": "Manual physics",
            "outer_diameter": 0.108,
            "pipe_length": 50.0,
            "wall_thickness": 0.004,
            "pipe_material": "carbon_steel",
            "insulation_layers": [
                {
                    "thickness": 0.05,
                    "material": "other",
                    "conductivity": 0.037,
                    "temperature_range": (-50.0, 250.0),
                }
            ],
            "ambient_temperature": -20.0,
            "process_temperature": 80.0,
            "min_switch_temperature": -20.0,
            "placement": "outdoor",
            "alpha_vnesh": 18.5,
            "wind_speed": 4.0,
            "insulation_temperature_basis": "outdoor_winter",
            "insulation_cover_material": "aluminum",
            "num_local_elements": 0,
        }
        content = build_objects_xlsx([SimpleNamespace(object_type="pipe", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "pipe"
        ]

        built, err = _build_pipe_params(rows[0])
        assert err is None
        assert built is not None
        prepared = prepare_project_object_params("pipe", built)

        assert "alpha_vnesh" not in prepared
        assert prepared["wind_speed"] == 4.0
        assert prepared["insulation_cover_material"] == "aluminum"
        assert prepared["insulation_layers"] == [
            {
                "thickness": 0.05,
                "material": "other",
                "conductivity": 0.037,
                "temperature_range": (-50.0, 250.0),
            }
        ]

    def test_tank_export_round_trips_canonical_underground_three_layers(self):
        from types import SimpleNamespace

        source = {
            "name": "UG tank",
            "shape": "rectangular",
            "length": 4.0,
            "width": 2.0,
            "height": 3.0,
            "wall_thickness": 0.01,
            "wall_lambda": 45.0,
            "insulation_layers": [
                {"thickness": 0.04, "material": "mineral_wool_boards_120"},
                {
                    "thickness": 0.03,
                    "material": "other",
                    "conductivity": 0.039,
                    "temperature_range": (-40.0, 180.0),
                },
                {"thickness": 0.02, "material": "polyurethane_products_50"},
            ],
            "ambient_temperature": -20.0,
            "ground_temperature": 5.0,
            "process_temperature": 80.0,
            "placement": "underground",
            "insulation_temperature_basis": "channel",
            "tank_buried_height": 1.25,
            "ground_type": "dry_sand",
            "ground_conductivity": 1.2,
            "wind_speed": 0.0,
            "safety_factor": 1.15,
            "q_additional": 0.0,
            "min_switch_temperature": -30.0,
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="tank", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "tank"
        ]
        params, err = _build_tank_params(rows[0])

        assert err is None
        assert params is not None
        assert params["insulation_layers"] == source["insulation_layers"]
        assert params["placement"] == "underground"
        assert params["ground_temperature"] == 5.0
        assert params["ground_temperature_source"] == "manual"
        assert params["ground_conductivity_source"] == "reference"
        assert params["tank_buried_height"] == pytest.approx(1.25)
        assert params["wall_thickness"] == pytest.approx(0.01)
        assert params["wall_lambda"] == 45.0
        assert params["q_additional"] == 0.0
        assert params["min_switch_temperature"] == -30.0
        for forbidden in (
            "location",
            "burial_depth",
            "insulation_thickness",
            "insulation_material",
            "insulation_layer_count",
        ):
            assert forbidden not in params


class TestExtendedRoundtripFields:
    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    @pytest.mark.parametrize(
        ("source_value", "expected_source"),
        [
            ("manual", "manual"),
            ("Вручную", "manual"),
            ("climate", "climate"),
            ("Климат", "climate"),
        ],
    )
    def test_parser_reads_ambient_temperature_source_aliases(
        self,
        builder,
        source_value: str,
        expected_source: str,
    ):
        params, err = builder(
            {
                "ambient_temperature": -7,
                "ambient_temperature_source": source_value,
            }
        )

        assert err is None
        assert params is not None
        assert params["ambient_temperature_source"] == expected_source

    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    def test_parser_rejects_unknown_ambient_temperature_source(self, builder):
        params, err = builder(
            {
                "ambient_temperature": -7,
                "ambient_temperature_source": "automatic",
            }
        )

        assert params is None
        assert err == "Не распознан источник температуры среды: automatic"

    def test_legacy_pipe_without_source_keeps_previous_climate_behavior(self):
        params, err = _build_pipe_params(
            {
                "outer_diameter_mm": 108,
                "ambient_temperature": -7,
                "climate_key": "Московская область|||Москва",
                "placement": "outdoor",
            }
        )

        assert err is None
        assert params is not None
        assert "ambient_temperature_source" not in params
        after_policy = apply_climate_policy("pipe", params)
        assert after_policy["ambient_temperature"] == pytest.approx(-23.0)
        assert after_policy["ambient_temperature_source"] == "climate"

    @pytest.mark.parametrize(
        ("source_field", "source_value", "expected_error"),
        [
            (
                "wind_speed_source",
                "climate",
                "Источник скорости ветра указан без скорости ветра",
            ),
            (
                "safety_factor_source",
                "manual",
                "Источник Kзап указан без Kзап",
            ),
        ],
    )
    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    def test_parser_rejects_source_without_its_value(
        self,
        builder,
        source_field: str,
        source_value: str,
        expected_error: str,
    ):
        params, err = builder({source_field: source_value})

        assert params is None
        assert err == expected_error

    def test_legacy_pipe_without_wind_or_safety_sources_keeps_manual_behavior(self):
        params, err = _build_pipe_params(
            {
                "wind_speed": 3.5,
                "safety_factor": 1.2,
                "placement": "outdoor",
            }
        )

        assert err is None
        assert params is not None
        assert "wind_speed_source" not in params
        assert "safety_factor_source" not in params

    @pytest.mark.parametrize(
        ("ground_type", "expected_source"),
        [(None, "manual"), ("custom", "manual"), ("clay", "reference")],
    )
    def test_pipe_parser_sets_ground_conductivity_provenance(
        self, ground_type: str | None, expected_source: str
    ):
        params, err = _build_pipe_params(
            {
                "placement": "underground",
                "ground_type": ground_type,
                "ground_conductivity": 1.5,
                "ground_temperature": 4,
            }
        )

        assert err is None
        assert params is not None
        assert params["ground_temperature_source"] == "manual"
        assert params["ground_conductivity_source"] == expected_source

    def test_pipe_parser_reads_climate_and_local_element_fields(self):
        params, err = _build_pipe_params(
            {
                "name": "pipe",
                "outer_diameter_mm": 108,
                "pipe_length": 50,
                "insulation_thickness_mm": 50,
                "insulation_material": "Минеральная вата",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "placement": "outdoor",
                "climate_region": "ХМАО",
                "climate_city": "Сургут",
                "climate_temperature_basis": "0,92",
                "safety_factor": 1.2,
                "min_switch_temperature": -35,
                "num_local_elements": 6,
                "local_element_equiv_length": 2.4,
            }
        )

        assert err is None
        assert params is not None
        assert params["climate_region"] == "ХМАО"
        assert params["climate_city"] == "Сургут"
        assert params["climate_temperature_basis"] == "t_0_92"
        assert params["min_switch_temperature"] == -35
        assert params["num_local_elements"] == 6
        assert params["local_element_equiv_length"] == 2.4

    def test_tank_parser_reads_climate_and_q_additional_fields(self):
        params, err = _build_tank_params(
            {
                "name": "tank",
                "shape": "Цилиндр",
                "diameter_mm": 2000,
                "height_mm": 3000,
                "insulation_thickness_mm": 80,
                "insulation_material": "Минеральная вата",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "climate_key": "ХМАО|||Сургут",
                "climate_temperature_basis": "Абс. мин.",
                "q_additional": 250,
            }
        )

        assert err is None
        assert params is not None
        assert params["climate_key"] == "ХМАО|||Сургут"
        assert params["climate_region"] == "ХМАО"
        assert params["climate_city"] == "Сургут"
        assert params["climate_temperature_basis"] == "t_abs_min"
        assert params["q_additional"] == 250


class TestToFloat:
    @pytest.mark.parametrize(
        "inp,expected",
        [
            (1, 1.0),
            (1.5, 1.5),
            ("2.5", 2.5),
            ("2,5", 2.5),  # русская запятая как десятичный
            ("  3.14  ", 3.14),
            (None, None),
            ("", None),
            ("abc", None),
        ],
    )
    def test_variants(self, inp, expected):
        assert _to_float(inp) == expected


class TestResolveMaterial:
    @pytest.mark.parametrize(
        "inp,expected",
        [
            ("mineral_wool_boards_120", "mineral_wool_boards_120"),
            ("polyurethane_products_50", "polyurethane_products_50"),
            ("k_flex_st", "k_flex_st"),
            ("другое", "other"),
            ("other", "other"),
        ],
    )
    def test_concrete_codes(self, inp, expected):
        assert _resolve_material(inp) == expected

    @pytest.mark.parametrize("inp", ["Минеральная вата", "мин. вата", "MINERAL_WOOL", "ППУ"])
    def test_generic_aliases_require_reselection(self, inp):
        resolution = _resolve_material_entry(inp)
        assert resolution.material is None
        assert resolution.needs_reselection is True

    @pytest.mark.parametrize("inp", ["", None, "некий материал"])
    def test_unknown_returns_none(self, inp):
        assert _resolve_material(inp) is None


class TestResolveShape:
    @pytest.mark.parametrize(
        "inp,expected",
        [
            ("Цилиндр", "cylindrical"),
            ("цилиндрический", "cylindrical"),
            ("cylindrical", "cylindrical"),
            ("Параллелепипед", "rectangular"),
            ("прямоугольный", "rectangular"),
        ],
    )
    def test_known(self, inp, expected):
        assert _resolve_shape(inp) == expected

    def test_unknown_returns_none(self):
        assert _resolve_shape("кубик") is None
        assert _resolve_shape("") is None
        assert _resolve_shape(None) is None


class TestBuildPipeParams:
    def test_full_row_ok(self):
        row = {
            "_row": 2,
            "name": "Т1",
            "outer_diameter_mm": 108,
            "pipe_length": 50,
            "insulation_thickness_mm": 50,
            "insulation_material": "mineral_wool_boards_120",
            "ambient_temperature": -20,
            "process_temperature": 80,
            "min_switch_temperature": -20,
        }
        params, err = _build_pipe_params(row)
        assert err is None
        # мм → м
        assert params["outer_diameter"] == pytest.approx(0.108)
        assert params["insulation_layers"] == [
            {"thickness": pytest.approx(0.05), "material": "mineral_wool_boards_120"}
        ]
        assert params["pipe_length"] == 50
        assert params["name"] == "Т1"
        for implicit_default in (
            "placement",
            "location",
            "wall_thickness",
            "pipe_material",
            "pipe_lambda",
            "wind_speed",
            "insulation_temperature_basis",
        ):
            assert implicit_default not in params

    def test_generic_material_is_preserved_as_reselection_request(self):
        row = {
            "_row": 2,
            "outer_diameter_mm": 108,
            "pipe_length": 50,
            "insulation_thickness_mm": 50,
            "insulation_material": "Минеральная вата",
            "ambient_temperature": -20,
            "process_temperature": 80,
        }
        params, err = _build_pipe_params(row)
        assert err is None
        assert params is not None
        assert params["insulation_layers"] == [{"thickness": pytest.approx(0.05)}]

    def test_missing_required_field_keeps_partial_params(self):
        row = {"_row": 5, "outer_diameter_mm": 108, "pipe_length": 50}
        params, err = _build_pipe_params(row)
        assert err is None
        assert params is not None
        assert params["outer_diameter"] == pytest.approx(0.108)
        assert params["pipe_length"] == 50
        assert "insulation_layers" not in params

    def test_unknown_pipe_material_is_rejected_explicitly(self):
        row = {
            "_row": 3,
            "outer_diameter_mm": 108,
            "pipe_length": 50,
            "insulation_thickness_mm": 50,
            "insulation_material": "mineral_wool_boards_120",
            "pipe_material": "какой-то-крутой",
            "ambient_temperature": -20,
            "process_temperature": 80,
        }
        params, err = _build_pipe_params(row)
        assert params is None
        assert err is not None
        assert "материал трубы" in err

    def test_unknown_placement_is_rejected_explicitly(self):
        params, err = _build_pipe_params({"_row": 2, "placement": "на Луне"})

        assert params is None
        assert err is not None
        assert "размещение" in err

    def test_pipe_material_and_lambda_are_mutually_exclusive(self):
        params, err = _build_pipe_params(
            {"_row": 2, "pipe_material": "carbon_steel", "pipe_lambda": 45}
        )

        assert params is None
        assert err is not None
        assert "один источник" in err

    def test_legacy_local_element_columns_are_not_imported(self):
        [(label, rows)] = _parse_csv(
            (
                "Тип;Задвижки;Фланцы;Опоры;Количество локальных элементов\n" "труба;1;2;3;4\n"
            ).encode()
        )

        assert label == "Трубопроводы (CSV)"
        assert rows == [{"_row": 2, "num_local_elements": "4"}]

    def test_name_stripped(self):
        row = {
            "_row": 2,
            "name": "  Спейс Т  ",
            "outer_diameter_mm": 108,
            "pipe_length": 50,
            "insulation_thickness_mm": 50,
            "insulation_material": "mineral_wool_boards_120",
            "ambient_temperature": -20,
            "process_temperature": 80,
        }
        params, _ = _build_pipe_params(row)
        assert params["name"] == "Спейс Т"

    def test_vapor_temperature_is_kept_in_common_params(self):
        row = {
            "_row": 2,
            "outer_diameter_mm": 108,
            "pipe_length": 50,
            "insulation_thickness_mm": 50,
            "insulation_material": "mineral_wool_boards_120",
            "insulation_temperature_basis": "outdoor_winter",
            "ambient_temperature": -20,
            "process_temperature": 80,
            "vapor_temperature": 140,
        }
        params, err = _build_pipe_params(row)
        assert err is None
        assert params["vapor_temperature"] == 140

    def test_russian_decimal_comma_works(self):
        """Пользователи часто пишут 108,5 вместо 108.5 — должно работать."""
        row = {
            "_row": 2,
            "outer_diameter_mm": "108,5",
            "pipe_length": "50,0",
            "insulation_thickness_mm": 50,
            "insulation_material": "mineral_wool_boards_120",
            "insulation_temperature_basis": "outdoor_winter",
            "ambient_temperature": -20,
            "process_temperature": 80,
        }
        params, err = _build_pipe_params(row)
        assert err is None
        assert params["outer_diameter"] == pytest.approx(0.1085)


class TestBuildTankParams:
    def test_cylindrical_full(self):
        row = {
            "_row": 2,
            "shape": "Цилиндр",
            "diameter_mm": 2000,
            "height_mm": 3000,
            "insulation_thickness_mm": 80,
            "insulation_material": "mineral_wool_boards_120",
            "ambient_temperature": -20,
            "process_temperature": 80,
        }
        params, err = _build_tank_params(row)
        assert err is None
        assert params["shape"] == "cylindrical"
        assert params["diameter"] == pytest.approx(2.0)
        assert params["height"] == pytest.approx(3.0)

    def test_rectangular_full(self):
        row = {
            "_row": 2,
            "shape": "Параллелепипед",
            "length_mm": 5000,
            "width_mm": 3000,
            "height_mm": 4000,
            "insulation_thickness_mm": 80,
            "insulation_material": "mineral_wool_boards_120",
            "ambient_temperature": -20,
            "process_temperature": 60,
        }
        params, err = _build_tank_params(row)
        assert err is None
        assert params["length"] == pytest.approx(5.0)
        assert params["width"] == pytest.approx(3.0)
        assert params["height"] == pytest.approx(4.0)

    def test_legacy_unsupported_shape_is_rejected(self):
        row = {
            "_row": 2,
            "shape": "Шар",
            "diameter_mm": 1500,
            "insulation_thickness_mm": 60,
            "insulation_material": "mineral_wool_boards_120",
            "ambient_temperature": -20,
            "process_temperature": 50,
        }
        params, err = _build_tank_params(row)
        assert params is None
        assert err is not None
        assert "форма" in err.lower()

    def test_missing_shape_keeps_partial_params(self):
        row = {"_row": 2, "insulation_thickness_mm": 60}
        params, err = _build_tank_params(row)
        assert err is None
        assert params is not None
        assert params["placement"] == "outdoor"
        assert params["insulation_layers"][0]["thickness"] == pytest.approx(0.06)
        assert "insulation_thickness" not in params

    def test_unknown_shape_rejected(self):
        row = {"_row": 2, "shape": "Куб", "insulation_thickness_mm": 60}
        params, err = _build_tank_params(row)
        assert params is None
        assert err is not None
        assert "форма" in err.lower()

    def test_cylindrical_missing_height_keeps_partial_params(self):
        row = {
            "_row": 2,
            "shape": "Цилиндр",
            "diameter_mm": 2000,
            "insulation_thickness_mm": 80,
            "insulation_material": "mineral_wool_boards_120",
            "ambient_temperature": -20,
            "process_temperature": 80,
        }
        params, err = _build_tank_params(row)
        assert err is None
        assert params is not None
        assert params["shape"] == "cylindrical"
        assert params["diameter"] == pytest.approx(2.0)
        assert "height" not in params

    def test_rectangular_missing_width_keeps_partial_params(self):
        row = {
            "_row": 2,
            "shape": "Параллелепипед",
            "length_mm": 5000,
            "height_mm": 4000,  # width пропущен
            "insulation_thickness_mm": 80,
            "insulation_material": "mineral_wool_boards_120",
            "ambient_temperature": -20,
            "process_temperature": 60,
        }
        params, err = _build_tank_params(row)
        assert err is None
        assert params is not None
        assert params["shape"] == "rectangular"
        assert params["length"] == pytest.approx(5.0)
        assert params["height"] == pytest.approx(4.0)
        assert "width" not in params

    def test_underground_rectangular_is_canonical_and_keeps_ground_provenance(self):
        params, err = _build_tank_params(
            {
                "_row": 2,
                "shape": "Параллелепипед",
                "length_mm": 5000,
                "width_mm": 3000,
                "height_mm": 4000,
                "wall_thickness_mm": 10,
                "wall_lambda": 45,
                "insulation_thickness_mm": 80,
                "insulation_material": "mineral_wool_boards_120",
                "ambient_temperature": -20,
                "ground_temperature": 5,
                "process_temperature": 80,
                "placement": "underground",
                "tank_buried_height": 1.5,
                "ground_type": "dry_sand",
                "ground_conductivity": 1.2,
                "wind_speed": 0,
            }
        )

        assert err is None
        assert params is not None
        assert params["placement"] == "underground"
        assert params["ambient_temperature"] == -20
        assert params["ground_temperature"] == 5
        assert params["ground_temperature_source"] == "manual"
        assert params["tank_buried_height"] == pytest.approx(1.5)
        assert params["ground_conductivity_source"] == "reference"
        assert params["wall_thickness"] == pytest.approx(0.01)
        assert params["wall_lambda"] == 45
        assert params["q_additional"] == 0.0
        assert params["insulation_layers"] == [
            {"thickness": pytest.approx(0.08), "material": "mineral_wool_boards_120"}
        ]
        for legacy_key in (
            "location",
            "burial_depth",
            "insulation_thickness",
            "insulation_material",
            "insulation_layer_count",
        ):
            assert legacy_key not in params


class TestParseCsv:
    def test_rejects_empty(self):
        with pytest.raises(ExcelImportError, match="пустой"):
            _parse_csv(b"")

    def test_rejects_missing_type_column(self):
        with pytest.raises(ExcelImportError, match="Тип"):
            _parse_csv(b"name;value\nfoo;1\n")

    def test_semicolon_delimiter(self):
        csv_data = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта\n"
            "труба;T1;108;50;50;Минеральная вата;-20;80\n"
        ).encode()
        sheets = _parse_csv(csv_data)
        assert len(sheets) == 1
        label, rows = sheets[0]
        assert "Трубопроводы" in label
        assert len(rows) == 1

    def test_blank_material_code_does_not_erase_material_alias(self):
        csv_data = (
            "Тип;Материал изоляции;Код материала изоляции;Размещение;Глубина заложения оси трубы, м;Грунт\n"
            "труба;other;;подземно;1.5;глина\n"
        ).encode()
        [(_label, rows)] = _parse_csv(csv_data)
        assert rows[0]["insulation_material"] == "other"
        assert rows[0]["pipe_centerline_depth"] == "1.5"
        assert rows[0]["ground_type"] == "глина"

    def test_comma_delimiter_autodetected(self):
        csv_data = (
            "Тип,Наименование,Форма,Диаметр, мм,Длина, мм,Ширина, мм,"
            "Высота, мм,Толщина изоляции, мм,Материал изоляции,T° среды,T° продукта\n"
        ).encode()
        # Только заголовок — результат пуст, но не падает
        sheets = _parse_csv(csv_data)
        assert sheets == []

    def test_cp1251_decoded(self):
        """Кодировка CP1251 — старый Excel, часто используется в РФ."""
        text = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта\n"
            "труба;T1;108;50;50;Минеральная вата;-20;80\n"
        )
        sheets = _parse_csv(text.encode("cp1251"))
        assert len(sheets) == 1

    def test_ignores_rows_without_type(self):
        csv_data = ("Тип;Наименование\n" "неизвестный_тип;foo\n" "труба;bar\n").encode()
        sheets = _parse_csv(csv_data)
        # Только «труба» попадает — раздел pipes
        # (но без других полей — распарсить в params нельзя, но это уже не дело _parse_csv)
        assert len(sheets) <= 1


class TestTemplateBuilders:
    def test_xlsx_has_both_sheets(self):
        import io

        from openpyxl import load_workbook

        data = build_template_xlsx()
        wb = load_workbook(io.BytesIO(data))
        assert "Трубопроводы" in wb.sheetnames
        assert "Резервуары" in wb.sheetnames
        # Есть хотя бы одна строка с примером данных
        assert wb["Трубопроводы"].max_row >= 2
        for sheet_name in ("Трубопроводы", "Резервуары"):
            headers = [cell.value for cell in wb[sheet_name][1]]
            assert "Мин. T° окр. среды" in headers
            assert "Макс. T° окр. среды" in headers
            assert "T° среды" not in headers
            assert "Источник T° среды" in headers
            assert "Источник скорости ветра" in headers
            assert "Источник Kзап" in headers

    def test_csv_starts_with_type_header(self):
        data = build_template_csv()
        text = data.decode("utf-8-sig")
        assert text.splitlines()[0].startswith("Тип")
        assert "T проп., °C" in text.splitlines()[0]
        assert "Мин. T° окр. среды" in text.splitlines()[0]
        assert "Макс. T° окр. среды" in text.splitlines()[0]
        assert ";T° среды;" not in text.splitlines()[0]


class TestAmbientBoundsImport:
    @pytest.mark.parametrize(
        "minimum_header",
        ["Мин. T° окр. среды", "T° среды", "Т° среды", "температура среды"],
    )
    @pytest.mark.parametrize(
        "maximum_header",
        ["Макс. T° окр. среды", "Макс T° окр. среды"],
    )
    def test_csv_accepts_canonical_and_legacy_ambient_headers(
        self,
        minimum_header: str,
        maximum_header: str,
    ):
        csv_data = (
            f"Тип;{minimum_header};{maximum_header};Размещение\n" "труба;-20;-5;outdoor\n"
        ).encode()

        [(_label, rows)] = _parse_csv(csv_data)
        params, err = _build_pipe_params(rows[0])

        assert err is None
        assert params is not None
        assert params["ambient_temperature"] == -20
        assert params["max_ambient_temperature"] == -5

    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    def test_blank_maximum_is_absent_and_zero_is_preserved(self, builder):
        blank, blank_error = builder(
            {
                "_row": 2,
                "ambient_temperature": -20,
                "max_ambient_temperature": None,
                "placement": "outdoor",
            }
        )
        zero, zero_error = builder(
            {
                "_row": 3,
                "ambient_temperature": -20,
                "max_ambient_temperature": 0,
                "placement": "outdoor",
            }
        )

        assert blank_error is None
        assert blank is not None
        assert "max_ambient_temperature" not in blank
        assert zero_error is None
        assert zero is not None
        assert zero["max_ambient_temperature"] == 0

    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    def test_negative_maximum_is_preserved(self, builder):
        params, err = builder(
            {
                "_row": 2,
                "ambient_temperature": -30,
                "max_ambient_temperature": -10,
                "placement": "outdoor",
            }
        )

        assert err is None
        assert params is not None
        assert params["max_ambient_temperature"] == -10

    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    def test_maximum_below_minimum_returns_field_aware_error(self, builder):
        params, err = builder(
            {
                "_row": 2,
                "ambient_temperature": -10,
                "max_ambient_temperature": -20,
                "placement": "outdoor",
            }
        )

        assert params is None
        assert err is not None
        assert "max_ambient_temperature" in err
        assert "ниже минимальной" in err

    def test_underground_pipe_ignores_inapplicable_ambient_bounds(self):
        params, err = _build_pipe_params(
            {
                "_row": 2,
                "ambient_temperature": -10,
                "max_ambient_temperature": -20,
                "placement": "underground",
            }
        )

        assert err is None
        assert params is not None
        assert "ambient_temperature" not in params
        assert "max_ambient_temperature" not in params


class TestAliasTables:
    """Консистентность таблиц алиасов."""

    def test_material_aliases_point_to_known_codes(self):
        """Все алиасы материалов должны указывать на коды из insulation.json."""
        from app.reference_data.loader import list_insulation_materials

        known = {m["material"] for m in list_insulation_materials()} | {"other"}
        for alias, code in {**GENERIC_MATERIAL_ALIASES, **SPECIAL_MATERIAL_ALIASES}.items():
            assert code in known, f"Алиас {alias!r} → {code!r} не найден в справочнике"

    def test_shape_aliases_point_to_known_shapes(self):
        for alias, code in SHAPE_ALIASES.items():
            assert code in {
                "cylindrical",
                "rectangular",
            }, f"Алиас формы {alias!r} → {code!r} неизвестен"


class TestAddRowsHelper:
    """Прямые unit-тесты на _add_rows (внутренний helper import flow)."""

    async def test_rejects_invalid_diameter_before_batch_state_and_dedupe(
        self, monkeypatch: pytest.MonkeyPatch
    ):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services.object_spreadsheet.persistence import _add_rows

        rows = [
            {
                "_row": 2,
                "name": "Невалидная",
                "outer_diameter_mm": 5000,
                "pipe_length": 50,
                "insulation_thickness_mm": 50,
                "insulation_material": "mineral_wool_boards_120",
                "insulation_temperature_basis": "outdoor_winter",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "min_switch_temperature": -20,
                "wall_thickness_mm": 4,
                "pipe_material": "carbon_steel",
                "placement": "outdoor",
                "wind_speed": 0,
            },
            {
                "_row": 3,
                "name": "Валидная",
                "outer_diameter_mm": 108,
                "pipe_length": 50,
                "insulation_thickness_mm": 50,
                "insulation_material": "mineral_wool_boards_120",
                "insulation_temperature_basis": "outdoor_winter",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "min_switch_temperature": -20,
                "wall_thickness_mm": 4,
                "pipe_material": "carbon_steel",
                "placement": "outdoor",
                "wind_speed": 0,
            },
        ]
        committed_rows: list[int] = []

        async def fake_commit(db, batch, sheet_label):
            committed_rows.extend(row["_row"] for _obj, row in batch)
            return len(batch), [uuid4() for _obj, _row in batch], []

        monkeypatch.setattr(persistence, "_commit_object_batch", fake_commit)
        dedupe_keys: set[str] = set()

        (
            created,
            next_sort,
            current_count,
            errors,
            object_ids,
            skipped,
            skipped_limit,
            invalid,
            validation_errors,
        ) = await _add_rows(
            AsyncMock(),
            uuid4(),
            "Трубопроводы",
            rows,
            "pipe",
            next_sort=7,
            current_count=3,
            dedupe_keys=dedupe_keys,
        )

        assert created == 1
        assert next_sort == 8
        assert current_count == 4
        assert len(object_ids) == 1
        assert errors == []
        assert skipped == 0
        assert skipped_limit == 0
        assert invalid == 1
        assert committed_rows == [3]
        assert len(dedupe_keys) == 1
        assert validation_errors == [
            {
                "sheet": "Трубопроводы",
                "row": 2,
                "field": "outer_diameter",
                "code": "OBJECT_PARAMS_INVALID",
                "message": "Наружный диаметр должен быть от 10,8 до 3000 мм",
            }
        ]

    async def test_commits_rows_in_batches(self, monkeypatch: pytest.MonkeyPatch):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services.object_spreadsheet.persistence import _add_rows

        rows = [
            {
                "_row": idx,
                "outer_diameter_mm": 108,
                "pipe_length": 50,
                "insulation_thickness_mm": 50,
                "insulation_material": "mineral_wool_boards_120",
                "insulation_temperature_basis": "outdoor_winter",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "min_switch_temperature": -20,
                "wall_thickness_mm": 4,
                "pipe_material": "carbon_steel",
                "placement": "outdoor",
                "wind_speed": 0,
            }
            for idx in range(2, 32)
        ]
        batch_sizes: list[int] = []
        stored_params: list[dict] = []

        async def fake_commit(db, batch, sheet_label):
            batch_sizes.append(len(batch))
            stored_params.extend(obj.params for obj, _row in batch)
            return len(batch), [uuid4() for _obj, _row in batch], []

        monkeypatch.setattr(persistence, "_commit_object_batch", fake_commit)
        db = AsyncMock()

        (
            created,
            next_sort,
            current_count,
            errors,
            object_ids,
            skipped,
            skipped_limit,
            invalid,
            validation_errors,
        ) = await _add_rows(
            db,
            uuid4(),
            "Pipes",
            rows,
            "pipe",
            next_sort=0,
            current_count=0,
        )

        assert created == 30
        assert next_sort == 30
        assert current_count == 30
        assert len(object_ids) == 30
        assert skipped == 0
        assert skipped_limit == 0
        assert invalid == 0
        assert validation_errors == []
        assert errors == []
        assert batch_sizes == [25, 5]
        assert all(params["insulation_layers"] for params in stored_params)
        assert all(params["placement"] == "outdoor" for params in stored_params)
        assert all(
            forbidden not in params
            for params in stored_params
            for forbidden in (
                "location",
                "burial_depth",
                "insulation_thickness",
                "insulation_material",
                "insulation_layer_count",
                "valve_count",
                "flange_count",
                "support_count",
            )
        )

    async def test_merge_mode_skips_existing_dedupe_key(self, monkeypatch: pytest.MonkeyPatch):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services.object_spreadsheet.persistence import _add_rows, _dedupe_key
        from app.services.project_object_params import prepare_project_object_params

        row = {
            "_row": 2,
            "name": " Line A ",
            "outer_diameter_mm": 108,
            "pipe_length": 50,
            "insulation_thickness_mm": 50,
            "insulation_material": "mineral_wool_boards_120",
            "insulation_temperature_basis": "outdoor_winter",
            "ambient_temperature": -20,
            "process_temperature": 80,
            "min_switch_temperature": -20,
            "wall_thickness_mm": 4,
            "pipe_material": "carbon_steel",
            "placement": "outdoor",
            "wind_speed": 0,
        }
        built, err = _build_pipe_params(row)
        assert err is None
        assert built is not None
        normalized = prepare_project_object_params("pipe", built)
        dedupe_keys = {_dedupe_key("pipe", normalized)}

        async def fake_commit(db, batch, sheet_label):
            return len(batch), ["should-not-create"] if batch else [], []

        monkeypatch.setattr(persistence, "_commit_object_batch", fake_commit)

        (
            created,
            next_sort,
            current_count,
            errors,
            object_ids,
            skipped,
            skipped_limit,
            invalid,
            validation_errors,
        ) = await _add_rows(
            AsyncMock(),
            uuid4(),
            "Pipes",
            [row],
            "pipe",
            next_sort=7,
            current_count=3,
            dedupe_keys=dedupe_keys,
        )

        assert created == 0
        assert next_sort == 7
        assert current_count == 3
        assert errors == []
        assert object_ids == []
        assert skipped == 1
        assert skipped_limit == 0
        assert invalid == 0
        assert validation_errors == []

    async def test_empty_batch_is_noop(self):
        from unittest.mock import AsyncMock

        created, object_ids, errors = await _commit_object_batch(AsyncMock(), [], "Pipes")

        assert created == 0
        assert object_ids == []
        assert errors == []

    async def test_commit_batch_success_adds_all_and_commits(self):
        from unittest.mock import AsyncMock, Mock
        from uuid import uuid4

        project_id = uuid4()
        objects = [
            ProjectObject(
                project_id=project_id,
                object_type="pipe",
                sort_order=idx,
                params={"name": f"P{idx}"},
            )
            for idx in range(2)
        ]
        db = AsyncMock()
        db.add_all = Mock()

        created, object_ids, errors = await _commit_object_batch(
            db,
            [(obj, {"_row": idx + 2}) for idx, obj in enumerate(objects)],
            "Pipes",
        )

        assert created == 2
        assert object_ids == [obj.id for obj in objects]
        assert errors == []
        db.add_all.assert_called_once_with(objects)
        db.flush.assert_awaited_once()
        db.commit.assert_awaited_once()

    async def test_commit_batch_falls_back_to_row_by_row_on_sql_error(
        self, monkeypatch: pytest.MonkeyPatch
    ):
        from unittest.mock import AsyncMock, Mock
        from uuid import uuid4

        async def fake_row_by_row(db, batch, sheet_label, batch_error):
            assert sheet_label == "Pipes"
            assert isinstance(batch_error, SQLAlchemyError)
            return 1, [uuid4()], [{"sheet": sheet_label, "row": 3, "message": "bad row"}]

        monkeypatch.setattr(persistence, "_commit_object_batch_row_by_row", fake_row_by_row)

        db = AsyncMock()
        db.add_all = Mock()
        db.flush.side_effect = SQLAlchemyError("batch failed")
        obj = ProjectObject(project_id=uuid4(), object_type="pipe", sort_order=0, params={})

        created, object_ids, errors = await _commit_object_batch(db, [(obj, {"_row": 2})], "Pipes")

        assert created == 1
        assert len(object_ids) == 1
        assert errors == [{"sheet": "Pipes", "row": 3, "message": "bad row"}]
        db.rollback.assert_awaited_once()

    async def test_row_by_row_fallback_keeps_successes_and_reports_failures(self):
        from unittest.mock import AsyncMock, Mock
        from uuid import uuid4

        db = AsyncMock()
        db.add = Mock()
        db.flush.side_effect = [None, RuntimeError("second failed")]
        batch = [
            (
                ProjectObject(
                    project_id=uuid4(),
                    object_type="pipe",
                    sort_order=0,
                    params={"name": "ok"},
                ),
                {"_row": 2},
            ),
            (
                ProjectObject(
                    project_id=uuid4(),
                    object_type="pipe",
                    sort_order=1,
                    params={"name": "bad"},
                ),
                {"_row": 3},
            ),
        ]

        created, object_ids, errors = await _commit_object_batch_row_by_row(
            db,
            batch,
            "Pipes",
            SQLAlchemyError("batch failed"),
        )

        assert created == 1
        assert len(object_ids) == 1
        assert errors == [{"sheet": "Pipes", "row": 3, "message": "RuntimeError: second failed"}]
        assert db.add.call_count == 2
        assert db.commit.await_count == 1
        assert db.rollback.await_count == 1

    async def test_import_csv_orchestrates_access_state_and_batches(
        self, monkeypatch: pytest.MonkeyPatch
    ):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services import excel_import_service as mod

        project_id = uuid4()
        first_id = uuid4()
        second_id = uuid4()
        calls: list[tuple[str, int, int]] = []

        async def fake_access(db, checked_project_id, principal):
            assert checked_project_id == project_id

        async def fake_state(db, checked_project_id):
            assert checked_project_id == project_id
            return 3, 4

        async def fake_dedupe_keys(db, checked_project_id):
            assert checked_project_id == project_id
            return set()

        async def fake_add_rows(
            db,
            checked_project_id,
            sheet_label,
            rows,
            object_type,
            next_sort,
            current_count,
            dedupe_keys=None,
            prepared_rows=None,
        ):
            calls.append((object_type, next_sort, current_count))
            object_id = first_id if object_type == "pipe" else second_id
            return 1, next_sort + 1, current_count + 1, [], [object_id], 0, 0, 0, []

        monkeypatch.setattr(mod, "_ensure_import_access", fake_access)
        monkeypatch.setattr(mod, "_project_import_state", fake_state)
        monkeypatch.setattr(mod, "_existing_dedupe_keys", fake_dedupe_keys)
        monkeypatch.setattr(
            mod,
            "_parse_csv",
            lambda content: [
                ("Трубопроводы (CSV)", [{"_row": 2}]),
                ("Резервуары (CSV)", [{"_row": 3}]),
            ],
        )
        monkeypatch.setattr(mod, "_add_rows", fake_add_rows)

        result = await import_objects_from_csv(AsyncMock(), project_id, object(), b"csv")

        assert result == {
            "created": 2,
            "skipped_duplicates": 0,
            "skipped_limit": 0,
            "invalid": 0,
            "mode": "merge",
            "errors": [],
            "validation_errors": [],
            "created_object_ids": [first_id, second_id],
        }
        assert calls == [("pipe", 4, 3), ("tank", 5, 4)]

    async def test_import_csv_rejects_file_without_recognized_rows(
        self, monkeypatch: pytest.MonkeyPatch
    ):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services import excel_import_service as mod

        async def fake_access(db, checked_project_id, principal):
            return None

        async def fake_state(db, checked_project_id):
            return 0, 0

        monkeypatch.setattr(mod, "_ensure_import_access", fake_access)
        monkeypatch.setattr(mod, "_project_import_state", fake_state)
        monkeypatch.setattr(mod, "_parse_csv", lambda content: [])

        with pytest.raises(ExcelImportError, match="распознанным типом"):
            await import_objects_from_csv(AsyncMock(), uuid4(), object(), b"csv")

    async def test_import_excel_orchestrates_known_sheets(self, monkeypatch: pytest.MonkeyPatch):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services import excel_import_service as mod
        from app.services.object_spreadsheet import parsing

        class FakeWorkbook:
            sheetnames = ["Трубопроводы", "Резервуары", "Неизвестный"]

            def __getitem__(self, name):
                return f"worksheet:{name}"

        project_id = uuid4()
        pipe_id = uuid4()
        tank_id = uuid4()
        read_sheets: list[str] = []
        add_calls: list[tuple[str, int, int]] = []

        async def fake_access(db, checked_project_id, principal):
            assert checked_project_id == project_id

        async def fake_state(db, checked_project_id):
            return 0, 0

        async def fake_dedupe_keys(db, checked_project_id):
            assert checked_project_id == project_id
            return set()

        def fake_read_sheet(worksheet, headers):
            read_sheets.append(worksheet)
            return [{"_row": len(read_sheets) + 1}]

        async def fake_add_rows(
            db,
            checked_project_id,
            sheet_label,
            rows,
            object_type,
            next_sort,
            current_count,
            dedupe_keys=None,
            prepared_rows=None,
        ):
            add_calls.append((object_type, next_sort, current_count))
            object_id = pipe_id if object_type == "pipe" else tank_id
            return 1, next_sort + 1, current_count + 1, [], [object_id], 0, 0, 0, []

        monkeypatch.setattr(mod, "_validate_xlsx_archive", lambda content: None)
        monkeypatch.setattr(parsing, "load_workbook", lambda *args, **kwargs: FakeWorkbook())
        monkeypatch.setattr(mod, "_ensure_import_access", fake_access)
        monkeypatch.setattr(mod, "_project_import_state", fake_state)
        monkeypatch.setattr(mod, "_existing_dedupe_keys", fake_dedupe_keys)
        monkeypatch.setattr(parsing, "_read_sheet", fake_read_sheet)
        monkeypatch.setattr(mod, "_add_rows", fake_add_rows)

        result = await import_objects_from_excel(AsyncMock(), project_id, object(), b"xlsx")

        assert result == {
            "created": 2,
            "skipped_duplicates": 0,
            "skipped_limit": 0,
            "invalid": 0,
            "mode": "merge",
            "errors": [],
            "validation_errors": [],
            "created_object_ids": [pipe_id, tank_id],
        }
        assert read_sheets == ["worksheet:Трубопроводы", "worksheet:Резервуары"]
        assert add_calls == [("pipe", 0, 0), ("tank", 1, 1)]

    async def test_import_excel_rejects_workbook_without_known_sheets(
        self, monkeypatch: pytest.MonkeyPatch
    ):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services import excel_import_service as mod
        from app.services.object_spreadsheet import parsing

        class FakeWorkbook:
            sheetnames = ["Лист1"]

        async def fake_access(db, checked_project_id, principal):
            return None

        async def fake_state(db, checked_project_id):
            return 0, 0

        monkeypatch.setattr(mod, "_validate_xlsx_archive", lambda content: None)
        monkeypatch.setattr(parsing, "load_workbook", lambda *args, **kwargs: FakeWorkbook())
        monkeypatch.setattr(mod, "_ensure_import_access", fake_access)
        monkeypatch.setattr(mod, "_project_import_state", fake_state)

        with pytest.raises(ExcelImportError, match="не найдены листы"):
            await import_objects_from_excel(AsyncMock(), uuid4(), object(), b"xlsx")

    async def test_skip_structural_row_error_continues_to_next(
        self, monkeypatch: pytest.MonkeyPatch
    ):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services.object_spreadsheet.persistence import _add_rows

        rows = [
            {"_row": 2, "shape": "куб"},
            {
                "_row": 3,
                "shape": "Цилиндр",
                "diameter_mm": 2000,
                "height_mm": 3000,
                "insulation_thickness_mm": 50,
                "insulation_material": "mineral_wool_boards_120",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "min_switch_temperature": -20,
                "heating_height": 3,
                "laying_step": 0.2,
                "wind_speed": 0,
            },
        ]

        async def fake_commit(db, batch, sheet_label):
            return len(batch), ["oid"], []

        monkeypatch.setattr(persistence, "_commit_object_batch", fake_commit)
        db = AsyncMock()
        (
            created,
            next_sort,
            current_count,
            errors,
            object_ids,
            skipped,
            skipped_limit,
            invalid,
            validation_errors,
        ) = await _add_rows(
            db,
            uuid4(),
            "Tanks",
            rows,
            "tank",
            next_sort=0,
            current_count=0,
        )
        assert created == 1
        assert len(errors) == 1
        assert next_sort == 1
        assert current_count == 1
        assert object_ids == ["oid"]
        assert skipped == 0
        assert skipped_limit == 0
        assert invalid == 0
        assert validation_errors == []

    async def test_project_limit_breaks_loop(self, monkeypatch: pytest.MonkeyPatch):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services.object_spreadsheet.persistence import _add_rows

        rows = [
            {
                "_row": 2,
                "outer_diameter_mm": 108,
                "pipe_length": 50,
                "insulation_thickness_mm": 50,
                "insulation_material": "mineral_wool_boards_120",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "min_switch_temperature": -20,
                "wall_thickness_mm": 4,
                "pipe_material": "carbon_steel",
                "placement": "outdoor",
                "wind_speed": 0,
                "insulation_temperature_basis": "outdoor_winter",
            },
            {
                "_row": 3,
                "outer_diameter_mm": 108,
                "pipe_length": 60,
                "insulation_thickness_mm": 50,
                "insulation_material": "mineral_wool_boards_120",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "min_switch_temperature": -20,
                "wall_thickness_mm": 4,
                "pipe_material": "carbon_steel",
                "placement": "outdoor",
                "wind_speed": 0,
                "insulation_temperature_basis": "outdoor_winter",
            },
        ]

        monkeypatch.setattr(settings, "GUEST_MAX_OBJECTS_PER_PROJECT", 0)
        db = AsyncMock()
        (
            created,
            _,
            current_count,
            errors,
            object_ids,
            skipped,
            skipped_limit,
            invalid,
            validation_errors,
        ) = await _add_rows(
            db,
            uuid4(),
            "Pipes",
            rows,
            "pipe",
            next_sort=0,
            current_count=0,
        )
        assert created == 0
        assert current_count == 0
        assert len(errors) == 1
        assert "Пропущено строк: 2" in errors[0]["message"]
        assert object_ids == []
        assert skipped == 0
        assert skipped_limit == 2
        assert invalid == 0
        assert validation_errors == []

    async def test_unexpected_exception_logged_to_errors(self, monkeypatch: pytest.MonkeyPatch):
        from unittest.mock import AsyncMock
        from uuid import uuid4

        from app.services.object_spreadsheet.persistence import _add_rows

        rows = [
            {
                "_row": 2,
                "outer_diameter_mm": 108,
                "pipe_length": 50,
                "insulation_thickness_mm": 50,
                "insulation_material": "mineral_wool_boards_120",
                "ambient_temperature": -20,
                "process_temperature": 80,
            },
        ]

        def boom_prepare(*args, **kwargs):
            raise RuntimeError("bad prepare")

        monkeypatch.setattr(
            preparation,
            "validate_and_canonicalize_project_object_params",
            boom_prepare,
        )
        db = AsyncMock()
        (
            created,
            _,
            current_count,
            errors,
            object_ids,
            skipped,
            skipped_limit,
            invalid,
            validation_errors,
        ) = await _add_rows(
            db,
            uuid4(),
            "X",
            rows,
            "pipe",
            next_sort=0,
            current_count=0,
        )
        assert created == 0
        assert current_count == 0
        assert "RuntimeError" in errors[0]["message"]
        assert object_ids == []
        assert skipped == 0
        assert skipped_limit == 0
        assert invalid == 0
        assert validation_errors == []


class TestParseCsvAdvanced:
    def test_only_header_no_data(self):
        csv_data = "Тип;Наименование\n".encode()
        assert _parse_csv(csv_data) == []

    def test_unknown_type_value_skipped(self):
        csv_data = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта\n"
            "странный;X;108;50;50;Минеральная вата;-20;80\n"
        ).encode()
        assert _parse_csv(csv_data) == []

    def test_mixed_pipe_and_tank_in_same_csv(self):
        csv_data = (
            "Тип;Наименование;Диаметр, мм;Длина, м;Толщина изоляции, мм;"
            "Материал изоляции;T° среды;T° продукта;Форма;Длина, мм;Ширина, мм;Высота, мм\n"
            "труба;P1;108;50;50;Минеральная вата;-20;80;;;;\n"
            "резервуар;T1;;;80;Минеральная вата;-20;60;Параллелепипед;5000;3000;4000\n"
        ).encode()
        result = _parse_csv(csv_data)
        labels = [label for label, _ in result]
        assert "Трубопроводы (CSV)" in labels
        assert "Резервуары (CSV)" in labels
