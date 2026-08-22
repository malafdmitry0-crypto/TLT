"""Owner-focused object spreadsheet tests.

Методология: table-driven testing. Эти функции получают произвольные строки
из Excel (локалями, запятыми, unicode) и должны корректно превращать их в
питоновские типы. Любая ошибка здесь = неправильный расчёт.
"""

import io

import pytest

from app.services.heat_loss_application import apply_climate_policy
from app.services.object_spreadsheet.export import build_objects_xlsx
from app.services.object_spreadsheet.mapping import (
    _mark_edited_climate_temperature_as_manual,
)
from app.services.object_spreadsheet.parsing import (
    _parse_excel_workbook,
)
from app.services.object_spreadsheet.pipe_mapping import _build_pipe_params
from app.services.object_spreadsheet.tank_mapping import _build_tank_params
from app.services.object_spreadsheet.templates import build_template_csv, build_template_xlsx


class TestBuildObjectsXlsxSafety:
    @pytest.mark.parametrize(
        ("object_type", "placement", "minimum", "maximum", "expected_maximum"),
        [
            ("pipe", "outdoor", -20.0, 0.0, 0.0),
            ("pipe", "indoor", -20.0, -5.0, -5.0),
            ("pipe", "underground", -20.0, 15.0, None),
            ("tank", "outdoor", -20.0, 0.0, 0.0),
            ("tank", "underground", -20.0, 15.0, 15.0),
        ],
    )
    def test_objects_xlsx_roundtrips_distinct_ambient_bounds(
        self,
        object_type: str,
        placement: str,
        minimum: float,
        maximum: float,
        expected_maximum: float | None,
    ):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        source = {
            "name": "Ambient bounds",
            "shape": "cylindrical",
            "ambient_temperature": minimum,
            "max_ambient_temperature": maximum,
            "ambient_temperature_source": "manual",
            "placement": placement,
        }
        content = build_objects_xlsx([SimpleNamespace(object_type=object_type, params=source)])
        sheet_name = "Трубопроводы" if object_type == "pipe" else "Резервуары"
        worksheet = load_workbook(io.BytesIO(content), data_only=True)[sheet_name]
        exported_row = dict(
            zip(
                [cell.value for cell in worksheet[1]],
                [cell.value for cell in worksheet[2]],
                strict=True,
            )
        )
        if placement == "underground" and object_type == "pipe":
            assert exported_row["Мин. T° окр. среды"] is None
            assert exported_row["Макс. T° окр. среды"] is None
            assert exported_row["Источник T° среды"] is None
        else:
            assert exported_row["Мин. T° окр. среды"] == minimum
            assert exported_row["Макс. T° окр. среды"] == maximum
            assert exported_row["Источник T° среды"] == "manual"
        [(label, parsed_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == object_type
        ]
        builder = _build_pipe_params if object_type == "pipe" else _build_tank_params
        params, err = builder(rows[0])

        assert label in {"Трубопроводы", "Резервуары"}
        assert parsed_type == object_type
        assert err is None
        assert params is not None
        if placement == "underground" and object_type == "pipe":
            assert "ambient_temperature" not in params
        else:
            assert params["ambient_temperature"] == minimum
        if expected_maximum is None:
            assert "max_ambient_temperature" not in params
        else:
            assert params["max_ambient_temperature"] == expected_maximum

    @pytest.mark.parametrize("object_type", ["pipe", "tank"])
    def test_objects_xlsx_keeps_blank_ambient_maximum_absent(self, object_type: str):
        from types import SimpleNamespace

        content = build_objects_xlsx(
            [
                SimpleNamespace(
                    object_type=object_type,
                    params={
                        "shape": "cylindrical",
                        "ambient_temperature": -20.0,
                        "placement": "outdoor",
                    },
                )
            ]
        )
        [(_label, _parsed_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == object_type
        ]
        builder = _build_pipe_params if object_type == "pipe" else _build_tank_params
        params, err = builder(rows[0])

        assert err is None
        assert params is not None
        assert "max_ambient_temperature" not in params

    def test_dangerous_object_strings_are_not_exported_as_formulas(self):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        obj = SimpleNamespace(
            object_type="pipe",
            params={
                "name": '=HYPERLINK("http://example.test","x")',
                "outer_diameter": 0.108,
                "pipe_length": "+SUM(1,1)",
                "wall_thickness": 0.004,
                "pipe_material": "carbon_steel",
                "insulation_layers": [{"thickness": 0.05, "material": "+SUM(1,1)"}],
                "ambient_temperature": "-2+3",
                "process_temperature": "=1+1",
                "vapor_temperature": "@cmd",
            },
        )

        wb = load_workbook(io.BytesIO(build_objects_xlsx([obj])), data_only=False)
        ws = wb["Трубопроводы"]
        cells_by_header = {
            header.value: ws.cell(row=2, column=index)
            for index, header in enumerate(ws[1], start=1)
        }
        for header in (
            "Наименование",
            "Длина, м",
            "Код материала изоляции",
            "Мин. T° окр. среды",
            "T° продукта",
            "T проп., °C",
        ):
            cell = cells_by_header[header]
            assert cell.data_type == "s"
            assert str(cell.value).startswith("'")

    def test_objects_xlsx_does_not_crash_on_malicious_dimension_strings(self):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        obj = SimpleNamespace(
            object_type="tank",
            params={
                "name": "tank",
                "shape": "rectangular",
                "diameter": "=1+1",
                "length": "+SUM(1,1)",
                "width": "-2+3",
                "height": "@cmd",
                "insulation_layers": [{"thickness": "=10", "material": "mineral_wool_boards_120"}],
                "ambient_temperature": -20,
                "process_temperature": 80,
            },
        )

        wb = load_workbook(io.BytesIO(build_objects_xlsx([obj])), data_only=False)
        ws = wb["Резервуары"]
        for cell_ref in ("C2", "D2", "E2", "F2", "G2"):
            cell = ws[cell_ref]
            assert cell.data_type == "s"
            assert str(cell.value).startswith("'")

    def test_objects_xlsx_exports_extended_roundtrip_fields(self):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        obj = SimpleNamespace(
            object_type="pipe",
            params={
                "name": "pipe",
                "outer_diameter": 0.108,
                "pipe_length": 50,
                "wall_thickness": 0.004,
                "pipe_material": "carbon_steel",
                "insulation_layers": [{"thickness": 0.05, "material": "mineral_wool_boards_120"}],
                "ambient_temperature": -20,
                "process_temperature": 80,
                "placement": "outdoor",
                "wind_speed": 3.5,
                "climate_region": "ХМАО",
                "climate_city": "Сургут",
                "climate_key": "ХМАО|||Сургут",
                "climate_temperature_basis": "t_0_92",
                "safety_factor": 1.2,
                "min_switch_temperature": -35,
                "num_local_elements": 6,
                "local_element_equiv_length": 2.4,
            },
        )

        wb = load_workbook(io.BytesIO(build_objects_xlsx([obj])), data_only=True)
        ws = wb["Трубопроводы"]
        headers = [cell.value for cell in ws[1]]
        row = dict(zip(headers, [cell.value for cell in ws[2]], strict=True))

        assert row["Ключ климата"] == "ХМАО|||Сургут"
        assert row["Обеспеченность климата"] == "t_0_92"
        assert row["Мин. T включения, °C"] == -35
        assert row["Количество локальных элементов"] == 6
        assert row["L экв., м"] == 2.4
        assert all(
            forbidden not in header.casefold()
            for header in headers
            for forbidden in ("location", "burial_depth", "количество слоёв")
        )

    @pytest.mark.parametrize(
        ("ambient_source", "expected_temperature"),
        [("manual", -7.0), ("climate", -33.0)],
    )
    def test_pipe_export_roundtrip_preserves_ambient_temperature_source(
        self,
        ambient_source: str,
        expected_temperature: float,
    ):
        from types import SimpleNamespace

        source = {
            "name": "Pipe climate provenance",
            "outer_diameter": 0.108,
            "pipe_length": 50.0,
            "ambient_temperature": -33.0 if ambient_source == "climate" else -7.0,
            "ambient_temperature_source": ambient_source,
            "process_temperature": 80.0,
            "placement": "outdoor",
            "climate_key": "Алтайский край|||Тогул",
            "climate_region": "Алтайский край",
            "climate_city": "Тогул",
            "climate_temperature_basis": "t_0_92",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="pipe", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "pipe"
        ]
        params, err = _build_pipe_params(rows[0])

        assert err is None
        assert params is not None
        assert params["ambient_temperature_source"] == ambient_source
        after_policy = apply_climate_policy("pipe", params)
        assert after_policy["ambient_temperature"] == pytest.approx(expected_temperature)
        assert after_policy["ambient_temperature_source"] == ambient_source

    @pytest.mark.parametrize("object_type", ["pipe", "tank"])
    @pytest.mark.parametrize("edited_temperature", [-10.0, 0.0])
    def test_exported_climate_temperature_edit_becomes_manual_override(
        self, object_type: str, edited_temperature: float
    ):
        from types import SimpleNamespace

        from openpyxl import load_workbook

        source = {
            "name": "Edited climate temperature",
            "outer_diameter": 0.108,
            "pipe_length": 50.0,
            "shape": "cylindrical",
            "diameter": 2.0,
            "height": 3.0,
            "ambient_temperature": -33.0,
            "ambient_temperature_source": "climate",
            "process_temperature": 80.0,
            "placement": "outdoor",
            "climate_key": "Алтайский край|||Тогул",
            "climate_region": "Алтайский край",
            "climate_city": "Тогул",
            "climate_temperature_basis": "t_0_92",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type=object_type, params=source)])
        workbook = load_workbook(io.BytesIO(content))
        sheet_name = "Трубопроводы" if object_type == "pipe" else "Резервуары"
        worksheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        worksheet.cell(row=2, column=headers["Мин. T° окр. среды"], value=edited_temperature)
        edited = io.BytesIO()
        workbook.save(edited)

        [(_label, _parsed_type, rows)] = [
            parsed
            for parsed in _parse_excel_workbook(edited.getvalue())
            if parsed[1] == object_type
        ]
        builder = _build_pipe_params if object_type == "pipe" else _build_tank_params
        params, err = builder(rows[0])

        assert err is None
        assert params is not None
        assert params["ambient_temperature"] == edited_temperature
        assert params["ambient_temperature_source"] == "manual"
        after_policy = apply_climate_policy(object_type, params)
        assert after_policy["ambient_temperature"] == edited_temperature
        assert after_policy["ambient_temperature_source"] == "manual"

    def test_underground_pipe_ignores_ambient_temperature_even_with_climate_source(self):
        params, err = _build_pipe_params(
            {
                "outer_diameter_mm": 108,
                "pipe_length": 50,
                "ambient_temperature": -33,
                "ambient_temperature_source": "climate",
                "placement": "underground",
                "pipe_centerline_depth": 1.0,
                "climate_key": "Алтайский край|||Тогул",
                "climate_region": "Алтайский край",
                "climate_city": "Тогул",
                "climate_temperature_basis": "t_0_92",
            }
        )

        assert err is None
        assert params is not None
        assert "ambient_temperature" not in params
        assert "ambient_temperature_source" not in params
        after_policy = apply_climate_policy("pipe", params)
        assert "ambient_temperature" not in after_policy
        assert "ambient_temperature_source" not in after_policy

    @pytest.mark.parametrize(
        ("source", "expected_source"),
        [(None, None), ("manual", "manual"), ("climate", "climate")],
    )
    def test_climate_temperature_provenance_does_not_guess_without_known_reference(
        self, source: str | None, expected_source: str | None
    ):
        params = {
            "outer_diameter": 0.108,
            "placement": "outdoor",
            "ambient_temperature": -10.0,
            "ambient_temperature_source": source,
        }
        if source == "climate":
            params["climate_key"] = "unknown|||Unknown"

        _mark_edited_climate_temperature_as_manual("pipe", params)

        assert params["ambient_temperature_source"] == expected_source

    def test_tank_export_roundtrip_preserves_manual_ambient_temperature(self):
        from types import SimpleNamespace

        source = {
            "name": "Tank climate provenance",
            "shape": "cylindrical",
            "diameter": 2.0,
            "height": 3.0,
            "ambient_temperature": -7.0,
            "ambient_temperature_source": "manual",
            "process_temperature": 80.0,
            "placement": "outdoor",
            "climate_key": "Москва|||Москва",
            "climate_region": "Москва",
            "climate_city": "Москва",
            "climate_temperature_basis": "t_0_92",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="tank", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "tank"
        ]
        params, err = _build_tank_params(rows[0])

        assert err is None
        assert params is not None
        assert params["ambient_temperature_source"] == "manual"
        after_policy = apply_climate_policy("tank", params)
        assert after_policy["ambient_temperature"] == pytest.approx(-7.0)
        assert after_policy["ambient_temperature_source"] == "manual"

    def test_pipe_export_roundtrip_preserves_climate_wind_and_safety_sources(self):
        from types import SimpleNamespace

        source = {
            "name": "Pipe climate provenance",
            "outer_diameter": 0.108,
            "pipe_length": 50.0,
            "process_temperature": 80.0,
            "placement": "outdoor",
            "climate_key": "Алтайский край|||Славгород",
            "climate_region": "Алтайский край",
            "climate_city": "Славгород",
            "climate_temperature_basis": "t_0_92",
            "wind_speed": 4.0,
            "wind_speed_source": "climate",
            "safety_factor": 1.1,
            "safety_factor_source": "climate_policy",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="pipe", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "pipe"
        ]
        params, err = _build_pipe_params(rows[0])

        assert err is None
        assert params is not None
        assert params["wind_speed_source"] == "climate"
        assert params["safety_factor_source"] == "climate_policy"

        params.update(
            climate_key="Алтайский край|||Солонешное",
            climate_region="Алтайский край",
            climate_city="Солонешное",
        )
        after_policy = apply_climate_policy("pipe", params)
        assert after_policy["wind_speed"] == pytest.approx(1.2)
        assert after_policy["wind_speed_source"] == "climate"
        assert after_policy["safety_factor"] == pytest.approx(1.1)
        assert after_policy["safety_factor_source"] == "climate_policy"

    def test_tank_export_roundtrip_preserves_wind_and_safety_sources(self):
        from types import SimpleNamespace

        source = {
            "name": "Tank climate provenance",
            "shape": "cylindrical",
            "diameter": 2.0,
            "height": 3.0,
            "process_temperature": 80.0,
            "placement": "outdoor",
            "wind_speed": 1.7,
            "wind_speed_source": "climate",
            "safety_factor": 1.1,
            "safety_factor_source": "climate_policy",
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="tank", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "tank"
        ]
        params, err = _build_tank_params(rows[0])

        assert err is None
        assert params is not None
        assert params["wind_speed"] == pytest.approx(1.7)
        assert params["wind_speed_source"] == "climate"
        assert params["safety_factor"] == pytest.approx(1.1)
        assert params["safety_factor_source"] == "climate_policy"

    def test_pipe_export_round_trips_canonical_underground_layers(self):
        from types import SimpleNamespace

        obj = SimpleNamespace(
            object_type="pipe",
            params={
                "name": "UG-1",
                "outer_diameter": 0.108,
                "pipe_length": 25.0,
                "wall_thickness": 0.004,
                "pipe_lambda": 45.0,
                "insulation_layers": [
                    {"thickness": 0.03, "material": "mineral_wool_boards_120"},
                    {"thickness": 0.02, "material": "polyurethane_products_50"},
                ],
                "process_temperature": 80.0,
                "placement": "underground",
                "pipe_centerline_depth": 1.4,
                "ground_temperature": 4.0,
                "ground_type": "clay",
                "ground_conductivity": 1.5,
                "num_local_elements": 2,
                "local_element_equiv_length": 1.2,
            },
        )

        content = build_objects_xlsx([obj])
        [(label, object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "pipe"
        ]
        params, err = _build_pipe_params(rows[0])

        assert label == "Трубопроводы"
        assert object_type == "pipe"
        assert err is None
        assert params is not None
        assert params["insulation_layers"] == obj.params["insulation_layers"]
        assert params["pipe_centerline_depth"] == 1.4
        assert params["ground_temperature"] == 4.0
        assert params["ground_temperature_source"] == "manual"
        assert params["ground_conductivity_source"] == "reference"
        assert params["num_local_elements"] == 2
        assert params["local_element_equiv_length"] == 1.2
        assert params["pipe_lambda"] == 45.0
        assert "pipe_material" not in params
        for forbidden in (
            "location",
            "burial_depth",
            "insulation_thickness",
            "insulation_material",
            "insulation_layer_count",
        ):
            assert forbidden not in params

    def test_pipe_export_omits_manual_alpha_and_round_trips_manual_insulation(self):
        from types import SimpleNamespace

        from app.services.project_object_params import (
            normalize_project_object_params,
            validate_and_canonicalize_project_object_params,
        )

        source = {
            "name": "Manual physics",
            "outer_diameter": 0.108,
            "pipe_length": 50.0,
            "wall_thickness": 0.004,
            "pipe_material": "carbon_steel",
            "insulation_layers": [
                {
                    "thickness": 0.05,
                    "material": "other",
                    "conductivity": 0.037,
                    "temperature_range": (-50.0, 250.0),
                }
            ],
            "ambient_temperature": -20.0,
            "process_temperature": 80.0,
            "min_switch_temperature": -20.0,
            "placement": "outdoor",
            "alpha_vnesh": 18.5,
            "wind_speed": 4.0,
            "insulation_temperature_basis": "outdoor_winter",
            "insulation_cover_material": "aluminum",
            "num_local_elements": 0,
        }
        content = build_objects_xlsx([SimpleNamespace(object_type="pipe", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "pipe"
        ]

        built, err = _build_pipe_params(rows[0])
        assert err is None
        assert built is not None
        normalized = normalize_project_object_params("pipe", built)
        preparation = validate_and_canonicalize_project_object_params("pipe", normalized)
        assert preparation.report.is_valid
        prepared = preparation.params

        assert "alpha_vnesh" not in prepared
        assert prepared["wind_speed"] == 4.0
        assert prepared["insulation_cover_material"] == "aluminum"
        assert prepared["insulation_layers"] == [
            {
                "thickness": 0.05,
                "material": "other",
                "conductivity": 0.037,
                "temperature_range": (-50.0, 250.0),
            }
        ]

    def test_tank_export_round_trips_canonical_underground_three_layers(self):
        from types import SimpleNamespace

        source = {
            "name": "UG tank",
            "shape": "rectangular",
            "length": 4.0,
            "width": 2.0,
            "height": 3.0,
            "wall_thickness": 0.01,
            "wall_lambda": 45.0,
            "insulation_layers": [
                {"thickness": 0.04, "material": "mineral_wool_boards_120"},
                {
                    "thickness": 0.03,
                    "material": "other",
                    "conductivity": 0.039,
                    "temperature_range": (-40.0, 180.0),
                },
                {"thickness": 0.02, "material": "polyurethane_products_50"},
            ],
            "ambient_temperature": -20.0,
            "ground_temperature": 5.0,
            "process_temperature": 80.0,
            "placement": "underground",
            "insulation_temperature_basis": "channel",
            "tank_buried_height": 1.25,
            "ground_type": "dry_sand",
            "ground_conductivity": 1.2,
            "wind_speed": 0.0,
            "safety_factor": 1.15,
            "q_additional": 0.0,
            "min_switch_temperature": -30.0,
        }

        content = build_objects_xlsx([SimpleNamespace(object_type="tank", params=source)])
        [(_label, _object_type, rows)] = [
            parsed for parsed in _parse_excel_workbook(content) if parsed[1] == "tank"
        ]
        params, err = _build_tank_params(rows[0])

        assert err is None
        assert params is not None
        assert params["insulation_layers"] == source["insulation_layers"]
        assert params["placement"] == "underground"
        assert params["ground_temperature"] == 5.0
        assert params["ground_temperature_source"] == "manual"
        assert params["ground_conductivity_source"] == "reference"
        assert params["tank_buried_height"] == pytest.approx(1.25)
        assert params["wall_thickness"] == pytest.approx(0.01)
        assert params["wall_lambda"] == 45.0
        assert params["q_additional"] == 0.0
        assert params["min_switch_temperature"] == -30.0
        for forbidden in (
            "location",
            "burial_depth",
            "insulation_thickness",
            "insulation_material",
            "insulation_layer_count",
        ):
            assert forbidden not in params


class TestExtendedRoundtripFields:
    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    @pytest.mark.parametrize(
        ("source_value", "expected_source"),
        [
            ("manual", "manual"),
            ("Вручную", "manual"),
            ("climate", "climate"),
            ("Климат", "climate"),
        ],
    )
    def test_parser_reads_ambient_temperature_source_aliases(
        self,
        builder,
        source_value: str,
        expected_source: str,
    ):
        params, err = builder(
            {
                "ambient_temperature": -7,
                "ambient_temperature_source": source_value,
            }
        )

        assert err is None
        assert params is not None
        assert params["ambient_temperature_source"] == expected_source

    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    def test_parser_rejects_unknown_ambient_temperature_source(self, builder):
        params, err = builder(
            {
                "ambient_temperature": -7,
                "ambient_temperature_source": "automatic",
            }
        )

        assert params is None
        assert err == "Не распознан источник температуры среды: automatic"

    def test_legacy_pipe_without_source_keeps_previous_climate_behavior(self):
        params, err = _build_pipe_params(
            {
                "outer_diameter_mm": 108,
                "ambient_temperature": -7,
                "climate_key": "Московская область|||Москва",
                "placement": "outdoor",
            }
        )

        assert err is None
        assert params is not None
        assert "ambient_temperature_source" not in params
        after_policy = apply_climate_policy("pipe", params)
        assert after_policy["ambient_temperature"] == pytest.approx(-23.0)
        assert after_policy["ambient_temperature_source"] == "climate"

    @pytest.mark.parametrize(
        ("source_field", "source_value", "expected_error"),
        [
            (
                "wind_speed_source",
                "climate",
                "Источник скорости ветра указан без скорости ветра",
            ),
            (
                "safety_factor_source",
                "manual",
                "Источник Kзап указан без Kзап",
            ),
        ],
    )
    @pytest.mark.parametrize("builder", [_build_pipe_params, _build_tank_params])
    def test_parser_rejects_source_without_its_value(
        self,
        builder,
        source_field: str,
        source_value: str,
        expected_error: str,
    ):
        params, err = builder({source_field: source_value})

        assert params is None
        assert err == expected_error

    def test_legacy_pipe_without_wind_or_safety_sources_keeps_manual_behavior(self):
        params, err = _build_pipe_params(
            {
                "wind_speed": 3.5,
                "safety_factor": 1.2,
                "placement": "outdoor",
            }
        )

        assert err is None
        assert params is not None
        assert "wind_speed_source" not in params
        assert "safety_factor_source" not in params

    @pytest.mark.parametrize(
        ("ground_type", "expected_source"),
        [(None, "manual"), ("custom", "manual"), ("clay", "reference")],
    )
    def test_pipe_parser_sets_ground_conductivity_provenance(
        self, ground_type: str | None, expected_source: str
    ):
        params, err = _build_pipe_params(
            {
                "placement": "underground",
                "ground_type": ground_type,
                "ground_conductivity": 1.5,
                "ground_temperature": 4,
            }
        )

        assert err is None
        assert params is not None
        assert params["ground_temperature_source"] == "manual"
        assert params["ground_conductivity_source"] == expected_source

    def test_pipe_parser_reads_climate_and_local_element_fields(self):
        params, err = _build_pipe_params(
            {
                "name": "pipe",
                "outer_diameter_mm": 108,
                "pipe_length": 50,
                "insulation_thickness_mm": 50,
                "insulation_material": "Минеральная вата",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "placement": "outdoor",
                "climate_region": "ХМАО",
                "climate_city": "Сургут",
                "climate_temperature_basis": "0,92",
                "safety_factor": 1.2,
                "min_switch_temperature": -35,
                "num_local_elements": 6,
                "local_element_equiv_length": 2.4,
            }
        )

        assert err is None
        assert params is not None
        assert params["climate_region"] == "ХМАО"
        assert params["climate_city"] == "Сургут"
        assert params["climate_temperature_basis"] == "t_0_92"
        assert params["min_switch_temperature"] == -35
        assert params["num_local_elements"] == 6
        assert params["local_element_equiv_length"] == 2.4

    def test_tank_parser_reads_climate_and_q_additional_fields(self):
        params, err = _build_tank_params(
            {
                "name": "tank",
                "shape": "Цилиндр",
                "diameter_mm": 2000,
                "height_mm": 3000,
                "insulation_thickness_mm": 80,
                "insulation_material": "Минеральная вата",
                "ambient_temperature": -20,
                "process_temperature": 80,
                "climate_key": "ХМАО|||Сургут",
                "climate_temperature_basis": "Абс. мин.",
                "q_additional": 250,
            }
        )

        assert err is None
        assert params is not None
        assert params["climate_key"] == "ХМАО|||Сургут"
        assert params["climate_region"] == "ХМАО"
        assert params["climate_city"] == "Сургут"
        assert params["climate_temperature_basis"] == "t_abs_min"
        assert params["q_additional"] == 250


class TestTemplateBuilders:
    def test_xlsx_has_both_sheets(self):
        import io

        from openpyxl import load_workbook

        data = build_template_xlsx()
        wb = load_workbook(io.BytesIO(data))
        assert "Трубопроводы" in wb.sheetnames
        assert "Резервуары" in wb.sheetnames
        # Есть хотя бы одна строка с примером данных
        assert wb["Трубопроводы"].max_row >= 2
        for sheet_name in ("Трубопроводы", "Резервуары"):
            headers = [cell.value for cell in wb[sheet_name][1]]
            assert "Мин. T° окр. среды" in headers
            assert "Макс. T° окр. среды" in headers
            assert "T° среды" not in headers
            assert "Источник T° среды" in headers
            assert "Источник скорости ветра" in headers
            assert "Источник Kзап" in headers

    def test_csv_starts_with_type_header(self):
        data = build_template_csv()
        text = data.decode("utf-8-sig")
        assert text.splitlines()[0].startswith("Тип")
        assert "T проп., °C" in text.splitlines()[0]
        assert "Мин. T° окр. среды" in text.splitlines()[0]
        assert "Макс. T° окр. среды" in text.splitlines()[0]
        assert ";T° среды;" not in text.splitlines()[0]
