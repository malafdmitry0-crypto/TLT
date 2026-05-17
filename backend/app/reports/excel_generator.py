"""Генерация XLSX-отчёта через openpyxl."""

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.spreadsheet_safety import set_safe_cell


def generate_xlsx(context: dict[str, Any]) -> bytes:
    wb = Workbook()
    enabled = set(
        context.get("sections") or ["summary", "pipes", "tanks", "electrical", "specification"]
    )

    # Sheet 1: Проект
    ws_info = wb.active
    assert ws_info is not None
    ws_info.title = "Проект"
    project = context["project"]
    ws_info["A1"] = "HeatCalc — отчёт"
    ws_info["A1"].font = Font(bold=True, size=14, color="1A5276")
    ws_info["A3"] = "Название"
    set_safe_cell(ws_info, 3, 2, project["name"])
    ws_info["A4"] = "ID"
    set_safe_cell(ws_info, 4, 2, project["id"])
    ws_info["A5"] = "Статус"
    set_safe_cell(ws_info, 5, 2, project["status"])
    if project.get("description"):
        ws_info["A6"] = "Описание"
        set_safe_cell(ws_info, 6, 2, project["description"])

    if not enabled & {"pipes", "tanks", "electrical", "summary", "specification"}:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Sheet 2: Объекты
    if enabled & {"pipes", "tanks", "electrical", "summary"}:
        ws_obj = wb.create_sheet("Объекты")
    else:
        ws_obj = None
    if ws_obj is not None:
        headers = ["#", "Тип", "Параметры", "Результат", "Валидно"]
        for col, h in enumerate(headers, start=1):
            cell = ws_obj.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EBF5FB")
        # Фильтруем объекты по выбранным разделам
        type_keys = {"pipes": "pipe", "tanks": "tank"}
        wanted_obj_types = {type_keys[s] for s in ("pipes", "tanks") if s in enabled}
        all_objects = context.get("objects", [])
        if wanted_obj_types:
            objects_to_show = [o for o in all_objects if o.get("object_type") in wanted_obj_types]
        else:
            objects_to_show = all_objects  # summary/electrical — оставляем все
        for idx, o in enumerate(objects_to_show, start=1):
            set_safe_cell(ws_obj, idx + 1, 1, idx)
            set_safe_cell(ws_obj, idx + 1, 2, o["object_type"])
            set_safe_cell(ws_obj, idx + 1, 3, str(o.get("params", {})))
            set_safe_cell(ws_obj, idx + 1, 4, str(o.get("results") or ""))
            set_safe_cell(ws_obj, idx + 1, 5, "да" if o.get("is_valid") else "нет")

    # Sheet 3: Спецификация
    if "specification" in enabled:
        ws_spec = wb.create_sheet("Спецификация")
        spec_headers = ["Категория", "Наименование", "Артикул", "Ед.", "Кол-во"]
        for col, h in enumerate(spec_headers, start=1):
            cell = ws_spec.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EBF5FB")
        for idx, item in enumerate(context.get("specification", {}).get("items", []), start=1):
            set_safe_cell(ws_spec, idx + 1, 1, item.get("category", ""))
            set_safe_cell(ws_spec, idx + 1, 2, item.get("name", ""))
            set_safe_cell(ws_spec, idx + 1, 3, item.get("article") or "")
            set_safe_cell(ws_spec, idx + 1, 4, item.get("unit", ""))
            set_safe_cell(ws_spec, idx + 1, 5, item.get("quantity", 0))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
