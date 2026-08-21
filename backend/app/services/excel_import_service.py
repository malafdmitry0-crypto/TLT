"""Импорт объектов проекта из Excel."""

from __future__ import annotations

import asyncio
import hashlib
import io
import json
import re
import zipfile
from dataclasses import dataclass
from typing import Any, Literal
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, func, select, update
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import use_fast_commit_for_current_transaction
from app.core.dependencies import CurrentPrincipal
from app.models.electrical_calculation import ElectricalCalculation
from app.models.project import Project
from app.models.project_object import ProjectObject
from app.models.specification import Specification
from app.reference_data.loader import list_insulation_materials
from app.services.project_object_params import (
    normalize_project_object_params,
    prepare_project_object_params,
    reject_legacy_specification_object_params,
)
from app.services.project_service import (
    ProjectAccessError,
    ProjectNotFoundError,
    ProjectService,
)
from app.services.spreadsheet_safety import append_safe_row

PIPE_SHEET_NAMES = {"трубопроводы", "трубы", "pipes"}
TANK_SHEET_NAMES = {"резервуары", "ёмкости", "емкости", "tanks"}
IMPORT_COMMIT_BATCH_SIZE = 25
ImportMode = Literal["append", "merge", "replace"]

# Алиасы для колонки «Тип» в CSV (различает трубу/резервуар в одном файле)
TYPE_ALIASES: dict[str, str] = {
    "труба": "pipe",
    "трубопровод": "pipe",
    "pipe": "pipe",
    "резервуар": "tank",
    "ёмкость": "tank",
    "емкость": "tank",
    "бак": "tank",
    "tank": "tank",
}

GENERIC_MATERIAL_ALIASES: dict[str, str] = {
    "минеральная вата": "mineral_wool",
    "мин вата": "mineral_wool",
    "мин. вата": "mineral_wool",
    "минвата": "mineral_wool",
    "mineral_wool": "mineral_wool",
    "пеностекло": "foam_glass",
    "foam_glass": "foam_glass",
    "пенополиуретан": "polyurethane",
    "ппу": "polyurethane",
    "polyurethane": "polyurethane",
    "пенополистирол": "polystyrene",
    "полистирол": "polystyrene",
    "polystyrene": "polystyrene",
    "аэрогель": "aerogel",
    "aerogel": "aerogel",
    "силикат кальция": "calcium_silicate",
    "calcium_silicate": "calcium_silicate",
}

SPECIAL_MATERIAL_ALIASES: dict[str, str] = {
    "другое": "other",
    "other": "other",
}

PIPE_MATERIAL_ALIASES: dict[str, str] = {
    "углеродистая сталь": "carbon_steel",
    "сталь углеродистая": "carbon_steel",
    "carbon_steel": "carbon_steel",
    "нержавеющая сталь": "stainless_304",
    "stainless_304": "stainless_304",
    "медь": "copper",
    "copper": "copper",
    "алюминий": "aluminum",
    "aluminum": "aluminum",
    "пластик": "plastic",
    "plastic": "plastic",
}

PLACEMENT_ALIASES: dict[str, str] = {
    "на открытом воздухе": "outdoor",
    "улица": "outdoor",
    "надземная": "outdoor",
    "outdoor": "outdoor",
    "в помещении": "indoor",
    "помещение": "indoor",
    "indoor": "indoor",
    "подземно": "underground",
    "подземная": "underground",
    "underground": "underground",
}

GROUND_ALIASES: dict[str, str] = {
    "сухой песок": "dry_sand",
    "dry_sand": "dry_sand",
    "влажный песок": "wet_sand",
    "wet_sand": "wet_sand",
    "глина": "clay",
    "clay": "clay",
    "другое": "custom",
    "custom": "custom",
}

CLIMATE_BASIS_ALIASES: dict[str, str] = {
    "0,92": "t_0_92",
    "0.92": "t_0_92",
    "t_0_92": "t_0_92",
    "t0.92": "t_0_92",
    "t 0.92": "t_0_92",
    "0,98": "t_0_98",
    "0.98": "t_0_98",
    "t_0_98": "t_0_98",
    "t0.98": "t_0_98",
    "t 0.98": "t_0_98",
    "абс. мин.": "t_abs_min",
    "абс мин": "t_abs_min",
    "абсолютный минимум": "t_abs_min",
    "t_abs_min": "t_abs_min",
}

INSULATION_TEMPERATURE_BASIS_ALIASES: dict[str, str] = {
    "indoor": "indoor",
    "помещение": "indoor",
    "в помещении": "indoor",
    "outdoor_summer": "outdoor_summer",
    "улица лето": "outdoor_summer",
    "открытый воздух лето": "outdoor_summer",
    "открытый воздух, лето": "outdoor_summer",
    "outdoor_winter": "outdoor_winter",
    "улица зима": "outdoor_winter",
    "открытый воздух зима": "outdoor_winter",
    "открытый воздух, зима": "outdoor_winter",
    "channel": "channel",
    "канал": "channel",
    "tunnel": "tunnel",
    "тоннель": "tunnel",
    "technical_subfloor": "technical_subfloor",
    "техническое подполье": "technical_subfloor",
    "подполье": "technical_subfloor",
    "attic": "attic",
    "чердак": "attic",
    "basement": "basement",
    "подвал": "basement",
}

SHAPE_ALIASES: dict[str, str] = {
    "цилиндр": "cylindrical",
    "цилиндрический": "cylindrical",
    "cylindrical": "cylindrical",
    "параллелепипед": "rectangular",
    "прямоугольный": "rectangular",
    "прямоуг": "rectangular",
    "rectangular": "rectangular",
}


# Заголовок колонки типа в CSV
TYPE_HEADERS: set[str] = {"тип", "type"}


# Возможные заголовки (нормализованные) → каноническое имя поля
PIPE_HEADERS: dict[str, str] = {
    "наименование": "name",
    "название": "name",
    "имя": "name",
    "диаметр, мм": "outer_diameter_mm",
    "диаметр мм": "outer_diameter_mm",
    "диаметр": "outer_diameter_mm",
    "ø, мм": "outer_diameter_mm",
    "ø мм": "outer_diameter_mm",
    "длина, м": "pipe_length",
    "длина м": "pipe_length",
    "длина": "pipe_length",
    "l, м": "pipe_length",
    "толщина изоляции, мм": "insulation_thickness_mm",
    "толщина изоляции мм": "insulation_thickness_mm",
    "толщина изоляции": "insulation_thickness_mm",
    "δ, мм": "insulation_thickness_mm",
    "δ мм": "insulation_thickness_mm",
    "материал изоляции": "insulation_material",
    "код материала изоляции": "insulation_material",
    "материал": "insulation_material",
    "т° среды": "ambient_temperature",
    "т среды": "ambient_temperature",
    "t° среды": "ambient_temperature",
    "t среды": "ambient_temperature",
    "температура среды": "ambient_temperature",
    "т° продукта": "process_temperature",
    "т продукта": "process_temperature",
    "t° продукта": "process_temperature",
    "t продукта": "process_temperature",
    "температура продукта": "process_temperature",
    "требуемая температура": "process_temperature",
    "требуемая температура трубы": "process_temperature",
    "температура поддержания": "process_temperature",
    "т° поддержания": "process_temperature",
    "t° поддержания": "process_temperature",
    "t проп": "vapor_temperature",
    "t проп.": "vapor_temperature",
    "t° проп": "vapor_temperature",
    "t° пропарки": "vapor_temperature",
    "t проп., °c": "vapor_temperature",
    "температура пропарки": "vapor_temperature",
    "t3": "maintain_temperature",
    "t3, °c": "maintain_temperature",
    "t3 поддержания": "maintain_temperature",
    "температура поддержания t3": "maintain_temperature",
    "макс. t° окр. среды": "max_ambient_temperature",
    "макс t° окр. среды": "max_ambient_temperature",
    "макс. допуст. t° продукта": "max_process_temperature",
    "макс допуст t° продукта": "max_process_temperature",
    "толщина стенки, мм": "wall_thickness_mm",
    "толщина стенки мм": "wall_thickness_mm",
    "толщина стенки": "wall_thickness_mm",
    "материал трубы": "pipe_material",
    "λ трубы": "pipe_lambda",
    "лямбда трубы": "pipe_lambda",
    "материал 2-го слоя": "second_insulation_material",
    "толщина 2-го слоя, мм": "second_insulation_thickness_mm",
    "толщина 2-го слоя": "second_insulation_thickness_mm",
    "λ 1-го слоя": "first_insulation_lambda",
    "диапазон температур 1-го слоя, °c": "first_insulation_temperature_range",
    "λ 2-го слоя": "second_insulation_lambda",
    "диапазон температур 2-го слоя, °c": "second_insulation_temperature_range",
    "материал 3-го слоя": "third_insulation_material",
    "толщина 3-го слоя, мм": "third_insulation_thickness_mm",
    "толщина 3-го слоя": "third_insulation_thickness_mm",
    "λ 3-го слоя": "third_insulation_lambda",
    "диапазон температур 3-го слоя, °c": "third_insulation_temperature_range",
    "материал покрытия": "insulation_cover_material",
    "размещение": "placement",
    "размещение трубопровода": "placement",
    "глубина прокладки": "pipe_centerline_depth",
    "глубина заложения оси трубы, м": "pipe_centerline_depth",
    "глубина заложения оси трубы": "pipe_centerline_depth",
    "грунт": "ground_type",
    "λ грунта": "ground_conductivity",
    "температура грунта": "ground_temperature",
    "t° грунта": "ground_temperature",
    "скорость ветра, м/с": "wind_speed",
    "скорость ветра": "wind_speed",
    "α внешней теплоотдачи": "alpha_vnesh",
    "alpha_vnesh": "alpha_vnesh",
    "kзап": "safety_factor",
    "k зап": "safety_factor",
    "коэффициент запаса": "safety_factor",
    "рабочее напряжение": "supply_voltage",
    "мин. t включения, °c": "min_switch_temperature",
    "мин t включения": "min_switch_temperature",
    "минимальная температура включения": "min_switch_temperature",
    "климатический регион": "climate_region",
    "регион климата": "climate_region",
    "климатический город": "climate_city",
    "город климата": "climate_city",
    "климат": "climate_city",
    "ключ климата": "climate_key",
    "climate_key": "climate_key",
    "обеспеченность климата": "climate_temperature_basis",
    "температура климата": "climate_temperature_basis",
    "режим температуры изоляции": "insulation_temperature_basis",
    "tm изоляции": "insulation_temperature_basis",
    "tм изоляции": "insulation_temperature_basis",
    "количество локальных элементов": "num_local_elements",
    "локальных элементов": "num_local_elements",
    "l экв., м": "local_element_equiv_length",
    "l экв. м": "local_element_equiv_length",
    "эквивалентная длина локального элемента": "local_element_equiv_length",
    "эквивалентная длина, м": "local_element_equiv_length",
}

TANK_HEADERS: dict[str, str] = {
    "наименование": "name",
    "название": "name",
    "форма": "shape",
    "диаметр, мм": "diameter_mm",
    "диаметр мм": "diameter_mm",
    "диаметр": "diameter_mm",
    "длина, мм": "length_mm",
    "длина мм": "length_mm",
    "длина": "length_mm",
    "ширина, мм": "width_mm",
    "ширина мм": "width_mm",
    "ширина": "width_mm",
    "высота, мм": "height_mm",
    "высота мм": "height_mm",
    "высота": "height_mm",
    "толщина стенки, мм": "wall_thickness_mm",
    "толщина стенки мм": "wall_thickness_mm",
    "λ стенки": "wall_lambda",
    "лямбда стенки": "wall_lambda",
    "толщина изоляции, мм": "insulation_thickness_mm",
    "толщина изоляции мм": "insulation_thickness_mm",
    "толщина изоляции": "insulation_thickness_mm",
    "δ, мм": "insulation_thickness_mm",
    "δ мм": "insulation_thickness_mm",
    "материал изоляции": "insulation_material",
    "код материала изоляции": "insulation_material",
    "материал": "insulation_material",
    "т° среды": "ambient_temperature",
    "т среды": "ambient_temperature",
    "t° среды": "ambient_temperature",
    "t среды": "ambient_temperature",
    "температура среды": "ambient_temperature",
    "т° продукта": "process_temperature",
    "т продукта": "process_temperature",
    "t° продукта": "process_temperature",
    "t продукта": "process_temperature",
    "температура продукта": "process_temperature",
    "требуемая температура": "process_temperature",
    "требуемая температура трубы": "process_temperature",
    "температура поддержания": "process_temperature",
    "т° поддержания": "process_temperature",
    "t° поддержания": "process_temperature",
    "t проп": "vapor_temperature",
    "t проп.": "vapor_temperature",
    "t° проп": "vapor_temperature",
    "t° пропарки": "vapor_temperature",
    "t проп., °c": "vapor_temperature",
    "температура пропарки": "vapor_temperature",
    "t3": "maintain_temperature",
    "t3, °c": "maintain_temperature",
    "t3 поддержания": "maintain_temperature",
    "температура поддержания t3": "maintain_temperature",
    "макс. t° окр. среды": "max_ambient_temperature",
    "макс t° окр. среды": "max_ambient_temperature",
    "макс. допуст. t° продукта": "max_process_temperature",
    "макс допуст t° продукта": "max_process_temperature",
    "материал 2-го слоя": "second_insulation_material",
    "толщина 2-го слоя, мм": "second_insulation_thickness_mm",
    "толщина 2-го слоя": "second_insulation_thickness_mm",
    "λ 1-го слоя": "first_insulation_lambda",
    "диапазон температур 1-го слоя, °c": "first_insulation_temperature_range",
    "λ 2-го слоя": "second_insulation_lambda",
    "диапазон температур 2-го слоя, °c": "second_insulation_temperature_range",
    "материал 3-го слоя": "third_insulation_material",
    "толщина 3-го слоя, мм": "third_insulation_thickness_mm",
    "толщина 3-го слоя": "third_insulation_thickness_mm",
    "λ 3-го слоя": "third_insulation_lambda",
    "диапазон температур 3-го слоя, °c": "third_insulation_temperature_range",
    "материал покрытия": "insulation_cover_material",
    "размещение": "placement",
    "размещение резервуара": "placement",
    "высота заглубленной части, м": "tank_buried_height",
    "высота заглублённой части, м": "tank_buried_height",
    "грунт": "ground_type",
    "λ грунта": "ground_conductivity",
    "температура грунта": "ground_temperature",
    "t° грунта": "ground_temperature",
    "скорость ветра, м/с": "wind_speed",
    "скорость ветра": "wind_speed",
    "α внешней теплоотдачи": "alpha_vnesh",
    "alpha_vnesh": "alpha_vnesh",
    "kзап": "safety_factor",
    "k зап": "safety_factor",
    "коэффициент запаса": "safety_factor",
    "рабочее напряжение": "supply_voltage",
    "мин. t включения, °c": "min_switch_temperature",
    "мин t включения": "min_switch_temperature",
    "минимальная температура включения": "min_switch_temperature",
    "климатический регион": "climate_region",
    "регион климата": "climate_region",
    "климатический город": "climate_city",
    "город климата": "climate_city",
    "климат": "climate_city",
    "ключ климата": "climate_key",
    "climate_key": "climate_key",
    "обеспеченность климата": "climate_temperature_basis",
    "температура климата": "climate_temperature_basis",
    "режим температуры изоляции": "insulation_temperature_basis",
    "tm изоляции": "insulation_temperature_basis",
    "tм изоляции": "insulation_temperature_basis",
    "q доп., вт": "q_additional",
    "q доп, вт": "q_additional",
    "дополнительные теплопотери, вт": "q_additional",
}


class ExcelImportError(Exception):
    """Ошибка импорта Excel/CSV."""


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > settings.MAX_XLSX_FILES:
                raise ExcelImportError(
                    f"XLSX содержит слишком много внутренних файлов: {len(infos)}"
                )
            total_uncompressed = sum(info.file_size for info in infos)
            if total_uncompressed > settings.MAX_XLSX_UNCOMPRESSED_BYTES:
                raise ExcelImportError(
                    "XLSX слишком большой после распаковки " f"({total_uncompressed // 1024} КБ)"
                )
            for info in infos:
                if info.compress_size and info.file_size / info.compress_size > 100:
                    raise ExcelImportError(
                        "XLSX похож на zip-bomb: слишком высокий коэффициент сжатия"
                    )
    except zipfile.BadZipFile as exc:
        raise ExcelImportError("Файл не является корректным XLSX-архивом") from exc


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, int | float):
        return float(v)
    s = str(v).replace(",", ".").strip()
    match = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if match:
        s = match.group(0)
    try:
        return float(s)
    except ValueError:
        return None


def _to_temperature_range(v: Any) -> tuple[float, float] | None:
    if isinstance(v, list | tuple) and len(v) == 2:
        lower = _to_float(v[0])
        upper = _to_float(v[1])
    elif v not in (None, ""):
        values = re.findall(r"[-+]?\d+(?:[.,]\d+)?", str(v))
        if len(values) != 2:
            return None
        lower, upper = (_to_float(value) for value in values)
    else:
        return None
    if lower is None or upper is None:
        return None
    return lower, upper


@dataclass(frozen=True)
class _MaterialResolution:
    material: str | None = None
    raw: str | None = None
    family: str | None = None
    needs_reselection: bool = False


def _selectable_insulation_entries() -> list[dict[str, Any]]:
    return [
        entry
        for entry in list_insulation_materials()
        if entry.get("selectable") is not False and entry.get("deprecated") is not True
    ]


def _resolve_material_entry(v: Any) -> _MaterialResolution:
    key = _norm(v)
    if not key:
        return _MaterialResolution()
    raw = str(v).strip()
    if key in SPECIAL_MATERIAL_ALIASES:
        return _MaterialResolution(material=SPECIAL_MATERIAL_ALIASES[key], raw=raw)
    if key in GENERIC_MATERIAL_ALIASES:
        return _MaterialResolution(
            raw=raw,
            family=GENERIC_MATERIAL_ALIASES[key],
            needs_reselection=True,
        )

    entries = _selectable_insulation_entries()
    by_code = {str(entry["material"]).lower(): entry for entry in entries}
    code_match = by_code.get(key)
    if code_match is not None:
        return _MaterialResolution(material=str(code_match["material"]), raw=raw)

    name_matches = [entry for entry in entries if _norm(entry.get("name")) == key]
    if len(name_matches) == 1:
        return _MaterialResolution(material=str(name_matches[0]["material"]), raw=raw)
    if len(name_matches) > 1:
        family = str(name_matches[0].get("material_family") or name_matches[0]["material"])
        return _MaterialResolution(raw=raw, family=family, needs_reselection=True)
    return _MaterialResolution()


def _resolve_material(v: Any) -> str | None:
    return _resolve_material_entry(v).material


def _resolve_pipe_material(v: Any) -> str | None:
    key = _norm(v)
    if not key:
        return None
    return PIPE_MATERIAL_ALIASES.get(key)


def _resolve_alias(v: Any, aliases: dict[str, str]) -> str | None:
    key = _norm(v)
    if not key:
        return None
    return aliases.get(key)


def _resolve_climate_basis(v: Any) -> str | None:
    key = _norm(v)
    if not key:
        return None
    return CLIMATE_BASIS_ALIASES.get(key)


def _resolve_shape(v: Any) -> str | None:
    key = _norm(v)
    if not key:
        return None
    # Первое слово тоже пробуем (если запись «цилиндрический бак»)
    return SHAPE_ALIASES.get(key) or SHAPE_ALIASES.get(key.split()[0] if key else "")


def _canonical_tank_insulation_layers(
    row: dict[str, Any],
    *,
    first_thickness_mm: float | None,
    first_material: _MaterialResolution,
) -> list[dict[str, Any]]:
    layers: list[dict[str, Any]] = []
    specs = (
        (
            0,
            first_thickness_mm,
            first_material,
            _to_float(row.get("first_insulation_lambda")),
            _to_temperature_range(row.get("first_insulation_temperature_range")),
        ),
        (
            1,
            _to_float(row.get("second_insulation_thickness_mm")),
            _resolve_material_entry(row.get("second_insulation_material")),
            _to_float(row.get("second_insulation_lambda")),
            _to_temperature_range(row.get("second_insulation_temperature_range")),
        ),
        (
            2,
            _to_float(row.get("third_insulation_thickness_mm")),
            _resolve_material_entry(row.get("third_insulation_material")),
            _to_float(row.get("third_insulation_lambda")),
            _to_temperature_range(row.get("third_insulation_temperature_range")),
        ),
    )
    for index, thickness_mm, material_resolution, conductivity, temperature_range in specs:
        if not any(
            value is not None
            for value in (
                thickness_mm,
                material_resolution.raw,
                conductivity,
                temperature_range,
            )
        ):
            continue
        while len(layers) < index:
            layers.append({})
        layer: dict[str, Any] = {}
        if thickness_mm is not None:
            layer["thickness"] = thickness_mm / 1000.0
        if material_resolution.material is not None:
            layer["material"] = material_resolution.material
        if material_resolution.material == "other":
            if conductivity is not None:
                layer["conductivity"] = conductivity
            if temperature_range is not None:
                layer["temperature_range"] = temperature_range
        layers.append(layer)
    return layers


def _read_sheet(ws: Any, header_map: dict[str, str]) -> list[dict[str, Any]]:
    """Читает лист, мапит заголовки → канонические имена, возвращает список строк-словарей.

    Строки возвращаются с ключом ``_row`` — номер строки в Excel (1-based).
    """
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    mapped_cols: list[tuple[int, str]] = []
    for idx, h in enumerate(header_row):
        key = header_map.get(_norm(h))
        if key:
            mapped_cols.append((idx, key))
    result: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        if len(result) >= settings.MAX_IMPORT_ROWS:
            raise ExcelImportError(f"Превышен лимит строк импорта: {settings.MAX_IMPORT_ROWS}")
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        item: dict[str, Any] = {"_row": row_idx}
        for idx, key in mapped_cols:
            if idx < len(row):
                item[key] = row[idx]
        result.append(item)
    return result


def _build_pipe_params(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    """Строит params-словарь для трубы (м, не мм). Возвращает (params, error)."""
    d_mm = _to_float(row.get("outer_diameter_mm"))
    L = _to_float(row.get("pipe_length"))
    ins_mm = _to_float(row.get("insulation_thickness_mm"))
    material_resolution = _resolve_material_entry(row.get("insulation_material"))
    t_a = _to_float(row.get("ambient_temperature"))
    t_p = _to_float(row.get("process_temperature"))

    placement_raw = row.get("placement")
    placement = _resolve_alias(placement_raw, PLACEMENT_ALIASES)
    if placement_raw not in (None, "") and placement is None:
        return None, f"Не распознано размещение трубопровода: {placement_raw}"

    pipe_material_raw = row.get("pipe_material")
    pipe_material = _resolve_pipe_material(pipe_material_raw)
    if pipe_material_raw not in (None, "") and pipe_material is None:
        return None, f"Не распознан материал трубы: {pipe_material_raw}"
    pipe_lambda = _to_float(row.get("pipe_lambda"))
    if pipe_material is not None and pipe_lambda is not None:
        return None, "Задайте только один источник λ трубы: материал или λ"

    params: dict[str, Any] = {}
    if d_mm is not None:
        params["outer_diameter"] = d_mm / 1000.0
    if L is not None:
        params["pipe_length"] = L
    layers: list[dict[str, Any]] = []
    first_layer: dict[str, Any] = {}
    if ins_mm is not None:
        first_layer["thickness"] = ins_mm / 1000.0
    if material_resolution.material:
        first_layer["material"] = material_resolution.material
    first_lambda = _to_float(row.get("first_insulation_lambda"))
    if first_lambda is not None:
        first_layer["conductivity"] = first_lambda
    first_temperature_range = _to_temperature_range(
        row.get("first_insulation_temperature_range")
    )
    if first_temperature_range is not None:
        first_layer["temperature_range"] = first_temperature_range
    if first_layer:
        layers.append(first_layer)

    for material_field, thickness_field, lambda_field, temperature_range_field in (
        (
            "second_insulation_material",
            "second_insulation_thickness_mm",
            "second_insulation_lambda",
            "second_insulation_temperature_range",
        ),
        (
            "third_insulation_material",
            "third_insulation_thickness_mm",
            "third_insulation_lambda",
            "third_insulation_temperature_range",
        ),
    ):
        resolution = _resolve_material_entry(row.get(material_field))
        thickness = _to_float(row.get(thickness_field))
        conductivity = _to_float(row.get(lambda_field))
        temperature_range = _to_temperature_range(row.get(temperature_range_field))
        if not (
            resolution.material
            or thickness is not None
            or conductivity is not None
            or temperature_range is not None
        ):
            continue
        layer: dict[str, Any] = {}
        if thickness is not None:
            layer["thickness"] = thickness / 1000.0
        if resolution.material:
            layer["material"] = resolution.material
        if conductivity is not None:
            layer["conductivity"] = conductivity
        if temperature_range is not None:
            layer["temperature_range"] = temperature_range
        layers.append(layer)
    if layers:
        params["insulation_layers"] = layers[:3]

    if t_p is not None:
        params["process_temperature"] = t_p
    wall_mm = _to_float(row.get("wall_thickness_mm"))
    if wall_mm is not None:
        params["wall_thickness"] = wall_mm / 1000.0
    if pipe_material:
        params["pipe_material"] = pipe_material
    if pipe_lambda is not None:
        params["pipe_lambda"] = pipe_lambda

    local_element_equiv_length = _to_float(row.get("local_element_equiv_length"))
    if local_element_equiv_length is not None:
        params["local_element_equiv_length"] = local_element_equiv_length
    explicit_local_count = _to_float(row.get("num_local_elements"))
    params["num_local_elements"] = int(explicit_local_count or 0)

    if placement:
        params["placement"] = placement
    if placement == "underground":
        depth = _to_float(row.get("pipe_centerline_depth"))
        if depth is not None:
            params["pipe_centerline_depth"] = depth
        ground_temperature = _to_float(row.get("ground_temperature"))
        if ground_temperature is not None:
            params["ground_temperature"] = ground_temperature
            params["ground_temperature_source"] = "manual"
        ground_conductivity = _to_float(row.get("ground_conductivity"))
        if ground_conductivity is not None:
            params["ground_conductivity"] = ground_conductivity
        ground_type = _resolve_alias(row.get("ground_type"), GROUND_ALIASES)
        if ground_type:
            params["ground_type"] = ground_type
        if ground_type or ground_conductivity is not None:
            params["ground_conductivity_source"] = (
                "reference" if ground_type not in (None, "custom") else "manual"
            )
    elif t_a is not None:
        params["ambient_temperature"] = t_a

    wind_speed = _to_float(row.get("wind_speed"))
    if wind_speed is not None:
        params["wind_speed"] = wind_speed
    alpha_vnesh = _to_float(row.get("alpha_vnesh"))
    if alpha_vnesh is not None:
        params["alpha_vnesh"] = alpha_vnesh

    safety_factor = _to_float(row.get("safety_factor"))
    if safety_factor is not None:
        params["safety_factor"] = safety_factor
    basis = _resolve_alias(
        row.get("insulation_temperature_basis"), INSULATION_TEMPERATURE_BASIS_ALIASES
    )
    if basis:
        params["insulation_temperature_basis"] = basis

    for field in (
        "vapor_temperature",
        "maintain_temperature",
        "max_ambient_temperature",
        "max_process_temperature",
        "min_switch_temperature",
        "supply_voltage",
    ):
        value = _to_float(row.get(field))
        if value is not None:
            params[field] = value
    for field in ("climate_key", "climate_region", "climate_city"):
        value = row.get(field)
        if value and str(value).strip():
            params[field] = str(value).strip()
    climate_basis = _resolve_climate_basis(row.get("climate_temperature_basis"))
    if climate_basis:
        params["climate_temperature_basis"] = climate_basis
    cover = row.get("insulation_cover_material")
    if cover and str(cover).strip():
        params["insulation_cover_material"] = str(cover).strip()
    name = row.get("name")
    if name and str(name).strip():
        params["name"] = str(name).strip()
    return params, None


def _build_tank_params(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    shape = _resolve_shape(row.get("shape"))
    if not shape and row.get("shape") not in (None, ""):
        return (
            None,
            "Форма резервуара больше не поддерживается. "
            "Допустимые формы: cylindrical, rectangular.",
        )

    ins_mm = _to_float(row.get("insulation_thickness_mm"))
    material_resolution = _resolve_material_entry(row.get("insulation_material"))
    t_a = _to_float(row.get("ambient_temperature"))
    t_p = _to_float(row.get("process_temperature"))

    placement_raw = row.get("placement")
    placement = _resolve_alias(placement_raw, PLACEMENT_ALIASES)
    if placement_raw not in (None, "") and placement is None:
        return None, f"Не распознано размещение резервуара: {placement_raw}"

    # The spreadsheet default is explicit outdoor placement; no other field
    # is allowed to infer the calculation mode.
    params: dict[str, Any] = {"placement": placement or "outdoor", "q_additional": 0.0}
    if shape:
        params["shape"] = shape
    if t_a is not None:
        params["ambient_temperature"] = t_a
    if t_p is not None:
        params["process_temperature"] = t_p

    if shape == "cylindrical":
        d_mm = _to_float(row.get("diameter_mm"))
        h_mm = _to_float(row.get("height_mm"))
        if d_mm is not None:
            params["diameter"] = d_mm / 1000.0
        if h_mm is not None:
            params["height"] = h_mm / 1000.0
    elif shape == "rectangular":
        L_mm = _to_float(row.get("length_mm"))
        W_mm = _to_float(row.get("width_mm"))
        H_mm = _to_float(row.get("height_mm"))
        if L_mm is not None:
            params["length"] = L_mm / 1000.0
        if W_mm is not None:
            params["width"] = W_mm / 1000.0
        if H_mm is not None:
            params["height"] = H_mm / 1000.0
    layers = _canonical_tank_insulation_layers(
        row,
        first_thickness_mm=ins_mm,
        first_material=material_resolution,
    )
    if layers:
        params["insulation_layers"] = layers

    wall_thickness_mm = _to_float(row.get("wall_thickness_mm"))
    wall_lambda = _to_float(row.get("wall_lambda"))
    if wall_thickness_mm is not None:
        params["wall_thickness"] = wall_thickness_mm / 1000.0
    if wall_lambda is not None:
        params["wall_lambda"] = wall_lambda

    if params["placement"] == "underground":
        buried_height = _to_float(row.get("tank_buried_height"))
        if buried_height is not None:
            params["tank_buried_height"] = buried_height
        ground_temperature = _to_float(row.get("ground_temperature"))
        if ground_temperature is not None:
            params["ground_temperature"] = ground_temperature
            params["ground_temperature_source"] = "manual"
        ground_type = _resolve_alias(row.get("ground_type"), GROUND_ALIASES)
        if ground_type:
            params["ground_type"] = ground_type
        ground_conductivity = _to_float(row.get("ground_conductivity"))
        if ground_conductivity is not None:
            params["ground_conductivity"] = ground_conductivity
        if ground_type or ground_conductivity is not None:
            params["ground_conductivity_source"] = (
                "reference" if ground_type not in (None, "custom") else "manual"
            )

    wind_speed = _to_float(row.get("wind_speed"))
    if wind_speed is not None:
        params["wind_speed"] = wind_speed
        params["wind_speed_source"] = "manual"
    alpha_vnesh = _to_float(row.get("alpha_vnesh"))
    if alpha_vnesh is not None:
        params["alpha_vnesh"] = alpha_vnesh
    safety_factor = _to_float(row.get("safety_factor"))
    if safety_factor is not None:
        params["safety_factor"] = safety_factor
        params["safety_factor_source"] = "manual"
    basis = _resolve_alias(
        row.get("insulation_temperature_basis"), INSULATION_TEMPERATURE_BASIS_ALIASES
    )
    if basis:
        params["insulation_temperature_basis"] = basis
    else:
        params["insulation_temperature_basis"] = {
            "indoor": "indoor",
            "outdoor": "outdoor_winter",
            "underground": "channel",
        }[params["placement"]]
    for field in (
        "vapor_temperature",
        "maintain_temperature",
        "max_ambient_temperature",
        "max_process_temperature",
        "min_switch_temperature",
        "supply_voltage",
    ):
        value = _to_float(row.get(field))
        if value is not None:
            params[field] = value
    for field in ("climate_key", "climate_region", "climate_city", "insulation_cover_material"):
        value = row.get(field)
        if value and str(value).strip():
            params[field] = str(value).strip()
    climate_key = params.get("climate_key")
    if isinstance(climate_key, str) and "|||" in climate_key:
        region, city = climate_key.split("|||", 1)
        params.setdefault("climate_region", region.strip())
        params.setdefault("climate_city", city.strip())
    climate_basis = _resolve_climate_basis(row.get("climate_temperature_basis"))
    if climate_basis:
        params["climate_temperature_basis"] = climate_basis
    q_additional = _to_float(row.get("q_additional"))
    if q_additional is not None:
        params["q_additional"] = q_additional

    name = row.get("name")
    if name and str(name).strip():
        params["name"] = str(name).strip()
    return params, None


def _parse_csv(content: bytes) -> list[tuple[str, list[dict[str, Any]]]]:
    """Парсит CSV-файл.

    Возвращает список пар (sheet_label, rows) по типу:
    [('Трубопроводы', [...]), ('Резервуары', [...])].
    CSV должен содержать колонку «Тип» со значениями «труба» / «резервуар».
    Автодетект разделителя (``,``, ``;``, ``\t``). Кодировки: UTF-8 / UTF-8-BOM / CP1251.
    """
    import csv

    # Определяем кодировку
    text: str | None = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ExcelImportError("Не удалось определить кодировку CSV (ожидается UTF-8 или CP1251)")

    # Автодетект разделителя по первой строке
    first_line = text.splitlines()[0] if text.strip() else ""
    if not first_line:
        raise ExcelImportError("CSV-файл пустой")
    counts = {d: first_line.count(d) for d in (";", ",", "\t")}
    delimiter = max(counts, key=counts.get) if any(counts.values()) else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        raise ExcelImportError("CSV-файл пустой")
    if len(all_rows) - 1 > settings.MAX_IMPORT_ROWS:
        raise ExcelImportError(f"Превышен лимит строк импорта: {settings.MAX_IMPORT_ROWS}")

    header = all_rows[0]
    # Индекс колонки «Тип»
    type_idx = next((i for i, h in enumerate(header) if _norm(h) in TYPE_HEADERS), None)
    if type_idx is None:
        raise ExcelImportError(
            "В CSV не найдена колонка «Тип» (со значениями «труба»/«резервуар»). "
            "Используйте шаблон CSV или файл Excel."
        )

    # Разделяем строки по типу
    pipe_rows_raw: list[tuple[int, list]] = []
    tank_rows_raw: list[tuple[int, list]] = []
    for row_idx, row in enumerate(all_rows[1:], start=2):
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        if type_idx >= len(row):
            continue
        t = TYPE_ALIASES.get(_norm(row[type_idx]))
        if t == "pipe":
            pipe_rows_raw.append((row_idx, row))
        elif t == "tank":
            tank_rows_raw.append((row_idx, row))

    def build_mapped(
        rows_raw: list[tuple[int, list]], header_map: dict[str, str]
    ) -> list[dict[str, Any]]:
        # Определяем маппинг колонок общего header
        mapped_cols: list[tuple[int, str]] = []
        for idx, h in enumerate(header):
            key = header_map.get(_norm(h))
            if key:
                mapped_cols.append((idx, key))
        out: list[dict[str, Any]] = []
        for row_idx, row in rows_raw:
            item: dict[str, Any] = {"_row": row_idx}
            for idx, key in mapped_cols:
                if idx < len(row):
                    val = row[idx]
                    # Пустые строки — None
                    item[key] = val if (val is not None and str(val).strip() != "") else None
            out.append(item)
        return out

    result: list[tuple[str, list[dict[str, Any]]]] = []
    if pipe_rows_raw:
        result.append(("Трубопроводы (CSV)", build_mapped(pipe_rows_raw, PIPE_HEADERS)))
    if tank_rows_raw:
        result.append(("Резервуары (CSV)", build_mapped(tank_rows_raw, TANK_HEADERS)))
    return result


async def _project_import_state(db: AsyncSession, project_id: UUID) -> tuple[int, int]:
    result = await db.execute(
        select(func.count(ProjectObject.id), func.max(ProjectObject.sort_order)).where(
            ProjectObject.project_id == project_id
        )
    )
    count, max_sort = result.one()
    return int(count or 0), int(max_sort if max_sort is not None else -1) + 1


def _normalize_name_for_dedupe(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _dedupe_key(object_type: str, params: dict[str, Any]) -> str:
    key_params = dict(params)
    name = _normalize_name_for_dedupe(key_params.pop("name", ""))
    payload = json.dumps(
        key_params,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    return f"{object_type}:{name}:{digest}"


def _object_type_for_dedupe(value: Any) -> str:
    return str(getattr(value, "value", value))


async def _existing_dedupe_keys(db: AsyncSession, project_id: UUID) -> set[str]:
    result = await db.execute(
        select(ProjectObject.object_type, ProjectObject.params).where(
            ProjectObject.project_id == project_id
        )
    )
    return {
        _dedupe_key(_object_type_for_dedupe(object_type), params or {})
        for object_type, params in result.all()
    }


async def _touch_project_updated_at(db: AsyncSession, project_id: UUID) -> None:
    """Кейс §4.4/§4.6: импорт объектов обновляет «Последнее изменение» проекта."""
    await db.execute(
        update(Project).where(Project.id == project_id).values(updated_at=func.now())
    )
    await db.commit()


async def _replace_project_objects(db: AsyncSession, project_id: UUID) -> None:
    await db.execute(
        delete(ElectricalCalculation).where(ElectricalCalculation.project_id == project_id)
    )
    await db.execute(delete(Specification).where(Specification.project_id == project_id))
    await db.execute(delete(ProjectObject).where(ProjectObject.project_id == project_id))
    await db.flush()


def _validate_import_mode(mode: str) -> ImportMode:
    normalized = (mode or "merge").strip().lower()
    if normalized not in {"append", "merge", "replace"}:
        raise ExcelImportError(
            "Некорректный режим импорта: " f"{mode!r} (допустимо: append, merge, replace)"
        )
    return normalized  # type: ignore[return-value]


async def _ensure_import_access(
    db: AsyncSession,
    project_id: UUID,
    principal: CurrentPrincipal,
) -> None:
    project_service = ProjectService(db)
    project = await project_service.get_project_basic(project_id, principal)
    project_service._check_owner(project, principal)


async def _commit_object_batch(
    db: AsyncSession,
    batch: list[tuple[ProjectObject, dict[str, Any]]],
    sheet_label: str,
) -> tuple[int, list[UUID], list[dict[str, Any]]]:
    if not batch:
        return 0, [], []
    objects = [item[0] for item in batch]
    try:
        db.add_all(objects)
        await db.flush()
        object_ids = [obj.id for obj in objects]
        await use_fast_commit_for_current_transaction(db)
        await db.commit()
        return len(objects), object_ids, []
    except SQLAlchemyError as exc:
        await db.rollback()
        return await _commit_object_batch_row_by_row(db, batch, sheet_label, exc)


async def _commit_object_batch_row_by_row(
    db: AsyncSession,
    batch: list[tuple[ProjectObject, dict[str, Any]]],
    sheet_label: str,
    batch_error: SQLAlchemyError,
) -> tuple[int, list[UUID], list[dict[str, Any]]]:
    created = 0
    object_ids: list[UUID] = []
    errors: list[dict[str, Any]] = []
    for obj, row in batch:
        retry_obj = ProjectObject(
            project_id=obj.project_id,
            object_type=obj.object_type,
            sort_order=obj.sort_order,
            params=obj.params,
        )
        try:
            db.add(retry_obj)
            await db.flush()
            object_ids.append(retry_obj.id)
            await use_fast_commit_for_current_transaction(db)
            await db.commit()
            created += 1
        except Exception as exc:
            await db.rollback()
            message = f"{type(exc).__name__}: {exc}"
            if not message.strip():
                message = f"{type(batch_error).__name__}: {batch_error}"
            errors.append({"sheet": sheet_label, "row": row["_row"], "message": message})
    return created, object_ids, errors


async def _add_rows(
    db: AsyncSession,
    project_id: UUID,
    sheet_label: str,
    rows: list[dict[str, Any]],
    object_type: str,
    next_sort: int,
    current_count: int,
    dedupe_keys: set[str] | None = None,
) -> tuple[int, int, int, list[dict[str, Any]], list[UUID], int, int]:
    """Создаёт объекты из распарсенных строк.

    Расчёт теплопотерь здесь намеренно не запускается: импорт должен быстро
    сохранить всё распознанное и отдать новые объекты в один фоновый batch.
    """
    created = 0
    skipped_duplicates = 0
    skipped_limit = 0
    created_object_ids: list[UUID] = []
    errors: list[dict[str, Any]] = []
    batch: list[tuple[ProjectObject, dict[str, Any]]] = []
    builder = _build_pipe_params if object_type == "pipe" else _build_tank_params

    async def flush_batch() -> None:
        nonlocal batch, created, current_count
        attempted = len(batch)
        batch_created, object_ids, batch_errors = await _commit_object_batch(
            db,
            batch,
            sheet_label,
        )
        created += batch_created
        created_object_ids.extend(object_ids)
        errors.extend(batch_errors)
        current_count -= attempted - batch_created
        batch = []

    for row_index, row in enumerate(rows):
        if current_count >= settings.GUEST_MAX_OBJECTS_PER_PROJECT:
            skipped_limit = len(rows) - row_index
            errors.append(
                {
                    "sheet": sheet_label,
                    "row": row["_row"],
                    "message": (
                        "Достигнут лимит объектов в проекте "
                        f"({settings.GUEST_MAX_OBJECTS_PER_PROJECT}). "
                        f"Пропущено строк: {skipped_limit}."
                    ),
                }
            )
            break
        params, err = builder(row)
        if err or params is None:
            errors.append(
                {"sheet": sheet_label, "row": row["_row"], "message": err or "Ошибка парсинга"}
            )
            continue
        try:
            reject_legacy_specification_object_params(params)
            normalized_params = (
                prepare_project_object_params(object_type, params)
                if object_type == "pipe"
                else normalize_project_object_params(object_type, params)
            )
            if dedupe_keys is not None:
                key = _dedupe_key(object_type, normalized_params)
                if key in dedupe_keys:
                    skipped_duplicates += 1
                    continue
                dedupe_keys.add(key)
            obj = ProjectObject(
                project_id=project_id,
                object_type=object_type,
                sort_order=next_sort,
                params=normalized_params,
            )
            batch.append((obj, row))
            current_count += 1
            next_sort += 1
            if len(batch) >= IMPORT_COMMIT_BATCH_SIZE:
                await flush_batch()
        except Exception as exc:
            errors.append(
                {
                    "sheet": sheet_label,
                    "row": row["_row"],
                    "message": f"{type(exc).__name__}: {exc}",
                }
            )
    await flush_batch()
    return (
        created,
        next_sort,
        current_count,
        errors,
        created_object_ids,
        skipped_duplicates,
        skipped_limit,
    )


async def import_objects_from_csv(
    db: AsyncSession,
    project_id: UUID,
    principal: CurrentPrincipal,
    content: bytes,
    mode: str = "merge",
) -> dict[str, Any]:
    """Импортирует объекты из CSV-файла. Требуется колонка «Тип»."""
    import_mode = _validate_import_mode(mode)
    await _ensure_import_access(db, project_id, principal)

    sheets = _parse_csv(content)
    if not sheets:
        raise ExcelImportError(
            "В CSV не найдено ни одной строки с распознанным типом (труба/резервуар)."
        )

    if import_mode == "replace":
        await _replace_project_objects(db, project_id)
        current_count, next_sort = 0, 0
        dedupe_keys = None
    else:
        current_count, next_sort = await _project_import_state(db, project_id)
        dedupe_keys = (
            await _existing_dedupe_keys(db, project_id) if import_mode == "merge" else None
        )

    total_created = 0
    skipped_duplicates = 0
    skipped_limit = 0
    all_errors: list[dict[str, Any]] = []
    created_object_ids: list[UUID] = []
    for sheet_label, rows in sheets:
        obj_type = "pipe" if "Трубопровод" in sheet_label else "tank"
        (
            created,
            next_sort,
            current_count,
            errors,
            object_ids,
            skipped,
            limit_skipped,
        ) = await _add_rows(
            db,
            project_id,
            sheet_label,
            rows,
            obj_type,
            next_sort,
            current_count,
            dedupe_keys=dedupe_keys,
        )
        total_created += created
        skipped_duplicates += skipped
        skipped_limit += limit_skipped
        all_errors.extend(errors)
        created_object_ids.extend(object_ids)

    if import_mode == "replace" and not created_object_ids:
        await db.commit()
    if created_object_ids or import_mode == "replace":
        await _touch_project_updated_at(db, project_id)

    return {
        "created": total_created,
        "skipped_duplicates": skipped_duplicates,
        "skipped_limit": skipped_limit,
        "mode": import_mode,
        "errors": all_errors,
        "created_object_ids": created_object_ids,
    }


def _parse_excel_workbook(content: bytes) -> list[tuple[str, str, list[dict[str, Any]]]]:
    """Синхронный разбор xlsx: открыть книгу, проверить лимит листов и прочитать
    листы трубопроводов/резервуаров.

    openpyxl полностью блокирующий (parsing + распаковка XML), поэтому функция
    запускается в отдельном потоке через ``asyncio.to_thread`` и не должна
    держать event loop на время разбора файла (см. import_objects_from_excel).
    Возвращает список ``(sheet_label, object_type, rows)`` по найденным листам.
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportError(f"Не удалось открыть файл: {exc}") from exc
    if len(wb.sheetnames) > settings.MAX_IMPORT_SHEETS:
        raise ExcelImportError(f"Превышен лимит листов импорта: {settings.MAX_IMPORT_SHEETS}")

    parsed_sheets: list[tuple[str, str, list[dict[str, Any]]]] = []
    for sheet in wb.sheetnames:
        norm = _norm(sheet)
        if norm in PIPE_SHEET_NAMES:
            parsed_sheets.append((sheet, "pipe", _read_sheet(wb[sheet], PIPE_HEADERS)))
        elif norm in TANK_SHEET_NAMES:
            parsed_sheets.append((sheet, "tank", _read_sheet(wb[sheet], TANK_HEADERS)))
    return parsed_sheets


async def import_objects_from_excel(
    db: AsyncSession,
    project_id: UUID,
    principal: CurrentPrincipal,
    content: bytes,
    mode: str = "merge",
) -> dict[str, Any]:
    """Импортирует объекты из xlsx-файла в проект.

    Возвращает сводку: {"created": N, "errors": [{"sheet", "row", "message"}]}.
    """
    import_mode = _validate_import_mode(mode)
    _validate_xlsx_archive(content)

    # Блокирующий разбор openpyxl уводим в поток, чтобы не стопорить event loop
    # на время парсинга (как уже сделано для генерации отчётов).
    parsed_sheets = await asyncio.to_thread(_parse_excel_workbook, content)

    # Проверяем доступ к проекту
    try:
        await _ensure_import_access(db, project_id, principal)
    except (ProjectNotFoundError, ProjectAccessError):
        raise

    created = 0
    skipped_duplicates = 0
    skipped_limit = 0
    errors: list[dict[str, Any]] = []
    created_object_ids: list[UUID] = []

    if not parsed_sheets:
        raise ExcelImportError(
            "В файле не найдены листы «Трубопроводы» или «Резервуары». "
            "Используйте шаблон (кнопка «Скачать шаблон»)."
        )

    if import_mode == "replace":
        await _replace_project_objects(db, project_id)
        current_count, next_sort = 0, 0
        dedupe_keys = None
    else:
        current_count, next_sort = await _project_import_state(db, project_id)
        dedupe_keys = (
            await _existing_dedupe_keys(db, project_id) if import_mode == "merge" else None
        )

    for sheet, object_type, rows in parsed_sheets:
        (
            added,
            next_sort,
            current_count,
            sheet_errors,
            object_ids,
            skipped,
            limit_skipped,
        ) = await _add_rows(
            db,
            project_id,
            sheet,
            rows,
            object_type,
            next_sort,
            current_count,
            dedupe_keys=dedupe_keys,
        )
        created += added
        skipped_duplicates += skipped
        skipped_limit += limit_skipped
        errors.extend(sheet_errors)
        created_object_ids.extend(object_ids)

    if import_mode == "replace" and not created_object_ids:
        await db.commit()
    if created_object_ids or import_mode == "replace":
        await _touch_project_updated_at(db, project_id)

    return {
        "created": created,
        "skipped_duplicates": skipped_duplicates,
        "skipped_limit": skipped_limit,
        "mode": import_mode,
        "errors": errors,
        "created_object_ids": created_object_ids,
    }


def _material_label(material: Any) -> str:
    if not material:
        return ""
    material_str = str(material)
    for entry in list_insulation_materials():
        if entry.get("material") == material_str:
            return str(entry.get("name") or material_str)
    return material_str


SHAPE_LABELS_RU: dict[str, str] = {
    "cylindrical": "Цилиндр",
    "rectangular": "Параллелепипед",
}


def _to_export_mm(value: Any) -> float | str:
    if value in (None, ""):
        return ""
    try:
        result = round(float(value) * 1000, 3)
    except (TypeError, ValueError):
        return str(value)
    return result or ""


def _format_temperature_range(value: Any) -> str:
    parsed = _to_temperature_range(value)
    if parsed is None:
        return ""
    lower, upper = parsed
    return f"{lower:g}..{upper:g}"


def build_objects_xlsx(objects: list[Any]) -> bytes:
    """Экспорт объектов проекта в формат, round-trip-совместимый с импортом.

    Пишет листы `Трубопроводы` и `Резервуары` с теми же колонками, что и
    `build_template_xlsx` — файл можно сохранить локально и через
    `POST /objects/import-excel` загрузить обратно.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_pipe = wb.active
    ws_pipe.title = "Трубопроводы"
    pipe_cols = [
        "Наименование",
        "Диаметр, мм",
        "Длина, м",
        "Толщина стенки, мм",
        "Материал трубы",
        "λ трубы",
        "Толщина изоляции, мм",
        "Код материала изоляции",
        "λ 1-го слоя",
        "Диапазон температур 1-го слоя, °C",
        "Толщина 2-го слоя, мм",
        "Материал 2-го слоя",
        "λ 2-го слоя",
        "Диапазон температур 2-го слоя, °C",
        "Толщина 3-го слоя, мм",
        "Материал 3-го слоя",
        "λ 3-го слоя",
        "Диапазон температур 3-го слоя, °C",
        "T° среды",
        "Температура грунта",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        "Глубина заложения оси трубы, м",
        "Грунт",
        "λ грунта",
        "Скорость ветра, м/с",
        "α внешней теплоотдачи",
        "Режим температуры изоляции",
        "Материал покрытия",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Количество локальных элементов",
        "L экв., м",
    ]
    for c, h in enumerate(pipe_cols, start=1):
        cell = ws_pipe.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")

    ws_tank = wb.create_sheet("Резервуары")
    tank_cols = [
        "Наименование",
        "Форма",
        "Диаметр, мм",
        "Длина, мм",
        "Ширина, мм",
        "Высота, мм",
        "Толщина изоляции, мм",
        "Материал изоляции",
        "Код материала изоляции",
        "Статус материала изоляции",
        "Комментарий материала изоляции",
        "T° среды",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        "Режим температуры изоляции",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Q доп., Вт",
        "Толщина стенки, мм",
        "λ стенки",
        "Температура грунта",
        "Высота заглубленной части, м",
        "Грунт",
        "λ грунта",
        "Скорость ветра, м/с",
        "α внешней теплоотдачи",
        "λ 1-го слоя",
        "Диапазон температур 1-го слоя, °C",
        "Толщина 2-го слоя, мм",
        "Материал 2-го слоя",
        "λ 2-го слоя",
        "Диапазон температур 2-го слоя, °C",
        "Толщина 3-го слоя, мм",
        "Материал 3-го слоя",
        "λ 3-го слоя",
        "Диапазон температур 3-го слоя, °C",
    ]
    for c, h in enumerate(tank_cols, start=1):
        cell = ws_tank.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")

    for obj in objects:
        params = obj.params or {}
        name = params.get("name") or ""
        layers = params.get("insulation_layers")
        layer_list = layers if isinstance(layers, list) else []
        first_layer = layer_list[0] if layer_list and isinstance(layer_list[0], dict) else {}
        material_code = first_layer.get("material", "")
        material = _material_label(material_code)
        if obj.object_type == "pipe":
            layers = params.get("insulation_layers")
            layer_list = layers if isinstance(layers, list) else []

            def layer_value(index: int, key: str, _layers: list[Any] = layer_list) -> Any:
                if index >= len(_layers) or not isinstance(_layers[index], dict):
                    return ""
                return _layers[index].get(key, "")

            append_safe_row(
                ws_pipe,
                [
                    name,
                    _to_export_mm(params.get("outer_diameter")),
                    params.get("pipe_length") or "",
                    _to_export_mm(params.get("wall_thickness")),
                    params.get("pipe_material", ""),
                    params.get("pipe_lambda", ""),
                    _to_export_mm(layer_value(0, "thickness")),
                    layer_value(0, "material"),
                    layer_value(0, "conductivity"),
                    _format_temperature_range(layer_value(0, "temperature_range")),
                    _to_export_mm(layer_value(1, "thickness")),
                    layer_value(1, "material"),
                    layer_value(1, "conductivity"),
                    _format_temperature_range(layer_value(1, "temperature_range")),
                    _to_export_mm(layer_value(2, "thickness")),
                    layer_value(2, "material"),
                    layer_value(2, "conductivity"),
                    _format_temperature_range(layer_value(2, "temperature_range")),
                    params.get("ambient_temperature", ""),
                    params.get("ground_temperature", ""),
                    params.get("process_temperature", ""),
                    params.get("vapor_temperature", ""),
                    params.get("placement", ""),
                    params.get("pipe_centerline_depth", ""),
                    params.get("ground_type", ""),
                    params.get("ground_conductivity", ""),
                    params.get("wind_speed", ""),
                    params.get("alpha_vnesh", ""),
                    params.get("insulation_temperature_basis", ""),
                    params.get("insulation_cover_material", ""),
                    params.get("climate_region", ""),
                    params.get("climate_city", ""),
                    params.get("climate_key", ""),
                    params.get("climate_temperature_basis", ""),
                    params.get("safety_factor", ""),
                    params.get("min_switch_temperature", ""),
                    params.get("num_local_elements", ""),
                    params.get("local_element_equiv_length", ""),
                ],
            )
        elif obj.object_type == "tank":
            shape_code = params.get("shape") or "cylindrical"
            shape = SHAPE_LABELS_RU.get(shape_code, shape_code)

            def to_mm(k, _params=params):
                return _to_export_mm(_params.get(k))

            def tank_layer_value(index: int, key: str, _layers: list[Any] = layer_list) -> Any:
                if index >= len(_layers) or not isinstance(_layers[index], dict):
                    return ""
                return _layers[index].get(key, "")

            append_safe_row(
                ws_tank,
                [
                    name,
                    shape,
                    to_mm("diameter"),
                    to_mm("length"),
                    to_mm("width"),
                    to_mm("height"),
                    _to_export_mm(first_layer.get("thickness")),
                    material,
                    material_code,
                    "Конкретный материал" if material_code else "",
                    "",
                    params.get("ambient_temperature", ""),
                    params.get("process_temperature", ""),
                    params.get("vapor_temperature", ""),
                    params.get("placement", ""),
                    params.get("insulation_temperature_basis", ""),
                    params.get("climate_region", ""),
                    params.get("climate_city", ""),
                    params.get("climate_key", ""),
                    params.get("climate_temperature_basis", ""),
                    params.get("safety_factor", ""),
                    params.get("min_switch_temperature", ""),
                    params.get("q_additional", ""),
                    to_mm("wall_thickness"),
                    params.get("wall_lambda", ""),
                    params.get("ground_temperature", ""),
                    params.get("tank_buried_height", ""),
                    params.get("ground_type", ""),
                    params.get("ground_conductivity", ""),
                    params.get("wind_speed", ""),
                    params.get("alpha_vnesh", ""),
                    tank_layer_value(0, "conductivity"),
                    _format_temperature_range(tank_layer_value(0, "temperature_range")),
                    _to_export_mm(tank_layer_value(1, "thickness")),
                    tank_layer_value(1, "material"),
                    tank_layer_value(1, "conductivity"),
                    _format_temperature_range(tank_layer_value(1, "temperature_range")),
                    _to_export_mm(tank_layer_value(2, "thickness")),
                    tank_layer_value(2, "material"),
                    tank_layer_value(2, "conductivity"),
                    _format_temperature_range(tank_layer_value(2, "temperature_range")),
                ],
            )

    ws_pipe.column_dimensions["A"].width = 24
    for col_idx in range(2, len(pipe_cols) + 1):
        ws_pipe.column_dimensions[get_column_letter(col_idx)].width = 18
    ws_tank.column_dimensions["A"].width = 24
    for col_idx in range(2, len(tank_cols) + 1):
        ws_tank.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_xlsx() -> bytes:
    """Формирует xlsx-шаблон с двумя листами и примерами."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_pipe = wb.active
    ws_pipe.title = "Трубопроводы"
    pipe_cols = [
        "Наименование",
        "Диаметр, мм",
        "Длина, м",
        "Толщина стенки, мм",
        "Материал трубы",
        "λ трубы",
        "Толщина изоляции, мм",
        "Код материала изоляции",
        "λ 1-го слоя",
        "Диапазон температур 1-го слоя, °C",
        "Толщина 2-го слоя, мм",
        "Материал 2-го слоя",
        "λ 2-го слоя",
        "Диапазон температур 2-го слоя, °C",
        "Толщина 3-го слоя, мм",
        "Материал 3-го слоя",
        "λ 3-го слоя",
        "Диапазон температур 3-го слоя, °C",
        "T° среды",
        "Температура грунта",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        "Глубина заложения оси трубы, м",
        "Грунт",
        "λ грунта",
        "Скорость ветра, м/с",
        "α внешней теплоотдачи",
        "Режим температуры изоляции",
        "Материал покрытия",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Количество локальных элементов",
        "L экв., м",
    ]
    for c, h in enumerate(pipe_cols, start=1):
        cell = ws_pipe.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")
    for example in (
        {
            "Наименование": "Пример DN100",
            "Диаметр, мм": 108,
            "Длина, м": 50,
            "Толщина стенки, мм": 4,
            "Материал трубы": "carbon_steel",
            "Толщина изоляции, мм": 50,
            "Код материала изоляции": "mineral_wool_boards_120",
            "T° среды": -20,
            "T° продукта": 80,
            "Размещение": "outdoor",
            "Скорость ветра, м/с": 0,
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Количество локальных элементов": 6,
            "L экв., м": 1.5,
        },
        {
            "Наименование": "Пример DN50",
            "Диаметр, мм": 57,
            "Длина, м": 20,
            "Толщина стенки, мм": 4,
            "Материал трубы": "carbon_steel",
            "Толщина изоляции, мм": 40,
            "Код материала изоляции": "polyurethane_products_50",
            "T° среды": -30,
            "T° продукта": 60,
            "Размещение": "outdoor",
            "Скорость ветра, м/с": 0,
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Количество локальных элементов": 0,
        },
    ):
        ws_pipe.append([example.get(header, "") for header in pipe_cols])
    ws_pipe.column_dimensions["A"].width = 24
    for col_idx in range(2, len(pipe_cols) + 1):
        ws_pipe.column_dimensions[get_column_letter(col_idx)].width = 18

    ws_tank = wb.create_sheet("Резервуары")
    tank_cols = [
        "Наименование",
        "Форма",
        "Диаметр, мм",
        "Длина, мм",
        "Ширина, мм",
        "Высота, мм",
        "Толщина изоляции, мм",
        "Материал изоляции",
        "Код материала изоляции",
        "Статус материала изоляции",
        "Комментарий материала изоляции",
        "T° среды",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        # Без скорости ветра размещение outdoor не рассчитывается, поэтому
        # колонка обязана быть и в шаблоне резервуаров — как у труб.
        "Скорость ветра, м/с",
        "Режим температуры изоляции",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Q доп., Вт",
    ]
    for c, h in enumerate(tank_cols, start=1):
        cell = ws_tank.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")
    for tank_example in (
        {
            "Наименование": "Бак цил.",
            "Форма": "Цилиндр",
            "Диаметр, мм": 2000,
            "Высота, мм": 3000,
            "Толщина изоляции, мм": 80,
            "Материал изоляции": "Плиты минераловатные прошивные",
            "Код материала изоляции": "mineral_wool_boards_120",
            "T° среды": -20,
            "T° продукта": 80,
            "Размещение": "outdoor",
            "Скорость ветра, м/с": 0,
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Q доп., Вт": 0,
        },
        {
            "Наименование": "Бак прям.",
            "Форма": "Параллелепипед",
            "Длина, мм": 5000,
            "Ширина, мм": 3000,
            "Высота, мм": 4000,
            "Толщина изоляции, мм": 80,
            "Материал изоляции": "Плиты минераловатные прошивные",
            "Код материала изоляции": "mineral_wool_boards_120",
            "T° среды": -20,
            "T° продукта": 80,
            "Размещение": "outdoor",
            "Скорость ветра, м/с": 0,
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Q доп., Вт": 0,
        },
        {
            "Наименование": "Шаровой",
            "Форма": "Шар",
            "Диаметр, мм": 1500,
            "Толщина изоляции, мм": 60,
            "Материал изоляции": "Изделия из ППУ",
            "Код материала изоляции": "polyurethane_products_50",
            "T° среды": -20,
            "T° продукта": 60,
            "Размещение": "outdoor",
            "Скорость ветра, м/с": 0,
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Q доп., Вт": 0,
        },
    ):
        ws_tank.append([tank_example.get(header, "") for header in tank_cols])
    ws_tank.column_dimensions["A"].width = 24
    for col_idx in range(2, len(tank_cols) + 1):
        ws_tank.column_dimensions[get_column_letter(col_idx)].width = 18

    ws_info = wb.create_sheet("Справка")
    ws_info["A1"] = "Подсказки по импорту"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A3"] = "Материалы изоляции: используйте конкретный код и плотность."
    ws_info["A3"].font = Font(bold=True)
    for i, mat in enumerate(
        [
            "mineral_wool_boards_120 — Плиты минераловатные прошивные, ρ 120",
            "polyurethane_products_50 — Изделия из ППУ, ρ 50",
            "polystyrene_products_50 — Изделия из полистирола, ρ 50",
            "mineral_wool_cylinders_100 — Цилиндры минераловатные, ρ 100",
            "fiberglass_mats_70 — Маты стеклянного штапельного волокна, ρ 70",
            "k_flex_st — Кайманфлекс K-Flex ST, ρ 60-80",
        ],
        start=4,
    ):
        ws_info.cell(row=i, column=1, value=mat)
    ws_info["A11"] = "Формы резервуара:"
    ws_info["A11"].font = Font(bold=True)
    ws_info["A12"] = "Цилиндр — требуются Диаметр и Высота"
    ws_info["A13"] = "Параллелепипед — требуются Длина, Ширина, Высота"
    ws_info["A14"] = "Шар — требуется Диаметр"
    ws_info.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_csv() -> bytes:
    """Формирует CSV-шаблон: один файл, колонка «Тип» разделяет трубы и резервуары."""
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", lineterminator="\n")
    headers = [
        "Тип",
        "Наименование",
        "Форма",
        "Диаметр, мм",
        "Длина, мм",
        "Ширина, мм",
        "Высота, мм",
        "Длина, м",  # для трубы — длина в метрах
        "Толщина изоляции, мм",
        "Материал изоляции",
        "Код материала изоляции",
        "Статус материала изоляции",
        "Комментарий материала изоляции",
        "T° среды",
        "T° продукта",
        "T проп., °C",
        "Размещение",
        "Режим температуры изоляции",
        "Климатический регион",
        "Климатический город",
        "Ключ климата",
        "Обеспеченность климата",
        "Kзап",
        "Мин. T включения, °C",
        "Толщина стенки, мм",
        "Материал трубы",
        "Скорость ветра, м/с",
        "Количество локальных элементов",
        "L экв., м",
        "Q доп., Вт",
    ]
    writer.writerow(headers)
    # Примеры задаём словарями: позиционные строки молча теряли значения при
    # правке состава колонок (так у резервуаров пропала скорость ветра).
    # Для трубы: диаметр в мм, длина в м, форма/ширина/высота пустые.
    for example in (
        {
            "Тип": "труба",
            "Наименование": "Пример DN100",
            "Диаметр, мм": 108,
            "Длина, м": 50,
            "Толщина изоляции, мм": 50,
            "Материал изоляции": "Плиты минераловатные прошивные",
            "Код материала изоляции": "mineral_wool_boards_120",
            "T° среды": -20,
            "T° продукта": 80,
            "Размещение": "outdoor",
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Толщина стенки, мм": 4,
            "Материал трубы": "carbon_steel",
            "Скорость ветра, м/с": 0,
            "Количество локальных элементов": 6,
            "L экв., м": 1.5,
        },
        {
            "Тип": "труба",
            "Наименование": "Пример DN50",
            "Диаметр, мм": 57,
            "Длина, м": 20,
            "Толщина изоляции, мм": 40,
            "Материал изоляции": "Теплоизоляционные изделия из пенополиуретана",
            "Код материала изоляции": "polyurethane_products_50",
            "T° среды": -30,
            "T° продукта": 60,
            "Размещение": "outdoor",
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Толщина стенки, мм": 4,
            "Материал трубы": "carbon_steel",
            "Скорость ветра, м/с": 0,
            "Количество локальных элементов": 0,
        },
        # Резервуары: заполняем Форма + нужные габариты в мм.
        {
            "Тип": "резервуар",
            "Наименование": "Бак цил.",
            "Форма": "Цилиндр",
            "Диаметр, мм": 2000,
            "Высота, мм": 3000,
            "Толщина изоляции, мм": 80,
            "Материал изоляции": "Плиты минераловатные прошивные",
            "Код материала изоляции": "mineral_wool_boards_120",
            "T° среды": -20,
            "T° продукта": 80,
            "Размещение": "outdoor",
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Скорость ветра, м/с": 0,
            "Q доп., Вт": 0,
        },
        {
            "Тип": "резервуар",
            "Наименование": "Бак прям.",
            "Форма": "Параллелепипед",
            "Длина, мм": 5000,
            "Ширина, мм": 3000,
            "Высота, мм": 4000,
            "Толщина изоляции, мм": 80,
            "Материал изоляции": "Теплоизоляционные изделия из пенополиуретана",
            "Код материала изоляции": "polyurethane_products_50",
            "T° среды": -20,
            "T° продукта": 60,
            "Размещение": "outdoor",
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Скорость ветра, м/с": 0,
            "Q доп., Вт": 0,
        },
        {
            "Тип": "резервуар",
            "Наименование": "Шар",
            "Форма": "Шар",
            "Диаметр, мм": 1500,
            "Толщина изоляции, мм": 60,
            "Материал изоляции": "Теплоизоляционные изделия из пенополиуретана",
            "Код материала изоляции": "polyurethane_products_50",
            "T° среды": -20,
            "T° продукта": 50,
            "Размещение": "outdoor",
            "Режим температуры изоляции": "outdoor_winter",
            "Kзап": 1.1,
            "Мин. T включения, °C": -20,
            "Скорость ветра, м/с": 0,
            "Q доп., Вт": 0,
        },
    ):
        writer.writerow([example.get(header, "") for header in headers])
    # UTF-8 BOM чтобы Excel правильно открывал файл
    return ("\ufeff" + buf.getvalue()).encode("utf-8")
