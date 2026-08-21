"""CSV and XLSX parsing boundaries for object imports."""

from __future__ import annotations

import io
import zipfile
from typing import Any

from openpyxl import load_workbook

from app.core.config import settings
from app.services.object_spreadsheet.mapping import TYPE_ALIASES, _norm
from app.services.spreadsheet_schema import PIPE_HEADERS, TANK_HEADERS, TYPE_HEADERS

PIPE_SHEET_NAMES = {"трубопроводы", "трубы", "pipes"}
TANK_SHEET_NAMES = {"резервуары", "ёмкости", "емкости", "tanks"}


class ExcelImportError(Exception):
    """Ошибка импорта Excel/CSV."""


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > settings.MAX_XLSX_FILES:
                raise ExcelImportError(
                    f"XLSX содержит слишком много внутренних файлов: {len(infos)}"
                )
            total_uncompressed = sum(info.file_size for info in infos)
            if total_uncompressed > settings.MAX_XLSX_UNCOMPRESSED_BYTES:
                raise ExcelImportError(
                    "XLSX слишком большой после распаковки " f"({total_uncompressed // 1024} КБ)"
                )
            for info in infos:
                if info.compress_size and info.file_size / info.compress_size > 100:
                    raise ExcelImportError(
                        "XLSX похож на zip-bomb: слишком высокий коэффициент сжатия"
                    )
    except zipfile.BadZipFile as exc:
        raise ExcelImportError("Файл не является корректным XLSX-архивом") from exc


def _read_sheet(ws: Any, header_map: dict[str, str]) -> list[dict[str, Any]]:
    """Читает лист, мапит заголовки → канонические имена, возвращает список строк-словарей.

    Строки возвращаются с ключом ``_row`` — номер строки в Excel (1-based).
    """
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    mapped_cols: list[tuple[int, str]] = []
    for idx, h in enumerate(header_row):
        key = header_map.get(_norm(h))
        if key:
            mapped_cols.append((idx, key))
    result: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        if len(result) >= settings.MAX_IMPORT_ROWS:
            raise ExcelImportError(f"Превышен лимит строк импорта: {settings.MAX_IMPORT_ROWS}")
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        item: dict[str, Any] = {"_row": row_idx}
        for idx, key in mapped_cols:
            if idx < len(row):
                item[key] = row[idx]
        result.append(item)
    return result


def _parse_csv(content: bytes) -> list[tuple[str, list[dict[str, Any]]]]:
    """Парсит CSV-файл.

    Возвращает список пар (sheet_label, rows) по типу:
    [('Трубопроводы', [...]), ('Резервуары', [...])].
    CSV должен содержать колонку «Тип» со значениями «труба» / «резервуар».
    Автодетект разделителя (``,``, ``;``, ``\t``). Кодировки: UTF-8 / UTF-8-BOM / CP1251.
    """
    import csv

    # Определяем кодировку
    text: str | None = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ExcelImportError("Не удалось определить кодировку CSV (ожидается UTF-8 или CP1251)")

    # Автодетект разделителя по первой строке
    first_line = text.splitlines()[0] if text.strip() else ""
    if not first_line:
        raise ExcelImportError("CSV-файл пустой")
    counts = {d: first_line.count(d) for d in (";", ",", "\t")}
    delimiter = max(counts, key=lambda item: counts[item]) if any(counts.values()) else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        raise ExcelImportError("CSV-файл пустой")
    if len(all_rows) - 1 > settings.MAX_IMPORT_ROWS:
        raise ExcelImportError(f"Превышен лимит строк импорта: {settings.MAX_IMPORT_ROWS}")

    header = all_rows[0]
    # Индекс колонки «Тип»
    type_idx = next((i for i, h in enumerate(header) if _norm(h) in TYPE_HEADERS), None)
    if type_idx is None:
        raise ExcelImportError(
            "В CSV не найдена колонка «Тип» (со значениями «труба»/«резервуар»). "
            "Используйте шаблон CSV или файл Excel."
        )

    # Разделяем строки по типу
    pipe_rows_raw: list[tuple[int, list[str]]] = []
    tank_rows_raw: list[tuple[int, list[str]]] = []
    for row_idx, row in enumerate(all_rows[1:], start=2):
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        if type_idx >= len(row):
            continue
        t = TYPE_ALIASES.get(_norm(row[type_idx]))
        if t == "pipe":
            pipe_rows_raw.append((row_idx, row))
        elif t == "tank":
            tank_rows_raw.append((row_idx, row))

    def build_mapped(
        rows_raw: list[tuple[int, list[str]]], header_map: dict[str, str]
    ) -> list[dict[str, Any]]:
        # Определяем маппинг колонок общего header
        mapped_cols: list[tuple[int, str]] = []
        for idx, h in enumerate(header):
            key = header_map.get(_norm(h))
            if key:
                mapped_cols.append((idx, key))
        out: list[dict[str, Any]] = []
        for row_idx, row in rows_raw:
            item: dict[str, Any] = {"_row": row_idx}
            for idx, key in mapped_cols:
                if idx < len(row):
                    val = row[idx]
                    # Пустые строки — None
                    normalized_value = val if (val is not None and str(val).strip() != "") else None
                    # «Материал изоляции» and its code are both aliases of
                    # one semantic field. Keep a meaningful earlier alias
                    # when a later optional display/code column is blank.
                    if normalized_value is not None or item.get(key) in (None, ""):
                        item[key] = normalized_value
            out.append(item)
        return out

    result: list[tuple[str, list[dict[str, Any]]]] = []
    if pipe_rows_raw:
        result.append(("Трубопроводы (CSV)", build_mapped(pipe_rows_raw, PIPE_HEADERS)))
    if tank_rows_raw:
        result.append(("Резервуары (CSV)", build_mapped(tank_rows_raw, TANK_HEADERS)))
    return result


def _parse_excel_workbook(content: bytes) -> list[tuple[str, str, list[dict[str, Any]]]]:
    """Синхронный разбор xlsx: открыть книгу, проверить лимит листов и прочитать
    листы трубопроводов/резервуаров.

    openpyxl полностью блокирующий (parsing + распаковка XML), поэтому функция
    запускается в отдельном потоке через ``asyncio.to_thread`` и не должна
    держать event loop на время разбора файла (см. import_objects_from_excel).
    Возвращает список ``(sheet_label, object_type, rows)`` по найденным листам.
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportError(f"Не удалось открыть файл: {exc}") from exc
    if len(wb.sheetnames) > settings.MAX_IMPORT_SHEETS:
        raise ExcelImportError(f"Превышен лимит листов импорта: {settings.MAX_IMPORT_SHEETS}")

    parsed_sheets: list[tuple[str, str, list[dict[str, Any]]]] = []
    for sheet in wb.sheetnames:
        norm = _norm(sheet)
        if norm in PIPE_SHEET_NAMES:
            parsed_sheets.append((sheet, "pipe", _read_sheet(wb[sheet], PIPE_HEADERS)))
        elif norm in TANK_SHEET_NAMES:
            parsed_sheets.append((sheet, "tank", _read_sheet(wb[sheet], TANK_HEADERS)))
    return parsed_sheets
