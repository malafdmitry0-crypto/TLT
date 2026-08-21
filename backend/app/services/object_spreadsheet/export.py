"""Object spreadsheet exports and downloadable import templates."""

from __future__ import annotations

import io
import re
from typing import Any

from app.reference_data.loader import list_insulation_materials
from app.services.spreadsheet_safety import append_safe_row
from app.services.spreadsheet_schema import (
    PIPE_XLSX_HEADERS,
    TANK_XLSX_HEADERS,
)


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, int | float):
        return float(value)
    normalized = str(value).replace(",", ".").strip()
    match = re.search(r"[-+]?\\d+(?:\\.\\d+)?", normalized)
    if match:
        normalized = match.group(0)
    try:
        return float(normalized)
    except ValueError:
        return None


def _to_temperature_range(value: Any) -> tuple[float, float] | None:
    if isinstance(value, list | tuple) and len(value) == 2:
        lower = _to_float(value[0])
        upper = _to_float(value[1])
    elif value not in (None, ""):
        values = re.findall(r"[-+]?\\d+(?:[.,]\\d+)?", str(value))
        if len(values) != 2:
            return None
        lower, upper = (_to_float(item) for item in values)
    else:
        return None
    if lower is None or upper is None:
        return None
    return lower, upper


def _material_label(material: Any) -> str:
    if not material:
        return ""
    material_str = str(material)
    for entry in list_insulation_materials():
        if entry.get("material") == material_str:
            return str(entry.get("name") or material_str)
    return material_str


SHAPE_LABELS_RU: dict[str, str] = {
    "cylindrical": "Цилиндр",
    "rectangular": "Параллелепипед",
}


def _to_export_mm(value: Any) -> float | str:
    if value in (None, ""):
        return ""
    try:
        result = round(float(value) * 1000, 3)
    except (TypeError, ValueError):
        return str(value)
    return result or ""


def _format_temperature_range(value: Any) -> str:
    parsed = _to_temperature_range(value)
    if parsed is None:
        return ""
    lower, upper = parsed
    return f"{lower:g}..{upper:g}"


def build_objects_xlsx(objects: list[Any]) -> bytes:
    """Экспорт объектов проекта в формат, round-trip-совместимый с импортом.

    Пишет листы `Трубопроводы` и `Резервуары` с теми же колонками, что и
    `build_template_xlsx` — файл можно сохранить локально и через
    `POST /objects/import-excel` загрузить обратно.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_pipe = wb.active
    ws_pipe.title = "Трубопроводы"
    pipe_cols = list(PIPE_XLSX_HEADERS)
    for c, h in enumerate(pipe_cols, start=1):
        cell = ws_pipe.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")

    ws_tank = wb.create_sheet("Резервуары")
    tank_cols = list(TANK_XLSX_HEADERS)
    for c, h in enumerate(tank_cols, start=1):
        cell = ws_tank.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")

    for obj in objects:
        params = obj.params or {}
        name = params.get("name") or ""
        layers = params.get("insulation_layers")
        layer_list = layers if isinstance(layers, list) else []
        first_layer = layer_list[0] if layer_list and isinstance(layer_list[0], dict) else {}
        material_code = first_layer.get("material", "")
        material = _material_label(material_code)
        if obj.object_type == "pipe":
            has_ambient_bounds = params.get("placement") != "underground"
            layers = params.get("insulation_layers")
            layer_list = layers if isinstance(layers, list) else []

            def layer_value(index: int, key: str, _layers: list[Any] = layer_list) -> Any:
                if index >= len(_layers) or not isinstance(_layers[index], dict):
                    return ""
                return _layers[index].get(key, "")

            append_safe_row(
                ws_pipe,
                [
                    name,
                    _to_export_mm(params.get("outer_diameter")),
                    params.get("pipe_length") or "",
                    _to_export_mm(params.get("wall_thickness")),
                    params.get("pipe_material", ""),
                    params.get("pipe_lambda", ""),
                    _to_export_mm(layer_value(0, "thickness")),
                    layer_value(0, "material"),
                    layer_value(0, "conductivity"),
                    _format_temperature_range(layer_value(0, "temperature_range")),
                    _to_export_mm(layer_value(1, "thickness")),
                    layer_value(1, "material"),
                    layer_value(1, "conductivity"),
                    _format_temperature_range(layer_value(1, "temperature_range")),
                    _to_export_mm(layer_value(2, "thickness")),
                    layer_value(2, "material"),
                    layer_value(2, "conductivity"),
                    _format_temperature_range(layer_value(2, "temperature_range")),
                    params.get("ambient_temperature", "") if has_ambient_bounds else "",
                    params.get("max_ambient_temperature", "") if has_ambient_bounds else "",
                    params.get("ambient_temperature_source", "") if has_ambient_bounds else "",
                    params.get("ground_temperature", ""),
                    params.get("process_temperature", ""),
                    params.get("vapor_temperature", ""),
                    params.get("placement", ""),
                    params.get("pipe_centerline_depth", ""),
                    params.get("ground_type", ""),
                    params.get("ground_conductivity", ""),
                    params.get("wind_speed", ""),
                    params.get("wind_speed_source", ""),
                    params.get("insulation_temperature_basis", ""),
                    params.get("insulation_cover_material", ""),
                    params.get("climate_region", ""),
                    params.get("climate_city", ""),
                    params.get("climate_key", ""),
                    params.get("climate_temperature_basis", ""),
                    params.get("safety_factor", ""),
                    params.get("safety_factor_source", ""),
                    params.get("min_switch_temperature", ""),
                    params.get("num_local_elements", ""),
                    params.get("local_element_equiv_length", ""),
                ],
            )
        elif obj.object_type == "tank":
            shape_code = params.get("shape") or "cylindrical"
            shape = SHAPE_LABELS_RU.get(shape_code, shape_code)

            def to_mm(k: str, _params: dict[str, Any] = params) -> Any:
                return _to_export_mm(_params.get(k))

            def tank_layer_value(index: int, key: str, _layers: list[Any] = layer_list) -> Any:
                if index >= len(_layers) or not isinstance(_layers[index], dict):
                    return ""
                return _layers[index].get(key, "")

            append_safe_row(
                ws_tank,
                [
                    name,
                    shape,
                    to_mm("diameter"),
                    to_mm("length"),
                    to_mm("width"),
                    to_mm("height"),
                    _to_export_mm(first_layer.get("thickness")),
                    material,
                    material_code,
                    "Конкретный материал" if material_code else "",
                    "",
                    params.get("ambient_temperature", ""),
                    params.get("max_ambient_temperature", ""),
                    params.get("ambient_temperature_source", ""),
                    params.get("process_temperature", ""),
                    params.get("vapor_temperature", ""),
                    params.get("placement", ""),
                    params.get("insulation_temperature_basis", ""),
                    params.get("climate_region", ""),
                    params.get("climate_city", ""),
                    params.get("climate_key", ""),
                    params.get("climate_temperature_basis", ""),
                    params.get("safety_factor", ""),
                    params.get("safety_factor_source", ""),
                    params.get("min_switch_temperature", ""),
                    params.get("heating_height", ""),
                    params.get("laying_step", ""),
                    params.get("q_additional", ""),
                    to_mm("wall_thickness"),
                    params.get("wall_lambda", ""),
                    params.get("ground_temperature", ""),
                    params.get("tank_buried_height", ""),
                    params.get("ground_type", ""),
                    params.get("ground_conductivity", ""),
                    params.get("wind_speed", ""),
                    params.get("wind_speed_source", ""),
                    tank_layer_value(0, "conductivity"),
                    _format_temperature_range(tank_layer_value(0, "temperature_range")),
                    _to_export_mm(tank_layer_value(1, "thickness")),
                    tank_layer_value(1, "material"),
                    tank_layer_value(1, "conductivity"),
                    _format_temperature_range(tank_layer_value(1, "temperature_range")),
                    _to_export_mm(tank_layer_value(2, "thickness")),
                    tank_layer_value(2, "material"),
                    tank_layer_value(2, "conductivity"),
                    _format_temperature_range(tank_layer_value(2, "temperature_range")),
                ],
            )

    ws_pipe.column_dimensions["A"].width = 24
    for col_idx in range(2, len(pipe_cols) + 1):
        ws_pipe.column_dimensions[get_column_letter(col_idx)].width = 18
    ws_tank.column_dimensions["A"].width = 24
    for col_idx in range(2, len(tank_cols) + 1):
        ws_tank.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
