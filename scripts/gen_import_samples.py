"""Генератор тестовых sample-файлов с ~100 записями (смесь труб и резервуаров)."""
import csv
import io
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

random.seed(42)

MATERIALS = ["Минеральная вата", "Пеностекло", "Пенополиуретан", "Пенополистирол", "Аэрогель", "Силикат кальция"]
# Стандартные DN для труб (наружный диаметр, мм)
DN_DIAMETERS = [21, 27, 34, 42, 48, 57, 76, 89, 108, 114, 133, 159, 219, 273, 325, 426]
SHAPES = ["Цилиндр", "Параллелепипед", "Шар"]


def gen_pipes(n: int) -> list[list]:
    rows = []
    for i in range(1, n + 1):
        d = random.choice(DN_DIAMETERS)
        L = round(random.uniform(5, 500), 1)
        ins = random.choice([30, 40, 50, 60, 80, 100])
        mat = random.choice(MATERIALS)
        t_amb = random.choice([-40, -30, -25, -20, -15, -10, 0, 5, 10])
        t_proc = random.choice([40, 50, 60, 70, 80, 90, 100, 110, 120])
        name = f"Труба DN{d} №{i}"
        rows.append([name, d, L, ins, mat, t_amb, t_proc])
    return rows


def gen_tanks(n: int) -> list[list]:
    rows = []
    for i in range(1, n + 1):
        shape = random.choice(SHAPES)
        ins = random.choice([40, 50, 60, 80, 100, 120, 150])
        mat = random.choice(MATERIALS)
        t_amb = random.choice([-40, -30, -20, -10, 0, 10])
        t_proc = random.choice([40, 60, 80, 100])
        if shape == "Цилиндр":
            d = random.choice([1000, 1500, 2000, 2500, 3000])
            h = random.choice([2000, 3000, 4000, 5000, 6000])
            rows.append([f"Цил. бак №{i}", shape, d, "", "", h, ins, mat, t_amb, t_proc])
        elif shape == "Параллелепипед":
            L = random.choice([2000, 3000, 4000, 5000, 6000])
            W = random.choice([1500, 2000, 3000, 4000])
            h = random.choice([2000, 3000, 4000])
            rows.append([f"Прям. бак №{i}", shape, "", L, W, h, ins, mat, t_amb, t_proc])
        else:  # Шар
            d = random.choice([1000, 1500, 2000, 2500])
            rows.append([f"Шар. бак №{i}", shape, d, "", "", "", ins, mat, t_amb, t_proc])
    return rows


def build_xlsx(pipes, tanks) -> bytes:
    wb = Workbook()
    ws_pipe = wb.active
    ws_pipe.title = "Трубопроводы"
    pipe_cols = ["Наименование", "Диаметр, мм", "Длина, м", "Толщина изоляции, мм",
                 "Материал изоляции", "T° среды", "T° продукта"]
    for c, h in enumerate(pipe_cols, 1):
        cell = ws_pipe.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")
    for row in pipes:
        ws_pipe.append(row)
    ws_pipe.column_dimensions["A"].width = 26
    for col in "BCDEFG":
        ws_pipe.column_dimensions[col].width = 18

    ws_tank = wb.create_sheet("Резервуары")
    tank_cols = ["Наименование", "Форма", "Диаметр, мм", "Длина, мм", "Ширина, мм",
                 "Высота, мм", "Толщина изоляции, мм", "Материал изоляции",
                 "T° среды", "T° продукта"]
    for c, h in enumerate(tank_cols, 1):
        cell = ws_tank.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEEF7")
    for row in tanks:
        ws_tank.append(row)
    ws_tank.column_dimensions["A"].width = 26
    for col in "BCDEFGHIJ":
        ws_tank.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(pipes, tanks) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\n")
    w.writerow(["Тип", "Наименование", "Форма", "Диаметр, мм", "Длина, мм",
                "Ширина, мм", "Высота, мм", "Длина, м", "Толщина изоляции, мм",
                "Материал изоляции", "T° среды", "T° продукта"])
    # Перемешиваем в один общий поток, чтобы трубы и резервуары чередовались
    all_rows = []
    for p in pipes:
        # name, d_mm, L_m, ins_mm, mat, t_a, t_p
        all_rows.append(["труба", p[0], "", p[1], "", "", "", p[2], p[3], p[4], p[5], p[6]])
    for t in tanks:
        # name, shape, d_mm, L_mm, W_mm, H_mm, ins_mm, mat, t_a, t_p
        all_rows.append(["резервуар", t[0], t[1], t[2], t[3], t[4], t[5], "", t[6], t[7], t[8], t[9]])
    random.shuffle(all_rows)
    for r in all_rows:
        w.writerow(r)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


# 60 труб + 40 резервуаров = 100 записей
pipes = gen_pipes(60)
tanks = gen_tanks(40)

with open("/tmp/sample_import.xlsx", "wb") as f:
    f.write(build_xlsx(pipes, tanks))
with open("/tmp/sample_import.csv", "wb") as f:
    f.write(build_csv(pipes, tanks))

print(f"Сгенерировано: {len(pipes)} труб + {len(tanks)} резервуаров = {len(pipes)+len(tanks)} записей")
